from flask import Flask, render_template, request, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_mail import Mail, Message
from sqlalchemy import func, text
from sqlalchemy.exc import ProgrammingError
from datetime import datetime
from werkzeug.security import check_password_hash, generate_password_hash
import os
import base64
import io
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.drawing.image import Image
    OPENPYXL_AVAILABLE = True
    print("✅ openpyxl está disponible")
except ImportError as e:
    OPENPYXL_AVAILABLE = False
    print(f"⚠️ openpyxl no está instalado. Error: {e}")
    print("⚠️ Asegúrate de activar el entorno virtual (env) y ejecutar: pip install openpyxl")
    print(f"⚠️ Python actual: {__import__('sys').executable}")

app = Flask(__name__)

# Configurar CORS manualmente si flask_cors no está instalado
@app.after_request
def after_request(response):
    # Solo agregar headers si no existen ya (para evitar duplicación)
    if 'Access-Control-Allow-Origin' not in response.headers:
        response.headers['Access-Control-Allow-Origin'] = '*'
    if 'Access-Control-Allow-Headers' not in response.headers:
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type,Authorization'
    if 'Access-Control-Allow-Methods' not in response.headers:
        response.headers['Access-Control-Allow-Methods'] = 'GET,PUT,POST,PATCH,DELETE,OPTIONS'
    return response

# Manejar solicitudes OPTIONS para CORS
@app.before_request
def handle_preflight():
    if request.method == "OPTIONS":
        response = jsonify({})
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type,Authorization'
        response.headers['Access-Control-Allow-Methods'] = 'GET,PUT,POST,PATCH,DELETE,OPTIONS'
        return response

# Configuración de la base de datos
app.config['SQLALCHEMY_DATABASE_URI'] = 'postgresql://postgres:UTEM2022@localhost/INJUV'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Configuración de Flask-Mail
# Para Gmail, necesitas usar una "Contraseña de aplicación" en lugar de tu contraseña normal
# Ve a: https://myaccount.google.com/apppasswords
# También puedes configurar estas variables en un archivo .env o como variables de entorno del sistema
app.config['MAIL_SERVER'] = os.environ.get('MAIL_SERVER', 'smtp.gmail.com')
app.config['MAIL_PORT'] = int(os.environ.get('MAIL_PORT', 587))
app.config['MAIL_USE_TLS'] = os.environ.get('MAIL_USE_TLS', 'True').lower() in ['true', '1', 'yes']
app.config['MAIL_USE_SSL'] = os.environ.get('MAIL_USE_SSL', 'False').lower() in ['true', '1', 'yes']
app.config['MAIL_USERNAME'] = os.environ.get('MAIL_USERNAME', '')  # Tu email
app.config['MAIL_PASSWORD'] = os.environ.get('MAIL_PASSWORD', '')  # Tu contraseña de aplicación
app.config['MAIL_DEFAULT_SENDER'] = os.environ.get('MAIL_DEFAULT_SENDER', app.config['MAIL_USERNAME'])

# Si no hay credenciales configuradas, Flask-Mail se inicializará pero no enviará emails
# Esto permite que la aplicación funcione sin email configurado (solo mostrará warnings)

db = SQLAlchemy(app)
mail = Mail(app)


# Modelo de prueba (opcional)
class Usuario(db.Model):
    __tablename__ = 'usuarios'
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100))
    apellido = db.Column(db.String(100))
    rut = db.Column(db.String(20), nullable=True)  # Hacer opcional
    email = db.Column(db.String(150))
    telefono = db.Column(db.String(20), nullable=True)
    region = db.Column(db.String(100), nullable=True)
    ciudad = db.Column(db.String(100), nullable=True)
    comuna = db.Column(db.String(100), nullable=True)
    sexo = db.Column(db.String(100), nullable=True)
    fecha_nacimiento = db.Column(db.Date, nullable=True)
    password_hash = db.Column(db.String(255))
    rol = db.Column(db.String(100), nullable=True)
    hora_voluntariado = db.Column(db.Integer, nullable=True)
    certificado_voluntariado = db.Column(db.JSON, default=list, nullable=True)
    certificado_personales = db.Column(db.JSON, default=list, nullable=True)
    organizacion_afiliada = db.Column(db.Integer)
    created_at = db.Column(db.DateTime, nullable=True)
    # Nombre de archivo guardado en disco (carpeta perfiles_usuarios/). Ejecutar migración SQL si la columna no existe.
    foto_perfil = db.Column(db.String(255), nullable=True)
       
class Organizacion(db.Model):
    __tablename__ = 'organizaciones'
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=True)
    siglas_nombre = db.Column(db.String(100), nullable=True)
    documentos_legales = db.Column(db.JSON, default=list, nullable=True)
    rut = db.Column(db.String(20), nullable=True)
    email_contacto = db.Column(db.String(150), nullable=True)
    fecha_creacion = db.Column(db.Date, nullable=True)
    telefono_contacto = db.Column(db.String(20), nullable=True)
    region = db.Column(db.String(100), nullable=True)
    ciudad = db.Column(db.String(100), nullable=True)
    comuna = db.Column(db.String(100), nullable=True)
    descripcion = db.Column(db.Text, nullable=True)
    descripcion_breve = db.Column(db.String(500), nullable=True)
    id_usuario_org = db.Column(db.Integer, db.ForeignKey('usuarios.id'), nullable=True)
    area_trabajo = db.Column(db.String(100), nullable=True)
    tipo_org = db.Column(db.String(100), nullable=True)
    sitio_web = db.Column(db.Text, nullable=True)
    redes_sociales = db.Column(db.JSON, default=list, nullable=True)
    experiencia_anios = db.Column(db.Integer, nullable=True)
    voluntarios_anuales = db.Column(db.String(100), nullable=True)
    certificacion = db.Column(db.JSON, default=list, nullable=True)
    # Nombre de archivo del logo en disco (carpeta logos_organizaciones/). Ejecutar migración SQL si la columna no existe.
    logo_filename = db.Column(db.String(255), nullable=True)
    # reseña_organizacion = db.Column(db.String(500), nullable=True)  # Comentado: columna no existe en BD
    created_at = db.Column(db.DateTime, nullable=True)

class SolicitudOrganizacion(db.Model):
    __tablename__ = 'solicitudes_organizacion'
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(150), nullable=False)
    rut = db.Column(db.String(20), nullable=True)
    email_contacto = db.Column(db.String(150), nullable=True)
    fecha_creacion = db.Column(db.Date, nullable=True)
    region = db.Column(db.String(100), nullable=False)
    ciudad = db.Column(db.String(100), nullable=False)
    comuna = db.Column(db.String(100), nullable=False)
    descripcion = db.Column(db.Text, nullable=False)
    sitio_web = db.Column(db.Text, nullable=True)
    redes_sociales = db.Column(db.JSON, default=list, nullable=True)
    id_usuario_org = db.Column(db.Integer, db.ForeignKey('usuarios.id'), nullable=False)
    estado = db.Column(db.String(20), nullable=False, default='pendiente')  # pendiente | aprobada | rechazada
    comentario_revision = db.Column(db.Text, nullable=True)
    revisado_por_admin_id = db.Column(db.Integer, db.ForeignKey('usuarios.id'), nullable=True)
    reviewed_at = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)

class Postulacion(db.Model):
    __tablename__ = 'postulaciones'
    id = db.Column(db.Integer, primary_key=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuarios.id'))
    oportunidad_id = db.Column(db.Integer, db.ForeignKey('oportunidades.id'))
    estado = db.Column(db.String(50))
    motivo_no_seleccion = db.Column(db.String(255))
    motivo_no_seleccion_otro = db.Column(db.Text)  # Corregido: debe coincidir con el nombre en la BD
    estado_confirmacion = db.Column(db.String(20))
    asistencia_capacitacion = db.Column(db.String(20))
    asistencia_actividad = db.Column(db.String(20))
    tiene_certificado = db.Column(db.Boolean)
    ruta_certificado_pdf = db.Column(db.Text)
    resena_org_sobre_voluntario = db.Column(db.Text)
    resena_org_publica = db.Column(db.Boolean)
    calificacion_org = db.Column(db.Numeric(3, 1))  # Permite valores de 0.0 a 5.0 con un decimal
    reseña_org = db.Column(db.Text)  # Reseña del usuario sobre la organización
    horas_voluntariado = db.Column(db.Integer, nullable=True)  # Horas de voluntariado específicas de esta postulación
    # calificacion_usuario_org se define condicionalmente o se maneja con getattr
    created_at = db.Column(db.DateTime)
    updated_at = db.Column(db.DateTime)

class Oportunidad(db.Model):
    __tablename__ = 'oportunidades'

    id = db.Column(db.Integer, primary_key=True)
    organizacion_id = db.Column(db.Integer, nullable=False)
    titulo = db.Column(db.String(200), nullable=False)
    descripcion = db.Column(db.Text, nullable=False)
    meta_postulantes = db.Column(db.Integer)
    cupo_maximo = db.Column(db.Integer)
    fecha_limite_postulacion = db.Column(db.Date)
    fecha_inicio_voluntariado = db.Column(db.Date, nullable=True)
    fecha_fin_voluntariado = db.Column(db.Date, nullable=True)
    horas_voluntariado = db.Column(db.Integer, nullable=True)  # Horas estimadas del voluntariado (definidas por organización)
    estado = db.Column(db.String(20))
    responsable_nombre = db.Column(db.String(50))
    responsable_apellido = db.Column(db.String(50))
    responsable_email = db.Column(db.String(255))
    responsable_email_institucional = db.Column(db.String(255))
    responsable_telefono = db.Column(db.String(30))
    region_opor = db.Column(db.String(255))
    ciudad_opor = db.Column(db.String(255))
    comuna_opor = db.Column(db.String(255))
    tipo_de_voluntariado = db.Column('tipo_de_voluntariado', db.String(100), nullable=True)
    # area_voluntariado es un alias que apunta a la misma columna tipo_de_voluntariado
    @property
    def area_voluntariado(self):
        return self.tipo_de_voluntariado
    
    @area_voluntariado.setter
    def area_voluntariado(self, value):
        self.tipo_de_voluntariado = value
    created_at = db.Column(db.DateTime)

class Noticia(db.Model):
    __tablename__ = 'noticias'
    
    id = db.Column(db.Integer, primary_key=True)
    titulo = db.Column(db.String(200), nullable=False)
    contenido = db.Column(db.Text, nullable=False)
    resumen = db.Column(db.String(500), nullable=True)
    autor_id = db.Column(db.Integer, db.ForeignKey('usuarios.id'), nullable=True)
    estado = db.Column(db.String(20), default='activa')  # activa, inactiva, borrador
    imagen_noticia = db.Column(db.String(500), nullable=True)
    fecha_publicacion = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)


# Repositorio: Biblioteca y Academia
class BibliotecaTematica(db.Model):
    __tablename__ = 'biblioteca_tematicas'
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(80), nullable=False, unique=True)


class BibliotecaDocumento(db.Model):
    __tablename__ = 'biblioteca_documentos'
    id = db.Column(db.Integer, primary_key=True)
    nombre_archivo = db.Column(db.String(255), nullable=False)
    archivo_filename = db.Column(db.String(255), nullable=False)
    archivo_mime = db.Column(db.String(120), nullable=True)
    archivo_tamano_bytes = db.Column(db.BigInteger, nullable=True)
    autor = db.Column(db.String(150), nullable=True)
    fecha_edicion = db.Column(db.DateTime, nullable=True)
    descripcion = db.Column(db.String(500), nullable=True)
    tematica_id = db.Column(db.Integer, db.ForeignKey('biblioteca_tematicas.id', ondelete='SET NULL'), nullable=True)
    organizacion_id = db.Column(db.Integer, db.ForeignKey('organizaciones.id', ondelete='SET NULL'), nullable=True)
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.now)
    updated_at = db.Column(db.DateTime, nullable=False, default=datetime.now, onupdate=datetime.now)


# Categorías válidas para documentos Academia (slug -> etiqueta en front)
ACADEMIA_CATEGORIAS = ('guias_manuales', 'plantillas_formatos', 'videos_tutoriales')


class AcademiaDocumento(db.Model):
    __tablename__ = 'academia_documentos'
    id = db.Column(db.Integer, primary_key=True)
    organizacion_id = db.Column(db.Integer, db.ForeignKey('organizaciones.id', ondelete='CASCADE'), nullable=True)  # null = documento global (admin)
    estado = db.Column(db.String(20), default='pendiente', nullable=False)  # pendiente | aprobado | rechazado
    categoria = db.Column(db.String(80), nullable=True)  # guias_manuales | plantillas_formatos | videos_tutoriales
    nombre_archivo = db.Column(db.String(255), nullable=False)
    archivo_filename = db.Column(db.String(255), nullable=False)
    archivo_mime = db.Column(db.String(120), nullable=True)
    archivo_tamano_bytes = db.Column(db.BigInteger, nullable=True)
    autor = db.Column(db.String(150), nullable=True)
    fecha_edicion = db.Column(db.DateTime, nullable=True)
    descripcion = db.Column(db.String(500), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)


@app.route("/")
def home():
    return "Conexion exitosa con Flask + PostgreSQL"

# Endpoint para registrar usuarios (útil para crear usuarios de prueba)
@app.route("/api/auth/register", methods=["POST"])
def register():
    try:
        data = request.json
        
        email = data.get('email')
        password = data.get('password')
        nombre = data.get('nombre', '')
        apellido = data.get('apellido', '')
        rut = data.get('rut', '').strip().upper()
        rol = data.get('rol', 'user')
        fecha_nacimiento_str = data.get('fecha_nacimiento')
        
        if not email or not password:
            return jsonify({
                'success': False,
                'error': 'Email y contraseña son requeridos'
            }), 400
        
        if not rut:
            return jsonify({
                'success': False,
                'error': 'El RUT es requerido'
            }), 400
        
        # Verificar si el RUT ya está registrado
        rut_limpio = rut.replace('.', '').replace('-', '')
        usuario_rut_existente = Usuario.query.filter_by(rut=rut_limpio).first()
        if usuario_rut_existente:
            return jsonify({
                'success': False,
                'error': 'El RUT ya está registrado'
            }), 400
        
        # Convertir fecha de nacimiento de string a Date si se proporciona
        fecha_nacimiento = None
        if fecha_nacimiento_str:
            try:
                from datetime import datetime
                fecha_nacimiento = datetime.strptime(fecha_nacimiento_str, '%Y-%m-%d').date()
            except ValueError:
                return jsonify({
                    'success': False,
                    'error': 'Formato de fecha de nacimiento inválido'
                }), 400
        
        # Verificar si el usuario ya existe
        usuario_existente = Usuario.query.filter_by(email=email).first()
        if usuario_existente:
            return jsonify({
                'success': False,
                'error': 'El email ya está registrado'
            }), 400
        
        # Crear nuevo usuario (campos opcionales pueden ser None)
        nuevo_usuario = Usuario(
            email=email,
            password_hash=generate_password_hash(password),
            nombre=nombre,
            apellido=apellido,
            rol=rol,
            rut=rut_limpio,  # RUT sin puntos ni guion
            telefono=None,
            region=None,
            comuna=None,
            sexo=None,
            fecha_nacimiento=fecha_nacimiento,
            created_at=datetime.now()
        )
        
        try:
            db.session.add(nuevo_usuario)
            db.session.commit()
        except Exception as db_error:
            # Si falla por restricción de NOT NULL en rut, intentar con valor por defecto
            if 'rut' in str(db_error).lower() or 'not null' in str(db_error).lower():
                # Intentar con un RUT temporal basado en el email
                rut_temporal = email.split('@')[0].replace('.', '').replace('-', '')[:12] or 'TEMP'
                nuevo_usuario.rut = rut_temporal
                db.session.rollback()
                db.session.add(nuevo_usuario)
                db.session.commit()
            else:
                raise
        
        return jsonify({
            'success': True,
            'message': 'Usuario registrado exitosamente',
            'user': {
                'id': nuevo_usuario.id,
                'email': nuevo_usuario.email,
                'nombre': nuevo_usuario.nombre,
                'apellido': nuevo_usuario.apellido,
                'rol': nuevo_usuario.rol
            }
        }), 201
        
    except Exception as e:
        db.session.rollback()
        # Registrar el error completo en la consola del servidor para depuración
        print("Error en /api/auth/register:", repr(e))
        return jsonify({
            'success': False,
            'error': 'Error interno al registrar usuario. Por favor, inténtalo nuevamente o contacta al administrador.'
        }), 500

# Endpoint de login
@app.route("/api/auth/login", methods=["POST"])
def login():
    try:
        data = request.json
        email = data.get('email')
        password = data.get('password')
        
        if not email or not password:
            return jsonify({
                'success': False,
                'error': 'Email y contraseña son requeridos'
            }), 400
        
        # Buscar usuario por email
        usuario = Usuario.query.filter_by(email=email).first()
        
        if not usuario:
            return jsonify({
                'success': False,
                'error': 'Email o contraseña incorrectos'
            }), 401
        
        # Verificar contraseña
        if not check_password_hash(usuario.password_hash, password):
            return jsonify({
                'success': False,
                'error': 'Email o contraseña incorrectos'
            }), 401
        
        # Asegurar que el rol siempre tenga un valor
        rol_usuario = usuario.rol if usuario.rol else 'user'
        
        return jsonify({
            'success': True,
            'message': 'Inicio de sesión exitoso',
            'user': {
                'id': usuario.id,
                'email': usuario.email,
                'nombre': usuario.nombre,
                'apellido': usuario.apellido,
                'rol': rol_usuario
            }
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Endpoint para actualizar información de contacto del usuario
@app.route("/api/usuario/contacto", methods=["PUT"])
def actualizar_contacto():
    try:
        data = request.json
        
        user_id = data.get('user_id')
        email = data.get('email')
        telefono = data.get('telefono')
        region = data.get('region')
        ciudad = data.get('ciudad')
        comuna = data.get('comuna')
        
        if not user_id:
            return jsonify({'success': False, 'error': 'user_id es requerido'}), 400
        
        # Buscar el usuario
        usuario = Usuario.query.get(user_id)
        
        if not usuario:
            return jsonify({'success': False, 'error': 'Usuario no encontrado'}), 404
        
        # Actualizar los campos (permitir valores vacíos para limpiar campos)
        if email is not None:
            usuario.email = email if email else None
        if telefono is not None:
            usuario.telefono = telefono if telefono else None
        if region is not None:
            usuario.region = region if region else None
        if ciudad is not None:
            usuario.ciudad = ciudad if ciudad else None
        if comuna is not None:
            usuario.comuna = comuna if comuna else None
        
        # Guardar en la base de datos
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Información de contacto actualizada exitosamente',
            'contacto': {
                'email': usuario.email,
                'telefono': usuario.telefono,
                'region': usuario.region,
                'ciudad': usuario.ciudad,
                'comuna': usuario.comuna
            }
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

# Endpoint para obtener información de contacto del usuario
@app.route("/api/usuario/<int:user_id>/contacto", methods=["GET"])
def obtener_contacto(user_id):
    try:
        usuario = Usuario.query.get(user_id)
        
        if not usuario:
            return jsonify({'success': False, 'error': 'Usuario no encontrado'}), 404
        
        return jsonify({
            'success': True,
            'contacto': {
                'email': usuario.email or '',
                'telefono': usuario.telefono or '',
                'region': usuario.region or '',
                'ciudad': usuario.ciudad or '',
                'comuna': usuario.comuna or ''
            }
        }), 200
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# Endpoint para obtener información completa del usuario
@app.route("/api/usuario/<int:user_id>", methods=["GET"])
def obtener_usuario(user_id):
    try:
        usuario = Usuario.query.get(user_id)
        
        if not usuario:
            return jsonify({'success': False, 'error': 'Usuario no encontrado'}), 404
        
        return jsonify({
            'success': True,
            'usuario': {
                'id': usuario.id,
                'nombre': usuario.nombre or '',
                'apellido': usuario.apellido or '',
                'email': usuario.email or '',
                'telefono': usuario.telefono or '',
                'region': usuario.region or '',
                'comuna': usuario.comuna or '',
                'rol': usuario.rol or '',
                'rut': usuario.rut or '',
                'fecha_nacimiento': usuario.fecha_nacimiento.isoformat() if usuario.fecha_nacimiento else None,
                'hora_voluntariado': usuario.hora_voluntariado if usuario.hora_voluntariado else 0,
                'foto_perfil': usuario.foto_perfil or None
            }
        }), 200
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route("/api/usuario/<int:user_id>/foto-perfil", methods=["GET"])
def obtener_foto_perfil_usuario(user_id):
    try:
        usuario = Usuario.query.get(user_id)
        if not usuario or not usuario.foto_perfil:
            return jsonify({'success': False, 'error': 'Sin foto de perfil'}), 404
        filepath = os.path.join(os.getcwd(), 'perfiles_usuarios', usuario.foto_perfil)
        if not os.path.isfile(filepath):
            return jsonify({'success': False, 'error': 'Archivo no encontrado'}), 404
        ext = os.path.splitext(usuario.foto_perfil)[1].lower()
        mimetype = 'image/png'
        if ext in ('.jpg', '.jpeg'):
            mimetype = 'image/jpeg'
        elif ext == '.webp':
            mimetype = 'image/webp'
        elif ext == '.gif':
            mimetype = 'image/gif'
        return send_file(filepath, mimetype=mimetype)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route("/api/usuario/<int:user_id>/foto-perfil", methods=["POST"])
def subir_foto_perfil_usuario(user_id):
    from werkzeug.utils import secure_filename
    try:
        usuario = Usuario.query.get(user_id)
        if not usuario:
            return jsonify({'success': False, 'error': 'Usuario no encontrado'}), 404
        if 'imagen' not in request.files:
            return jsonify({'success': False, 'error': 'No se proporcionó imagen'}), 400
        file = request.files['imagen']
        if not file or file.filename == '':
            return jsonify({'success': False, 'error': 'Archivo vacío'}), 400
        file_ext = os.path.splitext(file.filename)[1].lower()
        if file_ext not in {'.png', '.jpg', '.jpeg', '.webp', '.gif'}:
            file_ext = '.png'
        upload_dir = os.path.join(os.getcwd(), 'perfiles_usuarios')
        os.makedirs(upload_dir, exist_ok=True)
        if usuario.foto_perfil:
            old_path = os.path.join(upload_dir, usuario.foto_perfil)
            if os.path.isfile(old_path):
                try:
                    os.remove(old_path)
                except OSError:
                    pass
        filename = secure_filename(
            f"user_{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{os.urandom(4).hex()}{file_ext}"
        )
        filepath = os.path.join(upload_dir, filename)
        file.save(filepath)
        usuario.foto_perfil = filename
        db.session.commit()
        return jsonify({'success': True, 'filename': filename}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# Eliminar cuenta de usuario
@app.route("/api/usuarios/<int:user_id>", methods=["DELETE"])
def eliminar_usuario(user_id):
    try:
        data = request.json or {}
        confirmar_eliminacion = data.get('confirmar', False)
        
        if not confirmar_eliminacion:
            return jsonify({
                'success': False,
                'error': 'Se requiere confirmación para eliminar la cuenta'
            }), 400
        
        # Buscar el usuario
        usuario = Usuario.query.get(user_id)
        
        if not usuario:
            return jsonify({
                'success': False,
                'error': 'Usuario no encontrado'
            }), 404
        
        # Obtener todas las postulaciones del usuario
        postulaciones = Postulacion.query.filter_by(usuario_id=user_id).all()
        
        # Verificar si el usuario es administrador de una organización
        organizacion = Organizacion.query.filter_by(id_usuario_org=user_id).first()
        
        if organizacion:
            if organizacion.logo_filename:
                logo_path = os.path.join(os.getcwd(), 'logos_organizaciones', organizacion.logo_filename)
                if os.path.isfile(logo_path):
                    try:
                        os.remove(logo_path)
                    except OSError:
                        pass
            # Si tiene una organización, eliminar también las oportunidades y postulaciones asociadas
            oportunidades = Oportunidad.query.filter_by(organizacion_id=organizacion.id).all()
            
            for oportunidad in oportunidades:
                # Eliminar postulaciones de cada oportunidad
                postulaciones_oportunidad = Postulacion.query.filter_by(oportunidad_id=oportunidad.id).all()
                for post in postulaciones_oportunidad:
                    db.session.delete(post)
                # Eliminar la oportunidad
                db.session.delete(oportunidad)
            
            # Eliminar la organización
            db.session.delete(organizacion)
        
        # Eliminar todas las postulaciones del usuario
        for postulacion in postulaciones:
            db.session.delete(postulacion)
        
        if usuario.foto_perfil:
            foto_path = os.path.join(os.getcwd(), 'perfiles_usuarios', usuario.foto_perfil)
            if os.path.isfile(foto_path):
                try:
                    os.remove(foto_path)
                except OSError:
                    pass
        
        # Eliminar el usuario
        db.session.delete(usuario)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Cuenta eliminada exitosamente'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        import traceback
        error_trace = traceback.format_exc()
        print(f"Error al eliminar usuario: {error_trace}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Obtener organización por usuario admin
@app.route("/api/organizaciones/usuario/<int:usuario_id>", methods=["GET"])
def obtener_organizacion_por_usuario(usuario_id):
    try:
        organizacion = Organizacion.query.filter_by(id_usuario_org=usuario_id).first()
        
        if not organizacion:
            return jsonify({
                'success': False,
                'error': 'No se encontró una organización asociada a este usuario'
            }), 404
        
        # Procesar certificaciones (puede ser una lista JSON o None)
        certificaciones = organizacion.certificacion if organizacion.certificacion else []
        if isinstance(certificaciones, str):
            try:
                import json
                certificaciones = json.loads(certificaciones)
            except:
                certificaciones = []
        
        return jsonify({
            'success': True,
            'organizacion': {
                'id': organizacion.id,
                'nombre': organizacion.nombre,
                'rut': organizacion.rut or '',
                'email_contacto': organizacion.email_contacto,
                'telefono_contacto': organizacion.telefono_contacto,
                'region': organizacion.region,
                'comuna': organizacion.comuna,
                'descripcion': organizacion.descripcion or '',
                'id_usuario_org': organizacion.id_usuario_org,
                'certificacion': certificaciones,
                'logo_filename': organizacion.logo_filename or None,
                'created_at': organizacion.created_at.strftime('%Y-%m-%d %H:%M:%S') if organizacion.created_at else None
            }
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Obtener organización por ID
@app.route("/api/organizaciones/<int:organizacion_id>", methods=["GET"])
def obtener_organizacion(organizacion_id):
    try:
        organizacion = Organizacion.query.get(organizacion_id)
        
        if not organizacion:
            return jsonify({
                'success': False,
                'error': 'Organización no encontrada'
            }), 404
        
        # Procesar certificaciones (puede ser una lista JSON o None)
        certificaciones = organizacion.certificacion if organizacion.certificacion else []
        if isinstance(certificaciones, str):
            try:
                import json
                certificaciones = json.loads(certificaciones)
            except:
                certificaciones = []
        
        return jsonify({
            'success': True,
            'organizacion': {
                'id': organizacion.id,
                'nombre': organizacion.nombre,
                'rut': organizacion.rut or '',
                'email_contacto': organizacion.email_contacto,
                'telefono_contacto': organizacion.telefono_contacto,
                'region': organizacion.region,
                'comuna': organizacion.comuna,
                'descripcion': organizacion.descripcion or '',
                'id_usuario_org': organizacion.id_usuario_org,
                'certificacion': certificaciones,
                'logo_filename': organizacion.logo_filename or None,
                'created_at': organizacion.created_at.strftime('%Y-%m-%d %H:%M:%S') if organizacion.created_at else None
            }
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route("/api/organizaciones/<int:organizacion_id>/logo", methods=["GET"])
def obtener_logo_organizacion(organizacion_id):
    try:
        organizacion = Organizacion.query.get(organizacion_id)
        if not organizacion or not organizacion.logo_filename:
            return jsonify({'success': False, 'error': 'Sin logo'}), 404
        filepath = os.path.join(os.getcwd(), 'logos_organizaciones', organizacion.logo_filename)
        if not os.path.isfile(filepath):
            return jsonify({'success': False, 'error': 'Archivo no encontrado'}), 404
        ext = os.path.splitext(organizacion.logo_filename)[1].lower()
        mimetype = 'image/png'
        if ext in ('.jpg', '.jpeg'):
            mimetype = 'image/jpeg'
        elif ext == '.webp':
            mimetype = 'image/webp'
        elif ext == '.gif':
            mimetype = 'image/gif'
        return send_file(filepath, mimetype=mimetype)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route("/api/organizaciones/<int:organizacion_id>/logo", methods=["POST"])
def subir_logo_organizacion(organizacion_id):
    from werkzeug.utils import secure_filename
    try:
        organizacion = Organizacion.query.get(organizacion_id)
        if not organizacion:
            return jsonify({'success': False, 'error': 'Organización no encontrada'}), 404
        if 'imagen' not in request.files:
            return jsonify({'success': False, 'error': 'No se proporcionó imagen'}), 400
        file = request.files['imagen']
        if not file or file.filename == '':
            return jsonify({'success': False, 'error': 'Archivo vacío'}), 400
        file_ext = os.path.splitext(file.filename)[1].lower()
        if file_ext not in {'.png', '.jpg', '.jpeg', '.webp', '.gif'}:
            file_ext = '.png'
        upload_dir = os.path.join(os.getcwd(), 'logos_organizaciones')
        os.makedirs(upload_dir, exist_ok=True)
        if organizacion.logo_filename:
            old_path = os.path.join(upload_dir, organizacion.logo_filename)
            if os.path.isfile(old_path):
                try:
                    os.remove(old_path)
                except OSError:
                    pass
        filename = secure_filename(
            f"org_{organizacion_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{os.urandom(4).hex()}{file_ext}"
        )
        filepath = os.path.join(upload_dir, filename)
        file.save(filepath)
        organizacion.logo_filename = filename
        db.session.commit()
        return jsonify({'success': True, 'filename': filename}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# Actualizar información de la organización
@app.route("/api/organizaciones/<int:organizacion_id>", methods=["PUT"])
def actualizar_organizacion(organizacion_id):
    try:
        data = request.json
        
        # Buscar la organización
        organizacion = Organizacion.query.get(organizacion_id)
        
        if not organizacion:
            return jsonify({
                'success': False,
                'error': 'Organización no encontrada'
            }), 404
        
        # Verificar que el usuario que hace la solicitud es el administrador
        id_usuario_org = data.get('id_usuario_org')
        if id_usuario_org and organizacion.id_usuario_org != id_usuario_org:
            return jsonify({
                'success': False,
                'error': 'No tienes permisos para actualizar esta organización'
            }), 403
        
        # Actualizar campos (solo los que se proporcionen)
        if 'nombre' in data:
            organizacion.nombre = data.get('nombre')
        if 'rut' in data:
            rut = data.get('rut', '').strip().upper().replace('.', '').replace('-', '')
            # Verificar que el RUT no esté en uso por otra organización
            if rut:
                otra_org = Organizacion.query.filter_by(rut=rut).first()
                if otra_org and otra_org.id != organizacion_id:
                    return jsonify({
                        'success': False,
                        'error': 'El RUT ya está registrado por otra organización'
                    }), 400
            organizacion.rut = rut if rut else None
        if 'email_contacto' in data:
            email = data.get('email_contacto')
            # Verificar que el email no esté en uso por otra organización
            otra_org = Organizacion.query.filter_by(email_contacto=email).first()
            if otra_org and otra_org.id != organizacion_id:
                return jsonify({
                    'success': False,
                    'error': 'El correo electrónico ya está registrado por otra organización'
                }), 400
            organizacion.email_contacto = email
        if 'telefono_contacto' in data:
            organizacion.telefono_contacto = data.get('telefono_contacto')
        if 'region' in data:
            organizacion.region = data.get('region')
        if 'comuna' in data:
            organizacion.comuna = data.get('comuna')
        if 'descripcion' in data:
            organizacion.descripcion = data.get('descripcion')
        if 'certificacion' in data:
            # Actualizar certificaciones
            certificaciones = data.get('certificacion')
            print(f"Recibidas certificaciones para actualizar: {certificaciones}")
            print(f"Tipo de certificaciones recibidas: {type(certificaciones)}")
            
            if isinstance(certificaciones, list):
                organizacion.certificacion = certificaciones
                print(f"Certificaciones guardadas como lista: {organizacion.certificacion}")
            elif isinstance(certificaciones, str):
                try:
                    import json
                    organizacion.certificacion = json.loads(certificaciones)
                    print(f"Certificaciones parseadas desde string: {organizacion.certificacion}")
                except Exception as e:
                    print(f"Error al parsear certificaciones: {e}")
                    organizacion.certificacion = []
            else:
                print(f"Tipo de certificaciones no reconocido, estableciendo lista vacía")
                organizacion.certificacion = []
        
        db.session.commit()
        print(f"Certificaciones después del commit: {organizacion.certificacion}")
        
        # Procesar certificaciones para la respuesta
        certificaciones_respuesta = organizacion.certificacion if organizacion.certificacion else []
        if isinstance(certificaciones_respuesta, str):
            try:
                import json
                certificaciones_respuesta = json.loads(certificaciones_respuesta)
            except:
                certificaciones_respuesta = []
        
        return jsonify({
            'success': True,
            'message': 'Información de la organización actualizada exitosamente',
            'organizacion': {
                'id': organizacion.id,
                'nombre': organizacion.nombre,
                'rut': organizacion.rut or '',
                'email_contacto': organizacion.email_contacto,
                'telefono_contacto': organizacion.telefono_contacto,
                'region': organizacion.region,
                'comuna': organizacion.comuna,
                'descripcion': organizacion.descripcion or '',
                'certificacion': certificaciones_respuesta
            }
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Endpoint para subir archivo de certificación de organización
@app.route("/api/organizaciones/<int:organizacion_id>/certificacion/upload", methods=["POST"])
def subir_certificacion_organizacion(organizacion_id):
    try:
        # Verificar que se envió un archivo
        if 'archivo' not in request.files:
            return jsonify({
                'success': False,
                'error': 'No se proporcionó ningún archivo'
            }), 400
        
        file = request.files['archivo']
        
        if file.filename == '':
            return jsonify({
                'success': False,
                'error': 'No se seleccionó ningún archivo'
            }), 400
        
        # Verificar que la organización existe
        organizacion = Organizacion.query.get(organizacion_id)
        if not organizacion:
            return jsonify({
                'success': False,
                'error': 'Organización no encontrada'
            }), 404
        
        # Verificar permisos
        id_usuario_org = request.form.get('id_usuario_org')
        if id_usuario_org and organizacion.id_usuario_org != int(id_usuario_org):
            return jsonify({
                'success': False,
                'error': 'No tienes permisos para subir certificaciones a esta organización'
            }), 403
        
        # Validar tipo de archivo
        allowed_extensions = {'.pdf', '.jpg', '.jpeg', '.png'}
        file_ext = os.path.splitext(file.filename)[1].lower()
        if file_ext not in allowed_extensions:
            return jsonify({
                'success': False,
                'error': f'Tipo de archivo no permitido. Formatos permitidos: {", ".join(allowed_extensions)}'
            }), 400
        
        # Validar tamaño (máximo 10MB)
        file.seek(0, os.SEEK_END)
        file_size = file.tell()
        file.seek(0)
        max_size = 10 * 1024 * 1024  # 10MB
        if file_size > max_size:
            return jsonify({
                'success': False,
                'error': 'El archivo es demasiado grande. Tamaño máximo: 10MB'
            }), 400
        
        # Crear directorio de certificaciones de organizaciones si no existe
        certificaciones_dir = os.path.join(os.getcwd(), 'certificaciones_organizaciones')
        if not os.path.exists(certificaciones_dir):
            os.makedirs(certificaciones_dir)
        
        # Generar nombre único para el archivo
        from werkzeug.utils import secure_filename
        filename = f"cert_org_{organizacion_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{file_ext}"
        filename = secure_filename(filename)
        filepath = os.path.join(certificaciones_dir, filename)
        
        # Guardar el archivo
        file.save(filepath)
        
        # Retornar la ruta relativa para almacenar en la base de datos
        ruta_relativa = f"certificaciones_organizaciones/{filename}"
        
        return jsonify({
            'success': True,
            'message': 'Archivo subido exitosamente',
            'ruta': ruta_relativa,
            'filename': filename
        }), 200
        
    except Exception as e:
        import traceback
        print(f"Error al subir certificación: {str(e)}")
        print(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Endpoint para descargar archivo de certificación
@app.route("/api/organizaciones/certificacion/<path:filename>", methods=["GET"])
def descargar_certificacion_organizacion(filename):
    try:
        filepath = os.path.join(os.getcwd(), 'certificaciones_organizaciones', filename)
        if os.path.exists(filepath):
            return send_file(filepath, as_attachment=True)
        else:
            return jsonify({
                'success': False,
                'error': 'Archivo no encontrado'
            }), 404
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Endpoint para registrar una organización
@app.route("/api/organizacion/registrar", methods=["POST"])
def registrar_organizacion():
    try:
        data = request.json
        
        nombre = data.get('nombre')
        rut = data.get('rut', '').strip().upper().replace('.', '').replace('-', '') if data.get('rut') else ''
        fecha_creacion_str = data.get('fecha_creacion')
        region = data.get('region')
        ciudad = data.get('ciudad')
        comuna = data.get('comuna')
        descripcion = data.get('descripcion')
        sitio_web = data.get('sitio_web')
        redes_sociales = data.get('redes_sociales')
        id_usuario_org = data.get('id_usuario_org')
        
        # Validaciones de campos obligatorios
        if not nombre or not fecha_creacion_str or not region or not ciudad or not comuna or not descripcion:
            return jsonify({
                'success': False,
                'error': 'Los campos obligatorios son: nombre, fecha de fundación, región, ciudad, comuna y descripción'
            }), 400
        
        # Convertir fecha de creación de string a Date
        fecha_creacion = None
        if fecha_creacion_str:
            try:
                fecha_creacion = datetime.strptime(fecha_creacion_str, '%Y-%m-%d').date()
            except ValueError:
                return jsonify({
                    'success': False,
                    'error': 'Formato de fecha inválido. Use el formato YYYY-MM-DD'
                }), 400
        
        if not id_usuario_org:
            return jsonify({
                'success': False,
                'error': 'ID de usuario organizador es requerido'
            }), 400
        
        # Verificar que el usuario existe
        usuario = Usuario.query.get(id_usuario_org)
        if not usuario:
            return jsonify({
                'success': False,
                'error': 'Usuario no encontrado'
            }), 404
        
        # Usar el email del usuario como email_contacto si no se proporciona
        email_contacto = usuario.email
        
        # Evitar que un usuario envíe múltiples formularios pendientes o tenga ya una organización activa
        organizacion_usuario_existente = Organizacion.query.filter_by(id_usuario_org=id_usuario_org).first()
        if organizacion_usuario_existente:
            return jsonify({
                'success': False,
                'error': 'Este usuario ya administra una organización.'
            }), 400

        solicitud_pendiente = SolicitudOrganizacion.query.filter_by(
            id_usuario_org=id_usuario_org,
            estado='pendiente'
        ).first()
        if solicitud_pendiente:
            return jsonify({
                'success': False,
                'error': 'Ya tienes un formulario pendiente de revisión.'
            }), 400

        # Verificar si el RUT ya está registrado (si se proporciona)
        if rut:
            organizacion_existente = Organizacion.query.filter_by(rut=rut).first()
            if organizacion_existente:
                return jsonify({
                    'success': False,
                    'error': 'El RUT ya está registrado'
                }), 400
            solicitud_rut_existente = SolicitudOrganizacion.query.filter(
                SolicitudOrganizacion.rut == rut,
                SolicitudOrganizacion.estado.in_(['pendiente', 'aprobada'])
            ).first()
            if solicitud_rut_existente:
                return jsonify({
                    'success': False,
                    'error': 'Ya existe una solicitud activa con ese RUT'
                }), 400
        
        # Procesar redes sociales: mantener como lista o convertir a lista
        redes_sociales_data = None
        if redes_sociales:
            if isinstance(redes_sociales, list):
                redes_sociales_data = redes_sociales
            elif isinstance(redes_sociales, str):
                # Si es string, convertir a lista
                redes_sociales_data = [linea.strip() for linea in redes_sociales.split('\n') if linea.strip()]
            else:
                redes_sociales_data = [str(redes_sociales)]
        
        # Crear solicitud en estado pendiente para revisión del administrador
        nueva_solicitud = SolicitudOrganizacion(
            nombre=nombre,
            rut=rut if rut else None,
            email_contacto=email_contacto,
            fecha_creacion=fecha_creacion,
            region=region,
            ciudad=ciudad,
            comuna=comuna,
            descripcion=descripcion,
            sitio_web=sitio_web if sitio_web else None,
            redes_sociales=redes_sociales_data if redes_sociales_data else [],
            id_usuario_org=id_usuario_org,
            created_at=datetime.now()
        )
        
        db.session.add(nueva_solicitud)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Formulario enviado correctamente. Un administrador revisará tu solicitud.',
            'solicitud': {
                'id': nueva_solicitud.id,
                'nombre': nueva_solicitud.nombre,
                'email_contacto': nueva_solicitud.email_contacto,
                'region': nueva_solicitud.region,
                'ciudad': nueva_solicitud.ciudad,
                'comuna': nueva_solicitud.comuna,
                'estado': nueva_solicitud.estado
            }
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# ============================================
# ENDPOINTS PARA OPORTUNIDADES
# ============================================

# Listar todas las oportunidades (con filtros opcionales)
@app.route("/api/oportunidades", methods=["GET"])
def listar_oportunidades():
    try:
        estado = request.args.get('estado')  # activa, cerrada, todas, o None para todas
        region = request.args.get('region')
        area_voluntariado = request.args.get('area') or request.args.get('tipo_de_voluntariado')  # Filtro por área de voluntariado
        organizacion_id_str = request.args.get('organizacion_id')
        print(f"📥 Parámetros recibidos - estado: '{estado}', region: '{region}', area_voluntariado: '{area_voluntariado}', organizacion_id: '{organizacion_id_str}'")
        organizacion_id = None
        if organizacion_id_str:
            try:
                organizacion_id = int(organizacion_id_str)
                print(f"✅ organizacion_id convertido a int: {organizacion_id}")
            except (ValueError, TypeError) as e:
                print(f"❌ Error: organizacion_id '{organizacion_id_str}' no es un entero válido: {e}")
                organizacion_id = None
        else:
            print(f"⚠️ No se recibió parámetro organizacion_id")
        
        # Si se especifica organizacion_id, usar consulta directa simple (más confiable)
        if organizacion_id:
            print(f"🔍 Buscando oportunidades para organizacion_id = {organizacion_id} (tipo: {type(organizacion_id)})")
            
            # PRIMERO: Verificar todas las oportunidades en la BD sin filtro
            todas_oportunidades = Oportunidad.query.all()
            print(f"📊 Total de oportunidades en BD (sin filtro): {len(todas_oportunidades)}")
            for op in todas_oportunidades:
                print(f"  - ID: {op.id}, Título: {op.titulo}, Organizacion_id: {op.organizacion_id} (tipo: {type(op.organizacion_id)})")
            
            # SEGUNDO: Hacer la consulta con el filtro
            print(f"🔍 Ejecutando consulta: Oportunidad.query.filter(Oportunidad.organizacion_id == {organizacion_id})")
            oportunidades = Oportunidad.query.filter(Oportunidad.organizacion_id == organizacion_id).all()
            print(f"✅ Oportunidades encontradas con filtro: {len(oportunidades)}")
            
            # TERCERO: Si no encuentra nada, intentar con diferentes tipos de comparación
            if len(oportunidades) == 0:
                print(f"⚠️ No se encontraron resultados. Intentando con filter_by...")
                oportunidades = Oportunidad.query.filter_by(organizacion_id=organizacion_id).all()
                print(f"✅ Oportunidades encontradas con filter_by: {len(oportunidades)}")
                
                # Si aún no encuentra nada, verificar si hay un problema de tipo
                if len(oportunidades) == 0:
                    print(f"⚠️ Aún no se encontraron resultados. Verificando tipos...")
                    # Intentar con string también
                    oportunidades_str = Oportunidad.query.filter(Oportunidad.organizacion_id == str(organizacion_id)).all()
                    print(f"✅ Intentando con string '{str(organizacion_id)}': {len(oportunidades_str)}")
                    if len(oportunidades_str) > 0:
                        oportunidades = oportunidades_str
            
            # Si se especifica estado y NO es 'todas', filtrar después en Python
            # IMPORTANTE: Si NO se especifica estado, devolver TODAS las oportunidades de la organización (activas, abiertas, etc.)
            if estado and estado != 'todas' and estado != 'all':
                print(f"🔍 Aplicando filtro de estado: {estado}")
                oportunidades_antes = len(oportunidades)
                # Normalizar estado: 'abierto' -> 'abierta' para consistencia
                estado_normalizado = estado.lower()
                if estado_normalizado == 'abierto':
                    estado_normalizado = 'abierta'
                oportunidades = [op for op in oportunidades if op.estado and op.estado.lower() == estado_normalizado]
                print(f"✅ Oportunidades después del filtro de estado '{estado_normalizado}': {len(oportunidades)} (de {oportunidades_antes})")
            else:
                # Si no se especifica estado o es 'todas', devolver TODAS las oportunidades de la organización
                print(f"🔍 No se aplicará filtro de estado, devolviendo TODAS las oportunidades de la organización (activas, abiertas, cerradas, etc.)")
            
            # Mostrar detalles de las oportunidades encontradas
            if oportunidades:
                for op in oportunidades:
                    print(f"  ✓ ID: {op.id}, Título: {op.titulo}, Estado: '{op.estado}', Organizacion_id: {op.organizacion_id}")
            else:
                # Verificar todas las oportunidades en la BD para debug
                todas_sin_filtro = Oportunidad.query.all()
                print(f"  ⚠️ No se encontraron oportunidades. Total en BD: {len(todas_sin_filtro)}")
                for op in todas_sin_filtro:
                    print(f"    - ID: {op.id}, Título: {op.titulo}, Organizacion_id: {op.organizacion_id} (tipo: {type(op.organizacion_id)})")
        else:
            # Si no hay organizacion_id, usar query normal
            print(f"🔍 No hay organizacion_id, buscando todas las oportunidades")
            print(f"📥 Parámetro estado recibido: '{estado}'")
            
            # PRIMERO: Verificar todas las oportunidades en la BD sin filtro
            todas_oportunidades = Oportunidad.query.all()
            print(f"📊 Total de oportunidades en BD (sin filtro): {len(todas_oportunidades)}")
            for op in todas_oportunidades:
                print(f"  - ID: {op.id}, Título: {op.titulo}, Estado: '{op.estado}', Organizacion_id: {op.organizacion_id}")
            
            query = Oportunidad.query
            
            # Aplicar filtro de estado
            # Por defecto, mostrar activas Y abiertas a menos que se especifique 'todas' o 'all'
            if estado and estado not in ['todas', 'all']:
                print(f"🔍 Aplicando filtro de estado: '{estado}'")
                # Normalizar estado: 'abierto' -> 'abierta'
                estado_normalizado = estado.lower()
                if estado_normalizado == 'abierto':
                    estado_normalizado = 'abierta'
                query = query.filter_by(estado=estado_normalizado)
            elif estado in ['todas', 'all']:
                print(f"🔍 Estado es 'todas' o 'all', no aplicando filtro de estado")
            else:
                # Por defecto, mostrar activas Y abiertas si no se especifica estado
                print(f"🔍 No se especificó estado, usando filtro por defecto: 'activa' o 'abierta'")
                query = query.filter(Oportunidad.estado.in_(['activa', 'abierta']))
            
            # Si se especifica region, usar region_opor de la tabla oportunidades
            if region and region.strip():
                region_clean = region.strip()
                print(f"🔍 Aplicando filtro de región: '{region_clean}'")
                # Usar coincidencia exacta o que contenga el valor completo
                query = query.filter(
                    Oportunidad.region_opor.ilike(f'%{region_clean}%')
                )
                print(f"✅ Query después del filtro de región: {query}")
                print(f"📋 Buscando region_opor que contenga: '{region_clean}'")
            
            # Si se especifica area_voluntariado, filtrar por área (solo si la columna existe)
            if area_voluntariado and area_voluntariado.strip():
                area_voluntariado_clean = area_voluntariado.strip()
                print(f"🔍 Aplicando filtro de tipo de voluntariado: '{area_voluntariado_clean}'")
                try:
                    # Verificar si la columna existe antes de usarla
                    from sqlalchemy import inspect
                    inspector = inspect(Oportunidad)
                    column_names = [col.name for col in inspector.columns]
                    
                    # Intentar con tipo_de_voluntariado primero (nombre real en BD), luego area_voluntariado
                    if 'tipo_de_voluntariado' in column_names:
                        # Usar coincidencia parcial (ilike ya es case-insensitive)
                        query = query.filter(Oportunidad.tipo_de_voluntariado.ilike(f'%{area_voluntariado_clean}%'))
                        print(f"✅ Query después del filtro de tipo de voluntariado: {query}")
                        print(f"📋 Buscando tipo_de_voluntariado que contenga: '{area_voluntariado_clean}'")
                    elif 'area_voluntariado' in column_names:
                        query = query.filter(Oportunidad.area_voluntariado.ilike(f'%{area_voluntariado_clean}%'))
                        print(f"✅ Query después del filtro de área: {query}")
                        print(f"📋 Buscando area_voluntariado que contenga: '{area_voluntariado_clean}'")
                    else:
                        print(f"⚠️ La columna tipo_de_voluntariado/area_voluntariado no existe en la BD, ignorando filtro")
                except Exception as e:
                    # Si hay algún error, ignorar el filtro
                    print(f"⚠️ Error al aplicar filtro de área: {e}, ignorando filtro")
                    pass
            
            try:
                oportunidades = query.all()
                print(f"✅ Oportunidades encontradas después de aplicar filtros SQL: {len(oportunidades)}")
                
                # Filtrar adicionalmente en Python para asegurar que cumplan los criterios
                oportunidades_filtradas = []
                for op in oportunidades:
                    # Verificar estado (debe ser 'activa' si no se especifica otro)
                    op_estado = getattr(op, 'estado', None) or 'activa'
                    if estado and estado not in ['todas', 'all']:
                        if op_estado.lower() != estado.lower():
                            print(f"  ❌ Oportunidad {op.id} descartada: estado '{op_estado}' != '{estado}'")
                            continue
                    elif estado in ['todas', 'all']:
                        # Si se especifica 'todas' o 'all', mostrar todas sin filtrar por estado
                        pass
                    else:
                        # Por defecto, mostrar activas Y abiertas
                        if op_estado.lower() not in ['activa', 'abierta']:
                            print(f"  ❌ Oportunidad {op.id} descartada: estado '{op_estado}' no es 'activa' ni 'abierta' (por defecto)")
                            continue
                    
                    # Verificar región si se especificó (OBLIGATORIO si se especifica)
                    if region and region.strip():
                        op_region = getattr(op, 'region_opor', None)
                        if not op_region or not op_region.strip():
                            print(f"  ❌ Oportunidad {op.id} descartada: region_opor está vacío (filtro requiere: '{region}')")
                            continue
                        
                        op_region_lower = op_region.lower().strip()
                        region_lower = region.strip().lower()
                        
                        # Verificar coincidencia exacta o parcial (más estricto)
                        if region_lower not in op_region_lower:
                            print(f"  ❌ Oportunidad {op.id} descartada: region_opor '{op_region}' no contiene '{region}'")
                            continue
                        print(f"  ✓ Oportunidad {op.id} región OK: '{op_region}' contiene '{region}'")
                    
                    # Verificar tipo de voluntariado si se especificó (OBLIGATORIO si se especifica)
                    if area_voluntariado and area_voluntariado.strip():
                        op_tipo = getattr(op, 'tipo_de_voluntariado', None) or getattr(op, 'area_voluntariado', None)
                        if not op_tipo or not op_tipo.strip():
                            print(f"  ❌ Oportunidad {op.id} descartada: tipo_de_voluntariado está vacío (filtro requiere: '{area_voluntariado}')")
                            continue
                        
                        op_tipo_lower = op_tipo.lower().strip()
                        area_lower = area_voluntariado.strip().lower()
                        
                        # Verificar coincidencia parcial (más flexible)
                        # Comparar tanto si el filtro está contenido en el valor como si el valor está contenido en el filtro
                        if area_lower not in op_tipo_lower and op_tipo_lower not in area_lower:
                            print(f"  ❌ Oportunidad {op.id} descartada: tipo_de_voluntariado '{op_tipo}' no coincide con '{area_voluntariado}'")
                            print(f"     Comparación: '{op_tipo_lower}' vs '{area_lower}'")
                            continue
                        print(f"  ✓ Oportunidad {op.id} tipo OK: '{op_tipo}' coincide con '{area_voluntariado}'")
                    
                    oportunidades_filtradas.append(op)
                    print(f"  ✅ Oportunidad {op.id} cumple todos los filtros")
                
                oportunidades = oportunidades_filtradas
                print(f"✅ Oportunidades después de filtrado en Python: {len(oportunidades)}")
                
                # Debug: mostrar los valores de region_opor y tipo_de_voluntariado de las oportunidades encontradas
                if oportunidades:
                    print(f"📊 Valores de las oportunidades encontradas:")
                    for op in oportunidades[:5]:  # Mostrar solo las primeras 5
                        region_val = getattr(op, 'region_opor', None) or 'N/A'
                        tipo_val = getattr(op, 'tipo_de_voluntariado', None) or getattr(op, 'area_voluntariado', None) or 'N/A'
                        estado_val = getattr(op, 'estado', None) or 'N/A'
                        print(f"  - ID {op.id}: estado='{estado_val}', region_opor='{region_val}', tipo_de_voluntariado='{tipo_val}'")
                else:
                    print(f"⚠️ No se encontraron oportunidades con los filtros aplicados")
                    estado_filtro = estado if estado else 'activa'
                    print(f"🔍 Filtros aplicados: estado='{estado_filtro}', region='{region}', tipo_de_voluntariado='{area_voluntariado}'")
            except Exception as e:
                print(f"❌ Error al ejecutar query: {e}")
                import traceback
                print(traceback.format_exc())
                # Si hay error, intentar sin filtro de tipo de voluntariado
                if area_voluntariado:
                    print(f"🔄 Reintentando sin filtro de tipo de voluntariado...")
                    query = Oportunidad.query
                    if estado and estado != 'todas' and estado != 'all':
                        query = query.filter_by(estado=estado)
                    elif not estado:
                        query = query.filter_by(estado='activa')
                    if region:
                        query = query.filter(Oportunidad.region_opor.ilike(f'%{region}%'))
                    oportunidades = query.all()
                else:
                    raise
            print(f"✅ Oportunidades encontradas (sin filtro de organización): {len(oportunidades)}")
            
            # Si se aplicaron filtros y no hay resultados, NO mostrar todas las oportunidades
            # Solo mostrar todas si NO se aplicaron filtros
            filtros_aplicados = (region and region.strip()) or (area_voluntariado and area_voluntariado.strip()) or (estado and estado not in ['todas', 'all'])
            if len(oportunidades) == 0 and len(todas_oportunidades) > 0:
                if filtros_aplicados:
                    print(f"⚠️ La query con filtro no encontró resultados, pero hay {len(todas_oportunidades)} oportunidades en BD")
                    print(f"🔍 Filtros aplicados: region='{region}', area_voluntariado='{area_voluntariado}', estado='{estado}'")
                    print(f"❌ No se mostrarán todas las oportunidades porque se aplicaron filtros")
                    # Mantener lista vacía si se aplicaron filtros
                    oportunidades = []
                else:
                    print(f"⚠️ No se encontraron resultados sin filtros, mostrando todas las oportunidades")
                    oportunidades = todas_oportunidades
        
        resultado = []
        for op in oportunidades:
            try:
                organizacion = Organizacion.query.get(op.organizacion_id)
                # Contar postulaciones de forma segura (sin cargar el modelo completo para evitar errores de columnas)
                try:
                    num_postulaciones = db.session.query(func.count(Postulacion.id)).filter(Postulacion.oportunidad_id == op.id).scalar() or 0
                except Exception as count_error:
                    print(f"⚠️ Error al contar postulaciones para oportunidad {op.id}: {count_error}")
                    num_postulaciones = 0
                
                # Formatear fecha límite de forma segura
                fecha_limite_str = None
                if op.fecha_limite_postulacion:
                    try:
                        fecha_limite_str = op.fecha_limite_postulacion.strftime('%Y-%m-%d')
                    except Exception as date_error:
                        print(f"Error al formatear fecha_limite_postulacion: {date_error}")
                        fecha_limite_str = None
                
                # Formatear fechas del voluntariado de forma segura
                fecha_inicio_vol_str = None
                fecha_fin_vol_str = None
                if op.fecha_inicio_voluntariado:
                    try:
                        fecha_inicio_vol_str = op.fecha_inicio_voluntariado.strftime('%Y-%m-%d')
                    except Exception as date_error:
                        print(f"Error al formatear fecha_inicio_voluntariado: {date_error}")
                        fecha_inicio_vol_str = None
                if op.fecha_fin_voluntariado:
                    try:
                        fecha_fin_vol_str = op.fecha_fin_voluntariado.strftime('%Y-%m-%d')
                    except Exception as date_error:
                        print(f"Error al formatear fecha_fin_voluntariado: {date_error}")
                        fecha_fin_vol_str = None
                
                # Formatear created_at de forma segura
                created_at_str = None
                if op.created_at:
                    try:
                        created_at_str = op.created_at.strftime('%Y-%m-%d %H:%M:%S')
                    except Exception as date_error:
                        print(f"Error al formatear created_at: {date_error}")
                        created_at_str = None
                
                # Obtener tipo_de_voluntariado de forma segura usando getattr (intentar ambos nombres)
                tipo_voluntariado_val = getattr(op, 'tipo_de_voluntariado', None) or getattr(op, 'area_voluntariado', None) or ''
                area_voluntariado_val = tipo_voluntariado_val  # Mantener compatibilidad
                
                # Obtener ciudad de organizacion de forma segura usando getattr
                organizacion_ciudad_val = getattr(organizacion, 'ciudad', None) or '' if organizacion else ''
                
                resultado.append({
                    'id': op.id,
                    'titulo': op.titulo if op.titulo else '',
                    'descripcion': op.descripcion if op.descripcion else '',
                    'organizacion_id': op.organizacion_id,
                    'organizacion_nombre': organizacion.nombre if organizacion and organizacion.nombre else '',
                    'organizacion_region': organizacion.region if organizacion and organizacion.region else '',
                    'organizacion_ciudad': organizacion_ciudad_val,
                    'organizacion_comuna': organizacion.comuna if organizacion and organizacion.comuna else '',
                    'region_opor': op.region_opor if op.region_opor else '',
                    'ciudad_opor': op.ciudad_opor if op.ciudad_opor else '',
                    'comuna_opor': op.comuna_opor if op.comuna_opor else '',
                    'meta_postulantes': op.meta_postulantes,
                    'cupo_maximo': op.cupo_maximo,
                    'fecha_limite_postulacion': fecha_limite_str,
                    'fecha_inicio_voluntariado': fecha_inicio_vol_str,
                    'fecha_fin_voluntariado': fecha_fin_vol_str,
                    'horas_voluntariado': int(op.horas_voluntariado) if op.horas_voluntariado is not None else None,
                    'estado': op.estado if op.estado is not None else 'activa',
                    'num_postulaciones': num_postulaciones,
                    'created_at': created_at_str,
                    'responsable_nombre': op.responsable_nombre if op.responsable_nombre else '',
                    'responsable_apellido': op.responsable_apellido if op.responsable_apellido else '',
                    'responsable_email': op.responsable_email if op.responsable_email else '',
                    'responsable_email_institucional': op.responsable_email_institucional if op.responsable_email_institucional else '',
                    'responsable_telefono': op.responsable_telefono if op.responsable_telefono else '',
                    'area_voluntariado': area_voluntariado_val or '',
                    'tipo_de_voluntariado': tipo_voluntariado_val or ''
                })
            except Exception as op_error:
                print(f"Error al procesar oportunidad ID {op.id}: {op_error}")
                import traceback
                print(traceback.format_exc())
                # Continuar con la siguiente oportunidad en lugar de fallar completamente
                continue
        
        return jsonify({
            'success': True,
            'oportunidades': resultado
        }), 200
        
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"Error en listar_oportunidades: {str(e)}")
        print(error_trace)
        return jsonify({
            'success': False,
            'error': str(e),
            'details': error_trace if app.debug else None
        }), 500

# Obtener una oportunidad por ID
@app.route("/api/oportunidades/<int:oportunidad_id>", methods=["GET"])
def obtener_oportunidad(oportunidad_id):
    try:
        oportunidad = Oportunidad.query.get(oportunidad_id)
        
        if not oportunidad:
            return jsonify({
                'success': False,
                'error': 'Oportunidad no encontrada'
            }), 404
        
        organizacion = Organizacion.query.get(oportunidad.organizacion_id)
        num_postulaciones = Postulacion.query.filter_by(oportunidad_id=oportunidad.id).count()
        
        return jsonify({
            'success': True,
            'oportunidad': {
                'id': oportunidad.id,
                'titulo': oportunidad.titulo,
                'descripcion': oportunidad.descripcion,
                'organizacion_id': oportunidad.organizacion_id,
                'organizacion_nombre': organizacion.nombre if organizacion else '',
                'organizacion_region': organizacion.region if organizacion else '',
                'organizacion_comuna': organizacion.comuna if organizacion else '',
                'organizacion_email': organizacion.email_contacto if organizacion else '',
                'meta_postulantes': oportunidad.meta_postulantes,
                'cupo_maximo': oportunidad.cupo_maximo,
                'fecha_limite_postulacion': oportunidad.fecha_limite_postulacion.strftime('%Y-%m-%d') if oportunidad.fecha_limite_postulacion else None,
                'fecha_inicio_voluntariado': oportunidad.fecha_inicio_voluntariado.strftime('%Y-%m-%d') if oportunidad.fecha_inicio_voluntariado else None,
                'fecha_fin_voluntariado': oportunidad.fecha_fin_voluntariado.strftime('%Y-%m-%d') if oportunidad.fecha_fin_voluntariado else None,
                'horas_voluntariado': int(oportunidad.horas_voluntariado) if oportunidad.horas_voluntariado is not None else None,
                'estado': oportunidad.estado,
                'num_postulaciones': num_postulaciones,
                'created_at': oportunidad.created_at.strftime('%Y-%m-%d %H:%M:%S') if oportunidad.created_at else None,
                'region_opor': oportunidad.region_opor if oportunidad.region_opor else '',
                'ciudad_opor': oportunidad.ciudad_opor if oportunidad.ciudad_opor else '',
                'comuna_opor': oportunidad.comuna_opor if oportunidad.comuna_opor else '',
                'tipo_de_voluntariado': getattr(oportunidad, 'tipo_de_voluntariado', None) or getattr(oportunidad, 'area_voluntariado', None) or '',
                'area_voluntariado': getattr(oportunidad, 'tipo_de_voluntariado', None) or getattr(oportunidad, 'area_voluntariado', None) or '',
                'responsable_nombre': oportunidad.responsable_nombre if getattr(oportunidad, 'responsable_nombre', None) else '',
                'responsable_apellido': oportunidad.responsable_apellido if getattr(oportunidad, 'responsable_apellido', None) else '',
                'responsable_email': oportunidad.responsable_email if getattr(oportunidad, 'responsable_email', None) else '',
                'responsable_email_institucional': oportunidad.responsable_email_institucional if getattr(oportunidad, 'responsable_email_institucional', None) else '',
                'responsable_telefono': oportunidad.responsable_telefono if getattr(oportunidad, 'responsable_telefono', None) else ''
            }
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Crear una nueva oportunidad
@app.route("/api/oportunidades", methods=["POST"])
def crear_oportunidad():
    try:
        data = request.json
        
        organizacion_id = data.get('organizacion_id')
        titulo = data.get('titulo')
        descripcion = data.get('descripcion')
        meta_postulantes = data.get('meta_postulantes')
        cupo_maximo = data.get('cupo_maximo')
        fecha_limite_str = data.get('fecha_limite_postulacion')
        fecha_inicio_vol_str = data.get('fecha_inicio_voluntariado')
        fecha_fin_vol_str = data.get('fecha_fin_voluntariado')
        horas_voluntariado_raw = data.get('horas_voluntariado')
        
        # Campos del responsable
        responsable_nombre = data.get('responsable_nombre')
        responsable_apellido = data.get('responsable_apellido')
        responsable_email = data.get('responsable_email')
        responsable_email_institucional = data.get('responsable_email_institucional')
        responsable_telefono = data.get('responsable_telefono')
        
        # Campos de ubicación
        region_opor = data.get('region_opor')
        ciudad_opor = data.get('ciudad_opor')
        comuna_opor = data.get('comuna_opor')
        
        # Campo de área de voluntariado
        area_voluntariado = data.get('area_voluntariado') or data.get('tipo_de_voluntariado')
        
        if not organizacion_id or not titulo or not descripcion:
            return jsonify({
                'success': False,
                'error': 'Organización, título y descripción son requeridos'
            }), 400
        
        # Validar longitud de descripción (máximo 500 caracteres)
        if len(descripcion) > 500:
            return jsonify({
                'success': False,
                'error': 'La descripción no puede exceder 500 caracteres'
            }), 400
        
        # Verificar que la organización existe
        organizacion = Organizacion.query.get(organizacion_id)
        if not organizacion:
            return jsonify({
                'success': False,
                'error': 'Organización no encontrada'
            }), 404
        
        # Convertir fechas
        fecha_limite = None
        if fecha_limite_str:
            try:
                fecha_limite = datetime.strptime(fecha_limite_str, '%Y-%m-%d').date()
            except ValueError:
                return jsonify({
                    'success': False,
                    'error': 'Formato de fecha límite inválido (debe ser YYYY-MM-DD)'
                }), 400
        
        fecha_inicio_vol = None
        if fecha_inicio_vol_str:
            try:
                fecha_inicio_vol = datetime.strptime(fecha_inicio_vol_str, '%Y-%m-%d').date()
            except ValueError:
                return jsonify({
                    'success': False,
                    'error': 'Formato de fecha de inicio inválido (debe ser YYYY-MM-DD)'
                }), 400
        
        fecha_fin_vol = None
        if fecha_fin_vol_str:
            try:
                fecha_fin_vol = datetime.strptime(fecha_fin_vol_str, '%Y-%m-%d').date()
            except ValueError:
                return jsonify({
                    'success': False,
                    'error': 'Formato de fecha de fin inválido (debe ser YYYY-MM-DD)'
                }), 400
        
        # Validar que fecha de inicio sea anterior a fecha de fin
        if fecha_inicio_vol and fecha_fin_vol and fecha_inicio_vol > fecha_fin_vol:
            return jsonify({
                'success': False,
                'error': 'La fecha de inicio debe ser anterior a la fecha de fin'
            }), 400

        # Validar horas del voluntariado (si viene)
        horas_voluntariado = None
        if horas_voluntariado_raw is not None and horas_voluntariado_raw != '':
            try:
                horas_voluntariado = int(horas_voluntariado_raw)
                if horas_voluntariado < 0:
                    return jsonify({'success': False, 'error': 'Las horas de voluntariado no pueden ser negativas'}), 400
            except (ValueError, TypeError):
                return jsonify({'success': False, 'error': 'Horas de voluntariado inválidas'}), 400
        
        # Validar tipos de datos
        try:
            organizacion_id = int(organizacion_id)
        except (ValueError, TypeError):
            return jsonify({
                'success': False,
                'error': 'ID de organización inválido'
            }), 400
        
        # Crear la oportunidad
        nueva_oportunidad = Oportunidad(
            organizacion_id=organizacion_id,
            titulo=titulo.strip(),
            descripcion=descripcion.strip(),
            meta_postulantes=int(meta_postulantes) if meta_postulantes else None,
            cupo_maximo=int(cupo_maximo) if cupo_maximo else None,
            fecha_limite_postulacion=fecha_limite,
            fecha_inicio_voluntariado=fecha_inicio_vol,
            fecha_fin_voluntariado=fecha_fin_vol,
            horas_voluntariado=horas_voluntariado,
            estado='activa',
            responsable_nombre=responsable_nombre.strip() if responsable_nombre else None,
            responsable_apellido=responsable_apellido.strip() if responsable_apellido else None,
            responsable_email=responsable_email.strip() if responsable_email else None,
            responsable_email_institucional=responsable_email_institucional.strip() if responsable_email_institucional else None,
            responsable_telefono=responsable_telefono.strip() if responsable_telefono else None,
            region_opor=region_opor.strip() if region_opor else None,
            ciudad_opor=ciudad_opor.strip() if ciudad_opor else None,
            comuna_opor=comuna_opor.strip() if comuna_opor else None,
            tipo_de_voluntariado=area_voluntariado.strip() if area_voluntariado else None,
            created_at=datetime.now()
        )
        
        try:
            db.session.add(nueva_oportunidad)
            db.session.flush()  # Para obtener el ID sin hacer commit
            
            # Verificar que se creó correctamente
            oportunidad_id = nueva_oportunidad.id
            print(f"Oportunidad creada con ID: {oportunidad_id}")
            
            db.session.commit()
            print(f"Commit exitoso para oportunidad ID: {oportunidad_id}")
            
            # Verificar que se guardó correctamente
            oportunidad_guardada = Oportunidad.query.get(oportunidad_id)
            if not oportunidad_guardada:
                raise Exception('La oportunidad no se guardó correctamente en la base de datos')
            
            return jsonify({
                'success': True,
                'message': 'Oportunidad creada exitosamente',
                'oportunidad': {
                    'id': nueva_oportunidad.id,
                    'titulo': nueva_oportunidad.titulo,
                    'estado': nueva_oportunidad.estado,
                    'organizacion_id': nueva_oportunidad.organizacion_id
                }
            }), 201
            
        except Exception as db_error:
            db.session.rollback()
            error_msg = str(db_error)
            print(f"ERROR al guardar en BD: {error_msg}")
            print(f"Datos recibidos: organizacion_id={organizacion_id} (tipo: {type(organizacion_id)}), titulo={titulo[:50]}, descripcion={descripcion[:50]}...")
            import traceback
            print(traceback.format_exc())
            return jsonify({
                'success': False,
                'error': f'Error al guardar en la base de datos: {error_msg}'
            }), 500
        
    except Exception as e:
        db.session.rollback()
        import traceback
        error_trace = traceback.format_exc()
        print(f"Error completo: {error_trace}")
        return jsonify({
            'success': False,
            'error': str(e),
            'details': error_trace if app.debug else None
        }), 500

# Actualizar estado de una oportunidad
@app.route("/api/oportunidades/<int:oportunidad_id>/estado", methods=["PUT"])
def actualizar_estado_oportunidad(oportunidad_id):
    try:
        data = request.json
        nuevo_estado = data.get('estado')
        
        if not nuevo_estado:
            return jsonify({
                'success': False,
                'error': 'El nuevo estado es requerido'
            }), 400
        
        estados_validos = ['activa', 'cerrada', 'abierta']
        
        if nuevo_estado not in estados_validos:
            return jsonify({
                'success': False,
                'error': f'Estado inválido. Debe ser uno de: {", ".join(estados_validos)}'
            }), 400
        
        oportunidad = Oportunidad.query.get(oportunidad_id)
        
        if not oportunidad:
            return jsonify({
                'success': False,
                'error': 'Oportunidad no encontrada'
            }), 404
        
        estado_anterior = oportunidad.estado
        oportunidad.estado = nuevo_estado
        db.session.commit()
        
        # Si se cerró la oportunidad, notificar a todos los postulantes
        if nuevo_estado == 'cerrada' and estado_anterior != 'cerrada':
            try:
                # Obtener todas las postulaciones de esta oportunidad
                postulaciones = Postulacion.query.filter_by(oportunidad_id=oportunidad_id).all()
                
                for postulacion in postulaciones:
                    # Solo notificar a postulantes que no hayan sido rechazados
                    if postulacion.estado not in ['No seleccionado']:
                        try:
                            usuario = Usuario.query.get(postulacion.usuario_id)
                            if usuario and usuario.email:
                                # Enviar email de notificación de cierre
                                enviar_email_cierre_oportunidad(usuario, oportunidad, postulacion)
                        except Exception as email_error:
                            print(f"Error enviando email de cierre a usuario {postulacion.usuario_id}: {str(email_error)}")
            except Exception as e:
                print(f"Error notificando cierre de oportunidad: {str(e)}")
                # No fallar la actualización si el envío de emails falla
        
        return jsonify({
            'success': True,
            'message': f'Estado actualizado a {nuevo_estado} exitosamente',
            'oportunidad': {
                'id': oportunidad.id,
                'estado': oportunidad.estado
            }
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Actualizar una oportunidad
@app.route("/api/oportunidades/<int:oportunidad_id>", methods=["PUT"])
def actualizar_oportunidad(oportunidad_id):
    try:
        oportunidad = Oportunidad.query.get(oportunidad_id)
        
        if not oportunidad:
            return jsonify({
                'success': False,
                'error': 'Oportunidad no encontrada'
            }), 404
        
        data = request.get_json()
        
        # Actualizar campos permitidos
        if 'titulo' in data:
            oportunidad.titulo = data['titulo'].strip() if data['titulo'] else oportunidad.titulo
        if 'descripcion' in data:
            oportunidad.descripcion = data['descripcion'].strip() if data['descripcion'] else oportunidad.descripcion
        if 'meta_postulantes' in data:
            oportunidad.meta_postulantes = int(data['meta_postulantes']) if data['meta_postulantes'] else None
        if 'cupo_maximo' in data:
            oportunidad.cupo_maximo = int(data['cupo_maximo']) if data['cupo_maximo'] else None
        if 'fecha_limite_postulacion' in data:
            if data['fecha_limite_postulacion']:
                try:
                    oportunidad.fecha_limite_postulacion = datetime.strptime(data['fecha_limite_postulacion'], '%Y-%m-%d').date()
                except ValueError:
                    pass
            else:
                oportunidad.fecha_limite_postulacion = None
        if 'fecha_inicio_voluntariado' in data:
            if data['fecha_inicio_voluntariado']:
                try:
                    oportunidad.fecha_inicio_voluntariado = datetime.strptime(data['fecha_inicio_voluntariado'], '%Y-%m-%d').date()
                except ValueError:
                    pass
            else:
                oportunidad.fecha_inicio_voluntariado = None
        if 'fecha_fin_voluntariado' in data:
            if data['fecha_fin_voluntariado']:
                try:
                    oportunidad.fecha_fin_voluntariado = datetime.strptime(data['fecha_fin_voluntariado'], '%Y-%m-%d').date()
                except ValueError:
                    pass
            else:
                oportunidad.fecha_fin_voluntariado = None
        if oportunidad.fecha_inicio_voluntariado and oportunidad.fecha_fin_voluntariado and oportunidad.fecha_inicio_voluntariado > oportunidad.fecha_fin_voluntariado:
            return jsonify({'success': False, 'error': 'La fecha de inicio debe ser anterior a la fecha de fin'}), 400
        if 'horas_voluntariado' in data:
            if data['horas_voluntariado'] is None or data['horas_voluntariado'] == '':
                oportunidad.horas_voluntariado = None
            else:
                try:
                    horas_int = int(data['horas_voluntariado'])
                    if horas_int < 0:
                        return jsonify({'success': False, 'error': 'Las horas de voluntariado no pueden ser negativas'}), 400
                    oportunidad.horas_voluntariado = horas_int
                except (ValueError, TypeError):
                    return jsonify({'success': False, 'error': 'Horas de voluntariado inválidas'}), 400
        if 'region_opor' in data:
            oportunidad.region_opor = data['region_opor'].strip() if data['region_opor'] else None
        if 'ciudad_opor' in data:
            oportunidad.ciudad_opor = data['ciudad_opor'].strip() if data['ciudad_opor'] else None
        if 'comuna_opor' in data:
            oportunidad.comuna_opor = data['comuna_opor'].strip() if data['comuna_opor'] else None
        if 'tipo_de_voluntariado' in data or 'area_voluntariado' in data:
            tipo_val = data.get('tipo_de_voluntariado') or data.get('area_voluntariado')
            if tipo_val:
                oportunidad.tipo_de_voluntariado = tipo_val.strip()
            else:
                oportunidad.tipo_de_voluntariado = None
        
        # Actualizar información del responsable
        if 'responsable_nombre' in data:
            oportunidad.responsable_nombre = data['responsable_nombre'].strip() if data['responsable_nombre'] else None
        if 'responsable_apellido' in data:
            oportunidad.responsable_apellido = data['responsable_apellido'].strip() if data['responsable_apellido'] else None
        if 'responsable_email' in data:
            oportunidad.responsable_email = data['responsable_email'].strip() if data['responsable_email'] else None
        if 'responsable_email_institucional' in data:
            oportunidad.responsable_email_institucional = data['responsable_email_institucional'].strip() if data['responsable_email_institucional'] else None
        if 'responsable_telefono' in data:
            oportunidad.responsable_telefono = data['responsable_telefono'].strip() if data['responsable_telefono'] else None
        
        db.session.commit()
        
        # Obtener datos actualizados para la respuesta
        organizacion = Organizacion.query.get(oportunidad.organizacion_id)
        num_postulaciones = Postulacion.query.filter_by(oportunidad_id=oportunidad.id).count()
        tipo_voluntariado_val = getattr(oportunidad, 'tipo_de_voluntariado', None) or ''
        
        return jsonify({
            'success': True,
            'message': 'Oportunidad actualizada exitosamente',
            'oportunidad': {
                'id': oportunidad.id,
                'titulo': oportunidad.titulo,
                'descripcion': oportunidad.descripcion,
                'organizacion_id': oportunidad.organizacion_id,
                'organizacion_nombre': organizacion.nombre if organizacion else '',
                'meta_postulantes': oportunidad.meta_postulantes,
                'cupo_maximo': oportunidad.cupo_maximo,
                'fecha_limite_postulacion': oportunidad.fecha_limite_postulacion.strftime('%Y-%m-%d') if oportunidad.fecha_limite_postulacion else None,
                'fecha_inicio_voluntariado': oportunidad.fecha_inicio_voluntariado.strftime('%Y-%m-%d') if oportunidad.fecha_inicio_voluntariado else None,
                'fecha_fin_voluntariado': oportunidad.fecha_fin_voluntariado.strftime('%Y-%m-%d') if oportunidad.fecha_fin_voluntariado else None,
                'horas_voluntariado': int(oportunidad.horas_voluntariado) if oportunidad.horas_voluntariado is not None else None,
                'estado': oportunidad.estado,
                'num_postulaciones': num_postulaciones,
                'region_opor': oportunidad.region_opor if oportunidad.region_opor else '',
                'ciudad_opor': oportunidad.ciudad_opor if oportunidad.ciudad_opor else '',
                'comuna_opor': oportunidad.comuna_opor if oportunidad.comuna_opor else '',
                'tipo_de_voluntariado': tipo_voluntariado_val,
                'area_voluntariado': tipo_voluntariado_val
            }
        }), 200
        
    except Exception as e:
        db.session.rollback()
        import traceback
        print(f"Error al actualizar oportunidad: {str(e)}")
        print(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Cerrar una oportunidad (mantener para compatibilidad)
@app.route("/api/oportunidades/<int:oportunidad_id>/cerrar", methods=["PUT"])
def cerrar_oportunidad(oportunidad_id):
    try:
        oportunidad = Oportunidad.query.get(oportunidad_id)
        
        if not oportunidad:
            return jsonify({
                'success': False,
                'error': 'Oportunidad no encontrada'
            }), 404
        
        oportunidad.estado = 'cerrada'
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Oportunidad cerrada exitosamente'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# ============================================
# ENDPOINTS PARA NOTICIAS
# ============================================

# Listar todas las noticias
@app.route("/api/noticias", methods=["GET"])
def listar_noticias():
    try:
        estado = request.args.get('estado')  # activa, inactiva, todas
        noticias = Noticia.query
        
        if estado and estado not in ['todas', 'all']:
            noticias = noticias.filter_by(estado=estado)
        elif not estado:
            # Por defecto, mostrar solo activas
            noticias = noticias.filter_by(estado='activa')
        
        noticias = noticias.order_by(Noticia.created_at.desc()).all()
        
        resultado = []
        for noticia in noticias:
            autor = Usuario.query.get(noticia.autor_id) if noticia.autor_id else None
            resultado.append({
                'id': noticia.id,
                'titulo': noticia.titulo,
                'contenido': noticia.contenido,
                'resumen': noticia.resumen or '',
                'autor_id': noticia.autor_id,
                'autor_nombre': f"{autor.nombre} {autor.apellido}".strip() if autor else 'Administrador',
                'estado': noticia.estado,
                'imagen_url': noticia.imagen_noticia if noticia.imagen_noticia else '',  # Solo el nombre del archivo (filename)
                'imagen_filename': noticia.imagen_noticia if noticia.imagen_noticia else '',  # Solo el nombre del archivo (filename)
                'fecha_publicacion': noticia.fecha_publicacion.strftime('%Y-%m-%d %H:%M:%S') if noticia.fecha_publicacion else None,
                'created_at': noticia.created_at.strftime('%Y-%m-%d %H:%M:%S') if noticia.created_at else None,
                'updated_at': noticia.updated_at.strftime('%Y-%m-%d %H:%M:%S') if noticia.updated_at else None
            })
        
        return jsonify({
            'success': True,
            'noticias': resultado
        }), 200
        
    except Exception as e:
        import traceback
        print(f"Error en listar_noticias: {str(e)}")
        print(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Obtener una noticia por ID
@app.route("/api/noticias/<int:noticia_id>", methods=["GET"])
def obtener_noticia(noticia_id):
    try:
        noticia = Noticia.query.get(noticia_id)
        
        if not noticia:
            return jsonify({
                'success': False,
                'error': 'Noticia no encontrada'
            }), 404
        
        autor = Usuario.query.get(noticia.autor_id) if noticia.autor_id else None
        
        return jsonify({
            'success': True,
            'noticia': {
                'id': noticia.id,
                'titulo': noticia.titulo,
                'contenido': noticia.contenido,
                'resumen': noticia.resumen or '',
                'autor_id': noticia.autor_id,
                'autor_nombre': f"{autor.nombre} {autor.apellido}".strip() if autor else 'Administrador',
                'estado': noticia.estado,
                'imagen_url': noticia.imagen_noticia if noticia.imagen_noticia else '',  # Solo el nombre del archivo (filename)
                'imagen_filename': noticia.imagen_noticia if noticia.imagen_noticia else '',  # Solo el nombre del archivo (filename)
                'fecha_publicacion': noticia.fecha_publicacion.strftime('%Y-%m-%d %H:%M:%S') if noticia.fecha_publicacion else None,
                'created_at': noticia.created_at.strftime('%Y-%m-%d %H:%M:%S') if noticia.created_at else None,
                'updated_at': noticia.updated_at.strftime('%Y-%m-%d %H:%M:%S') if noticia.updated_at else None
            }
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Subir imagen de noticia
@app.route("/api/noticias/upload-imagen", methods=["POST"])
def subir_imagen_noticia():
    try:
        # Verificar que se envió un archivo
        if 'imagen' not in request.files:
            return jsonify({
                'success': False,
                'error': 'No se proporcionó ningún archivo'
            }), 400
        
        file = request.files['imagen']
        
        if file.filename == '':
            return jsonify({
                'success': False,
                'error': 'No se seleccionó ningún archivo'
            }), 400
        
        # Validar tipo de archivo (solo imágenes)
        allowed_extensions = {'.jpg', '.jpeg', '.png', '.gif', '.webp'}
        file_ext = os.path.splitext(file.filename)[1].lower()
        if file_ext not in allowed_extensions:
            return jsonify({
                'success': False,
                'error': f'Tipo de archivo no permitido. Formatos permitidos: {", ".join(allowed_extensions)}'
            }), 400
        
        # Validar tamaño (máximo 5MB para imágenes)
        file.seek(0, os.SEEK_END)
        file_size = file.tell()
        file.seek(0)
        max_size = 5 * 1024 * 1024  # 5MB
        if file_size > max_size:
            return jsonify({
                'success': False,
                'error': 'El archivo es demasiado grande. Tamaño máximo: 5MB'
            }), 400
        
        # Crear directorio de imágenes de noticias si no existe
        imagenes_dir = os.path.join(os.getcwd(), 'imagenes_noticias')
        if not os.path.exists(imagenes_dir):
            os.makedirs(imagenes_dir)
        
        # Generar nombre único para el archivo
        from werkzeug.utils import secure_filename
        filename = f"noticia_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{os.urandom(4).hex()}{file_ext}"
        filename = secure_filename(filename)
        filepath = os.path.join(imagenes_dir, filename)
        
        # Guardar el archivo
        file.save(filepath)
        
        # Retornar solo el filename para almacenar en la base de datos
        # La ruta siempre será imagenes_noticias/ pero solo guardamos el filename
        
        return jsonify({
            'success': True,
            'message': 'Imagen subida exitosamente',
            'filename': filename,  # Solo el nombre del archivo
            'ruta': f"imagenes_noticias/{filename}"  # Ruta completa solo para referencia
        }), 200
        
    except Exception as e:
        import traceback
        print(f"Error al subir imagen de noticia: {str(e)}")
        print(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Endpoint para servir imágenes de noticias
@app.route("/api/noticias/imagen/<path:filename>", methods=["GET"])
def obtener_imagen_noticia(filename):
    try:
        filepath = os.path.join(os.getcwd(), 'imagenes_noticias', filename)
        if os.path.exists(filepath):
            return send_file(filepath, as_attachment=False)
        else:
            return jsonify({
                'success': False,
                'error': 'Imagen no encontrada'
            }), 404
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Crear una nueva noticia
@app.route("/api/noticias", methods=["POST"])
def crear_noticia():
    try:
        data = request.get_json()
        
        titulo = data.get('titulo')
        contenido = data.get('contenido')
        resumen = data.get('resumen', '')
        autor_id = data.get('autor_id')
        estado = data.get('estado', 'activa')
        # El campo imagen_url viene como el filename del archivo subido (NO es una URL)
        imagen_filename = data.get('imagen_url', '').strip() if data.get('imagen_url') else None
        fecha_publicacion_str = data.get('fecha_publicacion')
        
        if not titulo or not contenido:
            return jsonify({
                'success': False,
                'error': 'Título y contenido son requeridos'
            }), 400
        
        # Si viene con ruta, extraer solo el filename (por compatibilidad)
        if imagen_filename:
            if '/' in imagen_filename:
                imagen_filename = imagen_filename.split('/')[-1]
            elif '\\' in imagen_filename:
                imagen_filename = imagen_filename.split('\\')[-1]
        
        fecha_publicacion = None
        if fecha_publicacion_str:
            try:
                fecha_publicacion = datetime.strptime(fecha_publicacion_str, '%Y-%m-%d %H:%M:%S')
            except ValueError:
                try:
                    fecha_publicacion = datetime.strptime(fecha_publicacion_str, '%Y-%m-%d')
                except ValueError:
                    pass
        
        nueva_noticia = Noticia(
            titulo=titulo.strip(),
            contenido=contenido.strip(),
            resumen=resumen.strip() if resumen else None,
            autor_id=int(autor_id) if autor_id else None,
            estado=estado,
            imagen_noticia=imagen_filename,  # Guardar solo el filename
            fecha_publicacion=fecha_publicacion or datetime.now(),
            created_at=datetime.now(),
            updated_at=datetime.now()
        )
        
        db.session.add(nueva_noticia)
        db.session.commit()
        
        autor = Usuario.query.get(nueva_noticia.autor_id) if nueva_noticia.autor_id else None
        
        return jsonify({
            'success': True,
            'message': 'Noticia creada exitosamente',
            'noticia': {
                'id': nueva_noticia.id,
                'titulo': nueva_noticia.titulo,
                'contenido': nueva_noticia.contenido,
                'resumen': nueva_noticia.resumen or '',
                'autor_id': nueva_noticia.autor_id,
                'autor_nombre': f"{autor.nombre} {autor.apellido}".strip() if autor else 'Administrador',
                'estado': nueva_noticia.estado,
                'imagen_url': nueva_noticia.imagen_noticia if nueva_noticia.imagen_noticia else '',  # Solo el filename del archivo
                'imagen_filename': nueva_noticia.imagen_noticia if nueva_noticia.imagen_noticia else '',  # Solo el filename del archivo
                'fecha_publicacion': nueva_noticia.fecha_publicacion.strftime('%Y-%m-%d %H:%M:%S') if nueva_noticia.fecha_publicacion else None,
                'created_at': nueva_noticia.created_at.strftime('%Y-%m-%d %H:%M:%S') if nueva_noticia.created_at else None
            }
        }), 201
        
    except Exception as e:
        db.session.rollback()
        import traceback
        print(f"Error en crear_noticia: {str(e)}")
        print(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Actualizar una noticia
@app.route("/api/noticias/<int:noticia_id>", methods=["PUT"])
def actualizar_noticia(noticia_id):
    try:
        print(f"DEBUG: Iniciando actualización de noticia ID: {noticia_id}")
        noticia = Noticia.query.get(noticia_id)
        
        if not noticia:
            print(f"DEBUG: Noticia {noticia_id} no encontrada")
            return jsonify({
                'success': False,
                'error': 'Noticia no encontrada'
            }), 404
        
        data = request.get_json()
        print(f"DEBUG: Datos recibidos: {data}")
        
        # Actualizar campos permitidos
        if 'titulo' in data:
            noticia.titulo = data['titulo'].strip() if data['titulo'] else noticia.titulo
        if 'contenido' in data:
            noticia.contenido = data['contenido'].strip() if data['contenido'] else noticia.contenido
        if 'resumen' in data:
            noticia.resumen = data['resumen'].strip() if data['resumen'] else None
        if 'estado' in data:
            noticia.estado = data['estado']
        if 'imagen_url' in data:
            # imagen_url es el filename del archivo, NO una URL externa
            imagen_filename = data['imagen_url']
            print(f"DEBUG: Recibido filename de imagen: '{imagen_filename}' (tipo: {type(imagen_filename)})")
            
            # Si es string vacío, eliminar la imagen
            if imagen_filename == '':
                print("DEBUG: Eliminando imagen (string vacío)")
                # Eliminar el archivo físico si existe
                if noticia.imagen_noticia:
                    try:
                        import os
                        filepath = os.path.join(os.getcwd(), 'imagenes_noticias', noticia.imagen_noticia)
                        if os.path.exists(filepath):
                            os.remove(filepath)
                            print(f"DEBUG: Archivo eliminado: {filepath}")
                    except Exception as e:
                        print(f"Error al eliminar archivo de imagen: {str(e)}")
                noticia.imagen_noticia = None
            elif imagen_filename:
                # Limpiar el filename - extraer solo el nombre si viene con ruta
                filename_clean = imagen_filename.strip()
                if '/' in filename_clean:
                    filename_clean = filename_clean.split('/')[-1]
                elif '\\' in filename_clean:
                    filename_clean = filename_clean.split('\\')[-1]
                
                if filename_clean:
                    noticia.imagen_noticia = filename_clean
                    print(f"DEBUG: Filename guardado: '{filename_clean}'")
                else:
                    print("DEBUG: Filename vacío después de limpiar")
                    noticia.imagen_noticia = None
            else:
                print("DEBUG: imagen_filename es None o falsy")
                noticia.imagen_noticia = None
        else:
            print("DEBUG: imagen_url no está en data, manteniendo valor existente")
        
        print(f"DEBUG: noticia.imagen_noticia después de procesar: '{noticia.imagen_noticia}'")
        if 'fecha_publicacion' in data:
            if data['fecha_publicacion']:
                try:
                    noticia.fecha_publicacion = datetime.strptime(data['fecha_publicacion'], '%Y-%m-%d %H:%M:%S')
                except ValueError:
                    try:
                        noticia.fecha_publicacion = datetime.strptime(data['fecha_publicacion'], '%Y-%m-%d')
                    except ValueError:
                        pass
            else:
                noticia.fecha_publicacion = None
        
        noticia.updated_at = datetime.now()
        
        print(f"DEBUG: Antes del commit - imagen_noticia: '{noticia.imagen_noticia}'")
        try:
            db.session.commit()
            print("DEBUG: Commit exitoso")
        except Exception as commit_error:
            print(f"DEBUG: Error en commit: {str(commit_error)}")
            import traceback
            print(traceback.format_exc())
            db.session.rollback()
            raise commit_error
        
        autor = Usuario.query.get(noticia.autor_id) if noticia.autor_id else None
        
        return jsonify({
            'success': True,
            'message': 'Noticia actualizada exitosamente',
            'noticia': {
                'id': noticia.id,
                'titulo': noticia.titulo,
                'contenido': noticia.contenido,
                'resumen': noticia.resumen or '',
                'autor_id': noticia.autor_id,
                'autor_nombre': f"{autor.nombre} {autor.apellido}".strip() if autor else 'Administrador',
                'estado': noticia.estado,
                'imagen_url': noticia.imagen_noticia if noticia.imagen_noticia else '',  # Solo el nombre del archivo (filename)
                'imagen_filename': noticia.imagen_noticia if noticia.imagen_noticia else '',  # Solo el nombre del archivo (filename)
                'fecha_publicacion': noticia.fecha_publicacion.strftime('%Y-%m-%d %H:%M:%S') if noticia.fecha_publicacion else None,
                'created_at': noticia.created_at.strftime('%Y-%m-%d %H:%M:%S') if noticia.created_at else None,
                'updated_at': noticia.updated_at.strftime('%Y-%m-%d %H:%M:%S') if noticia.updated_at else None
            }
        }), 200
        
    except Exception as e:
        db.session.rollback()
        import traceback
        print(f"Error en actualizar_noticia: {str(e)}")
        print(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Eliminar una noticia
@app.route("/api/noticias/<int:noticia_id>", methods=["DELETE"])
def eliminar_noticia(noticia_id):
    try:
        noticia = Noticia.query.get(noticia_id)
        
        if not noticia:
            return jsonify({
                'success': False,
                'error': 'Noticia no encontrada'
            }), 404
        
        db.session.delete(noticia)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Noticia eliminada exitosamente'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        import traceback
        print(f"Error en eliminar_noticia: {str(e)}")
        print(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ============================================
# ENDPOINTS BIBLIOTECA (repositorio de documentos)
# ============================================

# Directorio donde se guardan los archivos subidos de la biblioteca
BIBLIOTECA_UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'documentos_biblioteca')

@app.route("/api/biblioteca/tematicas", methods=["GET"])
def listar_tematicas_biblioteca():
    try:
        tematicas = BibliotecaTematica.query.order_by(BibliotecaTematica.nombre).all()
        return jsonify({
            'success': True,
            'tematicas': [{'id': t.id, 'nombre': t.nombre} for t in tematicas]
        }), 200
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route("/api/biblioteca/documentos", methods=["GET"])
def listar_documentos_biblioteca():
    try:
        organizacion_id = request.args.get('organizacion_id', type=int)
        tematica_id = request.args.get('tematica_id', type=int)
        fecha_desde = request.args.get('fecha_desde')  # YYYY-MM-DD
        fecha_hasta = request.args.get('fecha_hasta')
        busqueda = request.args.get('q', '').strip()

        query = BibliotecaDocumento.query
        if organizacion_id:
            query = query.filter_by(organizacion_id=organizacion_id)
        if tematica_id:
            query = query.filter_by(tematica_id=tematica_id)
        if fecha_desde:
            try:
                fd = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
                query = query.filter(db.func.date(BibliotecaDocumento.fecha_edicion) >= fd)
            except ValueError:
                pass
        if fecha_hasta:
            try:
                fh = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
                query = query.filter(db.func.date(BibliotecaDocumento.fecha_edicion) <= fh)
            except ValueError:
                pass
        if busqueda:
            term = f'%{busqueda}%'
            query = query.filter(
                db.or_(
                    BibliotecaDocumento.nombre_archivo.ilike(term),
                    BibliotecaDocumento.descripcion.ilike(term),
                    BibliotecaDocumento.autor.ilike(term)
                )
            )
        documentos = query.order_by(BibliotecaDocumento.fecha_edicion.desc().nulls_last(), BibliotecaDocumento.created_at.desc()).all()

        resultado = []
        for doc in documentos:
            tematica = BibliotecaTematica.query.get(doc.tematica_id) if doc.tematica_id else None
            org = Organizacion.query.get(doc.organizacion_id) if doc.organizacion_id else None
            resultado.append({
                'id': doc.id,
                'nombre_archivo': doc.nombre_archivo,
                'archivo_filename': doc.archivo_filename,
                'archivo_mime': doc.archivo_mime,
                'autor': doc.autor or '',
                'fecha_edicion': doc.fecha_edicion.strftime('%Y-%m-%d %H:%M') if doc.fecha_edicion else None,
                'descripcion': doc.descripcion or '',
                'tematica_id': doc.tematica_id,
                'tematica_nombre': tematica.nombre if tematica else None,
                'organizacion_id': doc.organizacion_id,
                'organizacion_nombre': org.nombre if org else None,
                'created_at': doc.created_at.strftime('%Y-%m-%d %H:%M') if doc.created_at else None,
            })
        return jsonify({'success': True, 'documentos': resultado}), 200
    except ProgrammingError as e:
        err_msg = str(e)
        if 'organizacion_id' in err_msg and 'biblioteca_documentos' in err_msg:
            return jsonify({
                'success': False,
                'error': 'Falta la columna organizacion_id en biblioteca_documentos. Ejecuta la migración: python backend/run_migration_organizacion_id.py',
                'code': 'RUN_MIGRATION'
            }), 503
        raise
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route("/api/biblioteca/documentos/<int:doc_id>", methods=["GET"])
def obtener_documento_biblioteca(doc_id):
    try:
        doc = BibliotecaDocumento.query.get(doc_id)
        if not doc:
            return jsonify({'success': False, 'error': 'Documento no encontrado'}), 404
        tematica = BibliotecaTematica.query.get(doc.tematica_id) if doc.tematica_id else None
        org = Organizacion.query.get(doc.organizacion_id) if doc.organizacion_id else None
        return jsonify({
            'success': True,
            'documento': {
                'id': doc.id,
                'nombre_archivo': doc.nombre_archivo,
                'archivo_filename': doc.archivo_filename,
                'autor': doc.autor or '',
                'fecha_edicion': doc.fecha_edicion.strftime('%Y-%m-%d %H:%M') if doc.fecha_edicion else None,
                'descripcion': doc.descripcion or '',
                'tematica_nombre': tematica.nombre if tematica else None,
                'organizacion_nombre': org.nombre if org else None,
            }
        }), 200
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route("/api/biblioteca/documentos/<int:doc_id>/descargar", methods=["GET"])
def descargar_documento_biblioteca(doc_id):
    try:
        doc = BibliotecaDocumento.query.get(doc_id)
        if not doc:
            return jsonify({'success': False, 'error': 'Documento no encontrado'}), 404
        if not os.path.isdir(BIBLIOTECA_UPLOAD_FOLDER):
            return jsonify({'success': False, 'error': 'Carpeta de documentos no disponible'}), 404
        filepath = os.path.join(BIBLIOTECA_UPLOAD_FOLDER, doc.archivo_filename)
        if not os.path.isfile(filepath):
            return jsonify({'success': False, 'error': 'Archivo no encontrado en el servidor'}), 404
        return send_file(filepath, as_attachment=True, download_name=doc.nombre_archivo or doc.archivo_filename)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route("/api/biblioteca/documentos", methods=["POST"])
def subir_documento_biblioteca():
    try:
        if 'archivo' not in request.files:
            return jsonify({'success': False, 'error': 'No se envió ningún archivo'}), 400
        file = request.files['archivo']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'Nombre de archivo vacío'}), 400

        nombre_archivo = request.form.get('nombre_archivo') or file.filename
        autor = request.form.get('autor', '').strip()
        descripcion = request.form.get('descripcion', '').strip()
        tematica_id = request.form.get('tematica_id', type=int)
        organizacion_id = request.form.get('organizacion_id', type=int)
        fecha_edicion_str = request.form.get('fecha_edicion')

        fecha_edicion = None
        if fecha_edicion_str:
            try:
                fecha_edicion = datetime.strptime(fecha_edicion_str, '%Y-%m-%d')
            except ValueError:
                pass

        if not os.path.isdir(BIBLIOTECA_UPLOAD_FOLDER):
            os.makedirs(BIBLIOTECA_UPLOAD_FOLDER, exist_ok=True)

        from werkzeug.utils import secure_filename
        ext = os.path.splitext(file.filename)[1] or ''
        unique_name = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{secure_filename(file.filename)}"
        filepath = os.path.join(BIBLIOTECA_UPLOAD_FOLDER, unique_name)
        file.save(filepath)
        size_bytes = os.path.getsize(filepath)

        doc = BibliotecaDocumento(
            nombre_archivo=nombre_archivo,
            archivo_filename=unique_name,
            archivo_mime=file.content_type,
            archivo_tamano_bytes=size_bytes,
            autor=autor or None,
            fecha_edicion=fecha_edicion,
            descripcion=descripcion or None,
            tematica_id=tematica_id,
            organizacion_id=organizacion_id,
        )
        db.session.add(doc)
        db.session.commit()

        return jsonify({
            'success': True,
            'documento': {
                'id': doc.id,
                'nombre_archivo': doc.nombre_archivo,
                'autor': doc.autor,
                'fecha_edicion': doc.fecha_edicion.strftime('%Y-%m-%d %H:%M') if doc.fecha_edicion else None,
                'descripcion': doc.descripcion,
                'tematica_id': doc.tematica_id,
                'organizacion_id': doc.organizacion_id,
            }
        }), 201
    except Exception as e:
        db.session.rollback()
        import traceback
        print(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e)}), 500


# Eliminar una oportunidad (voluntariado)
@app.route("/api/oportunidades/<int:oportunidad_id>", methods=["DELETE"])
def eliminar_oportunidad(oportunidad_id):
    try:
        data = request.json or {}
        organizacion_id = data.get('organizacion_id')
        es_admin = data.get('es_admin', False)  # Permitir que admin elimine sin verificación
        
        # Buscar la oportunidad
        oportunidad = Oportunidad.query.get(oportunidad_id)
        
        if not oportunidad:
            return jsonify({
                'success': False,
                'error': 'Oportunidad no encontrada'
            }), 404
        
        # Si es admin, permitir eliminar sin verificación de organizacion_id
        # Si no es admin, verificar organizacion_id
        if not es_admin:
            # Verificar que se proporcionó el organizacion_id
            if not organizacion_id:
                return jsonify({
                'success': False,
                'error': 'ID de organización es requerido para eliminar la oportunidad'
            }), 400
        
        # Verificar que la organización es la propietaria de la oportunidad
        if oportunidad.organizacion_id != organizacion_id:
            return jsonify({
                'success': False,
                'error': 'No tienes permisos para eliminar esta oportunidad'
            }), 403
        
        # Obtener todas las postulaciones asociadas
        postulaciones = Postulacion.query.filter_by(oportunidad_id=oportunidad_id).all()
        num_postulaciones = len(postulaciones)
        
        # Eliminar todas las postulaciones asociadas
        for postulacion in postulaciones:
            db.session.delete(postulacion)
        
        # Eliminar la oportunidad
        db.session.delete(oportunidad)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Oportunidad eliminada exitosamente. Se eliminaron {num_postulaciones} postulación(es) asociada(s).',
            'oportunidad_eliminada': {
                'id': oportunidad_id,
                'titulo': oportunidad.titulo,
                'postulaciones_eliminadas': num_postulaciones
            }
        }), 200
        
    except Exception as e:
        db.session.rollback()
        import traceback
        error_trace = traceback.format_exc()
        print(f"Error al eliminar oportunidad: {error_trace}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# ============================================
# FUNCIONES DE EMAIL
# ============================================

def enviar_email_cambio_estado(postulacion, nuevo_estado, motivo_no_seleccion=None, motivo_no_seleccion_otro=None):
    """
    Envía un email al usuario cuando cambia el estado de su postulación.
    """
    try:
        # Obtener información del usuario y la oportunidad
        usuario = Usuario.query.get(postulacion.usuario_id)
        oportunidad = Oportunidad.query.get(postulacion.oportunidad_id)
        
        if not usuario or not oportunidad:
            print(f"No se pudo obtener información del usuario u oportunidad para la postulación {postulacion.id}")
            return
        
        # Obtener información de la organización
        organizacion = Organizacion.query.get(oportunidad.organizacion_id)
        organizacion_nombre = organizacion.nombre if organizacion else 'la organización'
        
        # Verificar que el usuario tenga email
        if not usuario.email:
            print(f"Usuario {usuario.id} no tiene email configurado")
            return
        
        # Determinar el asunto y el color del header según el estado
        estado_info = {
            'Seleccionado': {
                'asunto': '🎉 ¡Felicidades! Has sido seleccionado',
                'color': '#10b981',  # Verde
                'icono': '🎉',
                'mensaje_principal': '¡Felicidades! Has sido seleccionado para participar en esta oportunidad de voluntariado.'
            },
            'No seleccionado': {
                'asunto': 'Actualización de tu postulación',
                'color': '#ef4444',  # Rojo
                'icono': 'ℹ️',
                'mensaje_principal': 'Lamentamos informarte que no fuiste seleccionado para esta oportunidad.'
            },
            'Pre-seleccionado': {
                'asunto': '¡Buenas noticias! Has sido pre-seleccionado',
                'color': '#3b82f6',  # Azul
                'icono': '✨',
                'mensaje_principal': '¡Buenas noticias! Has sido pre-seleccionado para esta oportunidad.'
            },
            'Etapa de entrevista': {
                'asunto': 'Siguiente paso: Etapa de entrevista',
                'color': '#8b5cf6',  # Morado
                'icono': '💼',
                'mensaje_principal': 'Has avanzado a la etapa de entrevista. La organización se pondrá en contacto contigo pronto.'
            },
            'En lista de espera': {
                'asunto': 'Actualización: Estás en lista de espera',
                'color': '#f59e0b',  # Naranja
                'icono': '⏳',
                'mensaje_principal': 'Tu postulación está en lista de espera. Te notificaremos si hay disponibilidad.'
            },
            'Pendiente de revisión': {
                'asunto': 'Actualización de tu postulación',
                'color': '#eab308',  # Amarillo
                'icono': '📋',
                'mensaje_principal': 'El estado de tu postulación ha sido actualizado a: Pendiente de revisión.'
            }
        }
        
        info = estado_info.get(nuevo_estado, {
            'asunto': 'Actualización de tu postulación',
            'color': '#6b7280',
            'icono': 'ℹ️',
            'mensaje_principal': f'El estado de tu postulación ha sido actualizado a: {nuevo_estado}.'
        })
        
        # Construir mensaje adicional según el estado
        mensaje_adicional = ''
        if nuevo_estado == 'No seleccionado' and motivo_no_seleccion:
            mensaje_adicional = f'<p><strong>Motivo:</strong> {motivo_no_seleccion}</p>'
            if motivo_no_seleccion_otro:
                mensaje_adicional += f'<p><strong>Detalle:</strong> {motivo_no_seleccion_otro}</p>'
        elif nuevo_estado == 'Seleccionado':
            # Generar enlace de confirmación
            from flask import request
            base_url = request.host_url.rstrip('/') if hasattr(request, 'host_url') else 'http://127.0.0.1:5000'
            confirmacion_url = f"{base_url}/api/postulaciones/{postulacion.id}/confirmar"
            mensaje_adicional = f'''
                <p>La organización se pondrá en contacto contigo próximamente con más detalles sobre tu participación.</p>
                <p><strong>Por favor, confirma tu participación haciendo clic en el siguiente enlace:</strong></p>
                <p style="text-align: center; margin: 20px 0;">
                    <a href="{confirmacion_url}" class="button" style="background-color: {info['color']}; color: white; padding: 12px 30px; text-decoration: none; border-radius: 5px; display: inline-block;">
                        Confirmar Participación
                    </a>
                </p>
                <p><small>O copia y pega este enlace en tu navegador: {confirmacion_url}</small></p>
            '''
        elif nuevo_estado == 'Etapa de entrevista':
            mensaje_adicional = '<p>Prepárate para la entrevista y mantén tu información de contacto actualizada.</p>'
        
        # Crear el mensaje
        msg = Message(
            subject=info['asunto'],
            recipients=[usuario.email],
            html=f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="UTF-8">
                <style>
                    body {{
                        font-family: Arial, sans-serif;
                        line-height: 1.6;
                        color: #333;
                    }}
                    .container {{
                        max-width: 600px;
                        margin: 0 auto;
                        padding: 20px;
                        background-color: #f9f9f9;
                    }}
                    .header {{
                        background-color: {info['color']};
                        color: white;
                        padding: 20px;
                        text-align: center;
                        border-radius: 5px 5px 0 0;
                    }}
                    .content {{
                        background-color: white;
                        padding: 30px;
                        border-radius: 0 0 5px 5px;
                    }}
                    .button {{
                        display: inline-block;
                        padding: 12px 30px;
                        background-color: {info['color']};
                        color: white;
                        text-decoration: none;
                        border-radius: 5px;
                        margin-top: 20px;
                    }}
                    .footer {{
                        text-align: center;
                        margin-top: 20px;
                        color: #666;
                        font-size: 12px;
                    }}
                    .estado-badge {{
                        display: inline-block;
                        padding: 8px 16px;
                        background-color: {info['color']}20;
                        color: {info['color']};
                        border-radius: 20px;
                        font-weight: 600;
                        margin: 10px 0;
                    }}
                </style>
            </head>
            <body>
                <div class="container">
                    <div class="header">
                        <h1>{info['icono']} {info['asunto'].replace('🎉 ', '').replace('✨ ', '').replace('💼 ', '').replace('⏳ ', '').replace('📋 ', '').replace('ℹ️ ', '')}</h1>
                    </div>
                    <div class="content">
                        <p>Hola <strong>{usuario.nombre or 'Usuario'}</strong>,</p>
                        
                        <p>{info['mensaje_principal']}</p>
                        
                        <div class="estado-badge">
                            Estado: {nuevo_estado}
                        </div>
                        
                        <h3>Detalles de tu postulación:</h3>
                        <ul>
                            <li><strong>Oportunidad:</strong> {oportunidad.titulo}</li>
                            <li><strong>Organización:</strong> {organizacion_nombre}</li>
                            <li><strong>Nuevo estado:</strong> {nuevo_estado}</li>
                        </ul>
                        
                        {mensaje_adicional}
                        
                        <p>Puedes revisar el estado de todas tus postulaciones en tu perfil de usuario.</p>
                        
                        <p>Saludos cordiales,<br>
                        <strong>Equipo INJUV</strong></p>
                    </div>
                    <div class="footer">
                        <p>Este es un email automático, por favor no respondas a este mensaje.</p>
                    </div>
                </div>
            </body>
            </html>
            """
        )
        
        # Enviar el email
        mail.send(msg)
        print(f"Email de cambio de estado enviado a {usuario.email} (Estado: {nuevo_estado})")
        
    except Exception as e:
        print(f"Error al enviar email de cambio de estado: {str(e)}")
        raise

def enviar_email_cierre_oportunidad(usuario, oportunidad, postulacion):
    """
    Envía un email al usuario cuando se cierra la oportunidad a la que está postulado.
    """
    try:
        # Obtener información de la organización
        organizacion = Organizacion.query.get(oportunidad.organizacion_id)
        organizacion_nombre = organizacion.nombre if organizacion else 'la organización'
        
        # Verificar que el usuario tenga email
        if not usuario.email:
            print(f"Usuario {usuario.id} no tiene email configurado")
            return
        
        # Determinar mensaje según el estado de la postulación
        if postulacion.estado == 'Seleccionado':
            mensaje_estado = "Como seleccionado en esta oportunidad, te informamos que la oportunidad ha sido cerrada."
            mensaje_adicional = "<p>Si tienes alguna pregunta sobre tu participación o certificados, por favor contacta a la organización.</p>"
        elif postulacion.estado in ['Pre-seleccionado', 'Etapa de entrevista', 'En lista de espera']:
            mensaje_estado = f"Tu postulación está en estado: <strong>{postulacion.estado}</strong>. Te informamos que la oportunidad ha sido cerrada."
            mensaje_adicional = "<p>La organización puede contactarte si hay nuevas oportunidades similares en el futuro.</p>"
        else:
            mensaje_estado = "Te informamos que la oportunidad a la que te postulaste ha sido cerrada."
            mensaje_adicional = ""
        
        # Crear el mensaje
        msg = Message(
            subject='Actualización: Oportunidad de voluntariado cerrada',
            recipients=[usuario.email],
            html=f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="UTF-8">
                <style>
                    body {{
                        font-family: Arial, sans-serif;
                        line-height: 1.6;
                        color: #333;
                    }}
                    .container {{
                        max-width: 600px;
                        margin: 0 auto;
                        padding: 20px;
                        background-color: #f9f9f9;
                    }}
                    .header {{
                        background-color: #6b7280;
                        color: white;
                        padding: 20px;
                        text-align: center;
                        border-radius: 5px 5px 0 0;
                    }}
                    .content {{
                        background-color: white;
                        padding: 30px;
                        border-radius: 0 0 5px 5px;
                    }}
                    .footer {{
                        text-align: center;
                        margin-top: 20px;
                        color: #666;
                        font-size: 12px;
                    }}
                </style>
            </head>
            <body>
                <div class="container">
                    <div class="header">
                        <h1>📢 Actualización de Oportunidad</h1>
                    </div>
                    <div class="content">
                        <p>Hola <strong>{usuario.nombre or 'Usuario'}</strong>,</p>
                        
                        <p>{mensaje_estado}</p>
                        
                        <h3>Detalles de la oportunidad:</h3>
                        <ul>
                            <li><strong>Oportunidad:</strong> {oportunidad.titulo}</li>
                            <li><strong>Organización:</strong> {organizacion_nombre}</li>
                            <li><strong>Estado de tu postulación:</strong> {postulacion.estado}</li>
                            <li><strong>Estado de la oportunidad:</strong> Cerrada</li>
                        </ul>
                        
                        {mensaje_adicional}
                        
                        <p>Puedes revisar el estado de todas tus postulaciones en tu perfil de usuario.</p>
                        
                        <p>Saludos cordiales,<br>
                        <strong>Equipo INJUV</strong></p>
                    </div>
                    <div class="footer">
                        <p>Este es un email automático, por favor no respondas a este mensaje.</p>
                    </div>
                </div>
            </body>
            </html>
            """
        )
        
        # Enviar el email
        mail.send(msg)
        print(f"Email de cierre de oportunidad enviado a {usuario.email} (Oportunidad: {oportunidad.titulo})")
        
    except Exception as e:
        print(f"Error al enviar email de cierre de oportunidad: {str(e)}")
        # No lanzar error para no interrumpir el flujo

def enviar_email_confirmacion_postulacion(usuario, oportunidad):
    """
    Envía un email de confirmación al usuario cuando realiza una postulación exitosa.
    """
    try:
        # Obtener información de la organización
        organizacion = Organizacion.query.get(oportunidad.organizacion_id)
        organizacion_nombre = organizacion.nombre if organizacion else 'la organización'
        
        # Verificar que el usuario tenga email
        if not usuario.email:
            print(f"Usuario {usuario.id} no tiene email configurado")
            return
        
        # Crear el mensaje
        msg = Message(
            subject='✅ Postulación Exitosa - INJUV',
            recipients=[usuario.email],
            html=f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="UTF-8">
                <style>
                    body {{
                        font-family: Arial, sans-serif;
                        line-height: 1.6;
                        color: #333;
                    }}
                    .container {{
                        max-width: 600px;
                        margin: 0 auto;
                        padding: 20px;
                        background-color: #f9f9f9;
                    }}
                    .header {{
                        background-color: #0052CC;
                        color: white;
                        padding: 20px;
                        text-align: center;
                        border-radius: 5px 5px 0 0;
                    }}
                    .content {{
                        background-color: white;
                        padding: 30px;
                        border-radius: 0 0 5px 5px;
                    }}
                    .button {{
                        display: inline-block;
                        padding: 12px 30px;
                        background-color: #0052CC;
                        color: white;
                        text-decoration: none;
                        border-radius: 5px;
                        margin-top: 20px;
                    }}
                    .footer {{
                        text-align: center;
                        margin-top: 20px;
                        color: #666;
                        font-size: 12px;
                    }}
                </style>
            </head>
            <body>
                <div class="container">
                    <div class="header">
                        <h1>🎉 ¡Postulación Exitosa!</h1>
                    </div>
                    <div class="content">
                        <p>Hola <strong>{usuario.nombre or 'Usuario'}</strong>,</p>
                        
                        <p>Te confirmamos que tu postulación ha sido recibida exitosamente.</p>
                        
                        <h3>Detalles de tu postulación:</h3>
                        <ul>
                            <li><strong>Oportunidad:</strong> {oportunidad.titulo}</li>
                            <li><strong>Organización:</strong> {organizacion_nombre}</li>
                            <li><strong>Estado:</strong> Pendiente de revisión</li>
                        </ul>
                        
                        <p>La organización revisará tu postulación y te notificará sobre el estado de tu aplicación.</p>
                        
                        <p>Puedes revisar el estado de tus postulaciones en tu perfil de usuario.</p>
                        
                        <p>¡Gracias por tu interés en participar como voluntario!</p>
                        
                        <p>Saludos cordiales,<br>
                        <strong>Equipo INJUV</strong></p>
                    </div>
                    <div class="footer">
                        <p>Este es un email automático, por favor no respondas a este mensaje.</p>
                    </div>
                </div>
            </body>
            </html>
            """
        )
        
        # Enviar el email
        mail.send(msg)
        print(f"Email de confirmación enviado a {usuario.email}")
        
    except Exception as e:
        print(f"Error al enviar email de confirmación: {str(e)}")
        raise

# ============================================
# ENDPOINTS PARA POSTULACIONES
# ============================================

# Crear una nueva postulación
@app.route("/api/postulaciones", methods=["POST"])
def crear_postulacion():
    try:
        data = request.json
        print(f"📥 Datos recibidos para postulación: {data}")
        
        usuario_id = data.get('usuario_id')
        oportunidad_id = data.get('oportunidad_id')
        
        print(f"🔍 usuario_id: {usuario_id} (tipo: {type(usuario_id)})")
        print(f"🔍 oportunidad_id: {oportunidad_id} (tipo: {type(oportunidad_id)})")
        
        if not usuario_id or not oportunidad_id:
            error_msg = f'Usuario y oportunidad son requeridos. Recibido: usuario_id={usuario_id}, oportunidad_id={oportunidad_id}'
            print(f"❌ {error_msg}")
            return jsonify({
                'success': False,
                'error': error_msg
            }), 400
        
        # Verificar que el usuario existe
        usuario = Usuario.query.get(usuario_id)
        print(f"👤 Usuario encontrado: {usuario is not None}")
        if not usuario:
            error_msg = f'Usuario no encontrado (ID: {usuario_id})'
            print(f"❌ {error_msg}")
            return jsonify({
                'success': False,
                'error': error_msg
            }), 404
        
        # Verificar que la oportunidad existe y está activa
        oportunidad = Oportunidad.query.get(oportunidad_id)
        print(f"🎯 Oportunidad encontrada: {oportunidad is not None}")
        if not oportunidad:
            error_msg = f'Oportunidad no encontrada (ID: {oportunidad_id})'
            print(f"❌ {error_msg}")
            return jsonify({
                'success': False,
                'error': error_msg
            }), 404
        
        print(f"📊 Estado de la oportunidad: '{oportunidad.estado}'")
        # Verificar que la oportunidad esté disponible (activa o abierta)
        if oportunidad.estado not in ['activa', 'abierta']:
            error_msg = f'La oportunidad no está disponible para postulaciones (estado: {oportunidad.estado})'
            print(f"❌ {error_msg}")
            return jsonify({
                'success': False,
                'error': error_msg
            }), 400
        
        # Verificar fecha límite
        fecha_limite = oportunidad.fecha_limite_postulacion
        fecha_actual = datetime.now().date()
        print(f"📅 Fecha límite: {fecha_limite}, Fecha actual: {fecha_actual}")
        if fecha_limite and fecha_limite < fecha_actual:
            error_msg = f'La fecha límite de postulación ha expirado (fecha límite: {fecha_limite})'
            print(f"❌ {error_msg}")
            return jsonify({
                'success': False,
                'error': error_msg
            }), 400
        
        # Verificar si ya postuló
        postulacion_existente = Postulacion.query.filter_by(
            usuario_id=usuario_id,
            oportunidad_id=oportunidad_id
        ).first()
        print(f"🔍 Postulación existente: {postulacion_existente is not None}")
        if postulacion_existente:
            error_msg = 'Ya has postulado a esta oportunidad'
            print(f"❌ {error_msg}")
            return jsonify({
                'success': False,
                'error': error_msg
            }), 400
        
        # Verificar si se alcanzó la meta de postulantes
        num_postulaciones = Postulacion.query.filter_by(oportunidad_id=oportunidad_id).count()
        meta_postulantes = oportunidad.meta_postulantes
        print(f"📈 Postulaciones actuales: {num_postulaciones}, Meta: {meta_postulantes}")
        if meta_postulantes and num_postulaciones >= meta_postulantes:
            # Cerrar automáticamente la oportunidad
            oportunidad.estado = 'cerrada'
            db.session.commit()
            error_msg = f'Se alcanzó el límite de postulaciones para esta oportunidad ({num_postulaciones}/{meta_postulantes})'
            print(f"❌ {error_msg}")
            return jsonify({
                'success': False,
                'error': error_msg
            }), 400
        
        print("✅ Todas las validaciones pasaron, creando postulación...")
        
        # Crear postulación
        nueva_postulacion = Postulacion(
            usuario_id=usuario_id,
            oportunidad_id=oportunidad_id,
            estado='Pendiente de revisión',
            estado_confirmacion='Pendiente',
            asistencia_capacitacion='No aplica',
            asistencia_actividad='No aplica',
            tiene_certificado=False,
            resena_org_publica=False,
            created_at=datetime.now(),
            updated_at=datetime.now()
        )
        
        db.session.add(nueva_postulacion)
        db.session.commit()
        
        # Enviar email de confirmación al usuario
        try:
            enviar_email_confirmacion_postulacion(usuario, oportunidad)
        except Exception as email_error:
            # No fallar la postulación si el email falla, solo loguear el error
            print(f"Error al enviar email de confirmación: {str(email_error)}")
        
        return jsonify({
            'success': True,
            'message': 'Postulación realizada exitosamente',
            'postulacion': {
                'id': nueva_postulacion.id,
                'estado': nueva_postulacion.estado,
                'oportunidad_titulo': oportunidad.titulo
            }
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Listar postulaciones de un usuario
@app.route("/api/usuarios/<int:usuario_id>/postulaciones", methods=["GET"])
def listar_postulaciones_usuario(usuario_id):
    try:
        print(f"🔍 listar_postulaciones_usuario llamado para usuario_id: {usuario_id}")
        # Usar SQL directo para evitar problemas con columnas que aún no existen
        from sqlalchemy import text
        
        # Asegurar que la sesión esté en un estado limpio (hacer rollback primero)
        try:
            db.session.rollback()
        except:
            pass
        
        # Usar SQL directo directamente para evitar problemas con columnas faltantes
        # (SQLAlchemy ORM intentará cargar todas las columnas del modelo, incluso si no existen)
        
        # Primero verificar si las columnas nuevas existen
        # Verificar con diferentes variaciones del nombre (con/sin tilde, con/sin mayúsculas)
        try:
            check_columns_query = text("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name = 'postulaciones' 
                AND (column_name = 'reseña_org' 
                     OR column_name = 'resena_org'
                     OR column_name = 'reseña_organizacion'
                     OR column_name = 'resena_organizacion'
                     OR column_name = 'resena_usuario_sobre_org'
                     OR column_name = 'calificacion_usuario_org')
            """)
            existing_columns = db.session.execute(check_columns_query).fetchall()
            existing_column_names = [row[0] for row in existing_columns]
        except Exception as check_error:
            # Si hay error (por ejemplo, transacción abortada), hacer rollback y reintentar
            db.session.rollback()
            try:
                existing_columns = db.session.execute(check_columns_query).fetchall()
                existing_column_names = [row[0] for row in existing_columns]
            except:
                # Si aún falla, continuar sin las columnas nuevas
                existing_column_names = []
        
        # Verificar variaciones del nombre de la columna de reseña
        has_resena_col = any(name in existing_column_names for name in [
            'reseña_org',
            'resena_org',
            'reseña_organizacion', 
            'resena_organizacion',
            'resena_usuario_sobre_org'
        ])
        # Obtener el nombre real de la columna si existe (priorizar reseña_org)
        resena_col_name = None
        for name in ['reseña_org', 'resena_org', 'reseña_organizacion', 'resena_organizacion', 'resena_usuario_sobre_org']:
            if name in existing_column_names:
                resena_col_name = name
                break
        
        has_calif_col = 'calificacion_usuario_org' in existing_column_names
        
        # Verificar si existe la columna resena_usuario_publica
        check_publica_query = text("""
            SELECT column_name 
            FROM information_schema.columns 
            WHERE table_name = 'postulaciones' 
            AND column_name = 'resena_usuario_publica'
        """)
        try:
            publica_col_exists = db.session.execute(check_publica_query).fetchone()
        except:
            publica_col_exists = None
        
        # Construir la consulta SQL base (incluir horas_voluntariado)
        base_query = """
            SELECT 
                p.id, p.oportunidad_id, p.estado, p.estado_confirmacion, 
                p.motivo_no_seleccion, p.tiene_certificado, p.ruta_certificado_pdf,
                p.resena_org_sobre_voluntario, p.resena_org_publica, p.calificacion_org,
                p.created_at, COALESCE(p.horas_voluntariado, 0) as horas_voluntariado
"""
        
        # Agregar columnas nuevas solo si existen, usando el nombre real
        if has_resena_col and resena_col_name:
            # Usar comillas dobles si el nombre tiene caracteres especiales
            if 'ñ' in resena_col_name or 'ó' in resena_col_name:
                base_query += f', p."{resena_col_name}"'
            else:
                base_query += f', p.{resena_col_name}'
        if has_calif_col:
            base_query += ", p.calificacion_usuario_org"
        
        if publica_col_exists:
            base_query += ", p.resena_usuario_publica"
        
        base_query += """
            FROM postulaciones p
            WHERE p.usuario_id = :usuario_id
        """
        
        # Ejecutar la consulta SQL
        try:
            print(f"📝 Ejecutando query SQL para usuario_id: {usuario_id}")
            print(f"📋 Query SQL completo:\n{base_query}")
            sql_query = text(base_query)
            results = db.session.execute(sql_query, {'usuario_id': usuario_id}).fetchall()
            print(f"✅ Query ejecutado. Filas encontradas: {len(results)}")
            if len(results) > 0:
                print(f"📊 Primera fila de resultados: {results[0]}")
        except Exception as sql_error:
            # Si hay error en la consulta SQL (por ejemplo, transacción abortada), hacer rollback y reintentar
            print(f"⚠️ Error en query SQL, reintentando: {sql_error}")
            error_str = str(sql_error).lower()
            db.session.rollback()
            # Reintentar después del rollback
            sql_query = text(base_query)
            results = db.session.execute(sql_query, {'usuario_id': usuario_id}).fetchall()
            print(f"✅ Reintento exitoso. Filas encontradas: {len(results)}")
        
        resultado = []
        print(f"🔄 Procesando {len(results)} filas de postulaciones")
        for row in results:
            post_id = row[0]
            oportunidad_id = row[1]
            
            # Usar SQL directo para obtener información de la oportunidad y organización
            # para evitar problemas con columnas faltantes en los modelos ORM
            try:
                oportunidad_query = text("""
                    SELECT id, titulo, estado, organizacion_id
                    FROM oportunidades
                    WHERE id = :oportunidad_id
                """)
                oportunidad_row = db.session.execute(oportunidad_query, {'oportunidad_id': oportunidad_id}).fetchone()
                
                if not oportunidad_row:
                    print(f"⚠️ Oportunidad {oportunidad_id} no encontrada, saltando postulación {post_id}")
                    continue
                
                oportunidad_titulo = oportunidad_row[1]
                oportunidad_estado = oportunidad_row[2]
                organizacion_id = oportunidad_row[3]
                oportunidad_cerrada = oportunidad_estado == 'cerrada'
                
                # Obtener información de la organización usando SQL directo
                organizacion_nombre = ''
                if organizacion_id:
                    try:
                        org_query = text("""
                            SELECT nombre
                            FROM organizaciones
                            WHERE id = :org_id
                        """)
                        org_row = db.session.execute(org_query, {'org_id': organizacion_id}).fetchone()
                        if org_row:
                            organizacion_nombre = org_row[0]
                    except:
                        pass
                
            except Exception as op_error:
                # Si hay error obteniendo la oportunidad, continuar con la siguiente postulación
                print(f"❌ Error obteniendo oportunidad {oportunidad_id} para postulación {post_id}: {op_error}")
                continue
            
            # Construir el diccionario de resultado
            post_data = {
                'id': post_id,
                'oportunidad_id': oportunidad_id,
                'oportunidad_titulo': oportunidad_titulo,
                'oportunidad_estado': oportunidad_estado,
                'oportunidad_cerrada': oportunidad_cerrada,
                'organizacion_id': organizacion_id,
                'organizacion_nombre': organizacion_nombre,
                'estado': row[2],
                'estado_confirmacion': row[3],
                'motivo_no_seleccion': row[4],
                'tiene_certificado': row[5],
                'ruta_certificado_pdf': row[6],
                'resena_org_sobre_voluntario': row[7],
                'resena_org_publica': row[8],
                'calificacion_org': float(row[9]) if row[9] is not None else None,
                'created_at': row[10].strftime('%Y-%m-%d %H:%M:%S') if row[10] else None,
                'horas_voluntariado': int(row[11]) if row[11] is not None else 0  # Horas de esta postulación específica
            }
            
            # Agregar campos nuevos solo si existen
            col_idx = 12
            if has_resena_col and resena_col_name:
                resena_valor = row[col_idx]
                post_data['reseña_org'] = resena_valor  # Nombre correcto de la columna
                post_data['resena_usuario_sobre_org'] = resena_valor  # Retrocompatibilidad
                post_data['reseña_organizacion'] = resena_valor  # Retrocompatibilidad
                col_idx += 1
            else:
                post_data['reseña_org'] = None
                post_data['resena_usuario_sobre_org'] = None
                post_data['reseña_organizacion'] = None
            
            if has_calif_col:
                post_data['calificacion_usuario_org'] = float(row[col_idx]) if row[col_idx] is not None else None
                col_idx += 1
            else:
                post_data['calificacion_usuario_org'] = None
            
            # Agregar campo resena_usuario_publica si existe
            if publica_col_exists and col_idx < len(row):
                post_data['resena_usuario_publica'] = bool(row[col_idx]) if row[col_idx] is not None else True
            else:
                post_data['resena_usuario_publica'] = True  # Por defecto es pública
            
            resultado.append(post_data)
        
        print(f"✅ Retornando {len(resultado)} postulaciones procesadas para usuario {usuario_id}")
        return jsonify({
            'success': True,
            'postulaciones': resultado
        }), 200
        
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"Error en listar_postulaciones_usuario: {str(e)}")
        print(error_trace)
        
        # Si el error es sobre columnas que no existen, dar un mensaje más claro
        error_str = str(e).lower()
        if 'no existe la columna' in error_str or ('column' in error_str and 'does not exist' in error_str):
            return jsonify({
                'success': False,
                'error': f'Error de base de datos: {str(e)}. Por favor verifica que la columna reseña_org existe en la tabla postulaciones.'
            }), 500
        
        return jsonify({
            'success': False,
            'error': f'Error al listar postulaciones: {str(e)}'
        }), 500

# Listar postulaciones de una oportunidad (para la organización)
@app.route("/api/oportunidades/<int:oportunidad_id>/postulaciones", methods=["GET"])
def listar_postulaciones_oportunidad(oportunidad_id):
    try:
        estado = request.args.get('estado')  # Filtro opcional por estado
        solo_seleccionados = request.args.get('solo_seleccionados', 'false').lower() == 'true'
        
        query = Postulacion.query.filter_by(oportunidad_id=oportunidad_id)
        
        if estado:
            query = query.filter_by(estado=estado)
        
        if solo_seleccionados:
            query = query.filter_by(estado='Seleccionado')
        
        postulaciones = query.all()
        
        resultado = []
        for post in postulaciones:
            usuario = Usuario.query.get(post.usuario_id)
            
            # Calcular edad si hay fecha de nacimiento
            edad = None
            if usuario and usuario.fecha_nacimiento:
                hoy = datetime.now().date()
                edad = hoy.year - usuario.fecha_nacimiento.year
                if (hoy.month, hoy.day) < (usuario.fecha_nacimiento.month, usuario.fecha_nacimiento.day):
                    edad -= 1
            
            # Obtener horas de voluntariado de esta postulación específica (no las horas totales del usuario)
            horas_voluntariado_postulacion = getattr(post, 'horas_voluntariado', None) or 0
            
            resultado.append({
                'id': post.id,
                'usuario_id': post.usuario_id,
                'usuario_nombre': usuario.nombre if usuario else '',
                'usuario_apellido': usuario.apellido if usuario else '',
                'usuario_nombre_completo': f"{usuario.nombre or ''} {usuario.apellido or ''}".strip() if usuario else '',
                'usuario_email': usuario.email if usuario else '',
                'usuario_telefono': usuario.telefono if usuario else '',
                'usuario_region': usuario.region if usuario else '',
                'usuario_comuna': usuario.comuna if usuario else '',
                'usuario_sexo': usuario.sexo if usuario else '',
                'usuario_rut': usuario.rut if usuario else '',
                'usuario_edad': edad,
                'estado': post.estado,
                'estado_confirmacion': post.estado_confirmacion,
                'asistencia_capacitacion': post.asistencia_capacitacion,
                'asistencia_actividad': post.asistencia_actividad,
                'tiene_certificado': post.tiene_certificado,
                'ruta_certificado_pdf': post.ruta_certificado_pdf if hasattr(post, 'ruta_certificado_pdf') else None,
                'horas_voluntariado': int(horas_voluntariado_postulacion) if horas_voluntariado_postulacion is not None else 0,  # Horas específicas de esta postulación
                'motivo_no_seleccion': post.motivo_no_seleccion,
                'motivo_no_seleccion_otro': post.motivo_no_seleccion_otro,
                'resena_org_sobre_voluntario': post.resena_org_sobre_voluntario,
                'resena_org_publica': post.resena_org_publica,
                'calificacion_org': float(getattr(post, 'calificacion_org', None)) if getattr(post, 'calificacion_org', None) is not None else None,
                'created_at': post.created_at.strftime('%Y-%m-%d %H:%M:%S') if post.created_at else None
            })
        
        return jsonify({
            'success': True,
            'postulaciones': resultado,
            'total': len(resultado)
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Actualizar estado de una postulación
@app.route("/api/postulaciones/<int:postulacion_id>/estado", methods=["PUT"])
def actualizar_estado_postulacion(postulacion_id):
    try:
        data = request.json
        
        nuevo_estado = data.get('estado')
        motivo_no_seleccion = data.get('motivo_no_seleccion')
        motivo_no_seleccion_otro = data.get('motivo_no_seleccion_otro')
        
        if not nuevo_estado:
            return jsonify({
                'success': False,
                'error': 'El nuevo estado es requerido'
            }), 400
        
        estados_validos = [
            'Pendiente de revisión',
            'No seleccionado',
            'Pre-seleccionado',
            'Etapa de entrevista',
            'En lista de espera',
            'Seleccionado'
        ]
        
        if nuevo_estado not in estados_validos:
            return jsonify({
                'success': False,
                'error': f'Estado inválido. Debe ser uno de: {", ".join(estados_validos)}'
            }), 400
        
        postulacion = Postulacion.query.get(postulacion_id)
        if not postulacion:
            return jsonify({
                'success': False,
                'error': 'Postulación no encontrada'
            }), 404
        
        postulacion.estado = nuevo_estado
        postulacion.updated_at = datetime.now()
        
        if nuevo_estado == 'No seleccionado' and motivo_no_seleccion:
            postulacion.motivo_no_seleccion = motivo_no_seleccion
            if motivo_no_seleccion_otro:
                postulacion.motivo_no_seleccion_otro = motivo_no_seleccion_otro
        
        db.session.commit()
        
        # Enviar email de notificación al usuario
        try:
            enviar_email_cambio_estado(postulacion, nuevo_estado, motivo_no_seleccion, motivo_no_seleccion_otro)
        except Exception as email_error:
            # No fallar la actualización si el email falla, solo loguear el error
            print(f"Error al enviar email de cambio de estado: {str(email_error)}")
        
        return jsonify({
            'success': True,
            'message': 'Estado actualizado exitosamente',
            'postulacion': {
                'id': postulacion.id,
                'estado': postulacion.estado
            }
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Declinar postulación (desde el usuario)
@app.route("/api/postulaciones/<int:postulacion_id>/declinar", methods=["PUT"])
def declinar_postulacion(postulacion_id):
    try:
        postulacion = Postulacion.query.get(postulacion_id)
        
        if not postulacion:
            return jsonify({
                'success': False,
                'error': 'Postulación no encontrada'
            }), 404
        
        postulacion.estado = 'Declinada por usuario'
        postulacion.updated_at = datetime.now()
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Postulación declinada exitosamente'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Actualizar confirmación de asistencia
@app.route("/api/postulaciones/<int:postulacion_id>/confirmacion", methods=["PUT"])
def actualizar_confirmacion(postulacion_id):
    try:
        data = request.json
        estado_confirmacion = data.get('estado_confirmacion')  # Pendiente, Confirmado, No confirmado
        
        if estado_confirmacion not in ['Pendiente', 'Confirmado', 'No confirmado']:
            return jsonify({
                'success': False,
                'error': 'Estado de confirmación inválido'
            }), 400
        
        postulacion = Postulacion.query.get(postulacion_id)
        if not postulacion:
            return jsonify({
                'success': False,
                'error': 'Postulación no encontrada'
            }), 404
        
        postulacion.estado_confirmacion = estado_confirmacion
        postulacion.updated_at = datetime.now()
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Confirmación actualizada exitosamente'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Confirmar postulación desde enlace de correo (GET para que funcione desde el enlace del email)
@app.route("/api/postulaciones/<int:postulacion_id>/confirmar", methods=["GET", "POST"])
def confirmar_postulacion(postulacion_id):
    try:
        postulacion = Postulacion.query.get(postulacion_id)
        if not postulacion:
            # Si es GET, retornar HTML de error
            if request.method == 'GET':
                return '''
                <!DOCTYPE html>
                <html>
                <head>
                    <meta charset="UTF-8">
                    <title>Error - Postulación no encontrada</title>
                    <style>
                        body { font-family: Arial, sans-serif; text-align: center; padding: 50px; }
                        .error { color: #ef4444; }
                    </style>
                </head>
                <body>
                    <h1 class="error">Error</h1>
                    <p>No se encontró la postulación.</p>
                </body>
                </html>
                ''', 404
            
            return jsonify({
                'success': False,
                'error': 'Postulación no encontrada'
            }), 404
        
        # Si es GET, mostrar página de confirmación
        if request.method == 'GET':
            usuario = Usuario.query.get(postulacion.usuario_id)
            oportunidad = Oportunidad.query.get(postulacion.oportunidad_id)
            
            return f'''
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="UTF-8">
                <title>Confirmar Participación</title>
                <style>
                    body {{ font-family: Arial, sans-serif; text-align: center; padding: 50px; background: #f3f4f6; }}
                    .container {{ max-width: 500px; margin: 0 auto; background: white; padding: 30px; border-radius: 10px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); }}
                    h1 {{ color: #10b981; }}
                    .button {{ display: inline-block; padding: 12px 30px; background: #10b981; color: white; text-decoration: none; border-radius: 5px; margin: 10px; }}
                    .button:hover {{ background: #059669; }}
                    .button-danger {{ background: #ef4444; }}
                    .button-danger:hover {{ background: #dc2626; }}
                </style>
            </head>
            <body>
                <div class="container">
                    <h1>Confirmar Participación</h1>
                    <p>Hola <strong>{usuario.nombre if usuario else 'Usuario'}</strong>,</p>
                    <p>¿Deseas confirmar tu participación en la oportunidad:</p>
                    <p><strong>{oportunidad.titulo if oportunidad else 'N/A'}</strong>?</p>
                    <div>
                        <a href="/api/postulaciones/{postulacion_id}/confirmar?accion=confirmar" class="button">✓ Confirmar</a>
                        <a href="/api/postulaciones/{postulacion_id}/confirmar?accion=rechazar" class="button button-danger">✗ No Confirmar</a>
                    </div>
                </div>
            </body>
            </html>
            '''
        
        # Si es POST o tiene parámetro accion, actualizar confirmación
        accion = request.args.get('accion') or (request.json.get('accion') if request.is_json else None)
        
        if accion == 'confirmar':
            postulacion.estado_confirmacion = 'Confirmado'
            mensaje = '¡Tu participación ha sido confirmada exitosamente!'
            color = '#10b981'
        elif accion == 'rechazar':
            postulacion.estado_confirmacion = 'No confirmado'
            mensaje = 'Has rechazado la participación en esta oportunidad.'
            color = '#ef4444'
        else:
            return jsonify({
                'success': False,
                'error': 'Acción inválida. Use "confirmar" o "rechazar"'
            }), 400
        
        postulacion.updated_at = datetime.now()
        db.session.commit()
        
        # Retornar HTML de confirmación
        return f'''
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>Confirmación Realizada</title>
            <style>
                body {{ font-family: Arial, sans-serif; text-align: center; padding: 50px; background: #f3f4f6; }}
                .container {{ max-width: 500px; margin: 0 auto; background: white; padding: 30px; border-radius: 10px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); }}
                h1 {{ color: {color}; }}
            </style>
        </head>
        <body>
            <div class="container">
                <h1>{mensaje}</h1>
                <p>Puedes cerrar esta ventana.</p>
            </div>
        </body>
        </html>
        ''', 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Actualizar asistencia (capacitación o actividad)
@app.route("/api/postulaciones/<int:postulacion_id>/asistencia", methods=["PUT"])
def actualizar_asistencia(postulacion_id):
    try:
        data = request.json or {}
        
        # Soporte para ambos formatos:
        # Formato nuevo: asistencia_capacitacion y asistencia_actividad directamente
        # Formato antiguo: tipo y asistencia
        asistencia_capacitacion = data.get('asistencia_capacitacion')
        asistencia_actividad = data.get('asistencia_actividad')
        horas_voluntariado = data.get('horas_voluntariado')  # Horas específicas de esta postulación
        tipo = data.get('tipo')  # 'capacitacion' o 'actividad' (formato antiguo)
        asistencia = data.get('asistencia')  # 'SI', 'No', 'No aplica' (formato antiguo)
        
        postulacion = Postulacion.query.get(postulacion_id)
        if not postulacion:
            return jsonify({
                'success': False,
                'error': 'Postulación no encontrada'
            }), 404
        
        # Validar valores válidos
        valores_validos = ['SI', 'No', 'No aplica', 'Sí', 'sí', 'si', 'no', 'NO', 'N/A', None]
        
        # Si se enviaron los campos directamente (formato nuevo)
        if asistencia_capacitacion is not None or asistencia_actividad is not None:
            if asistencia_capacitacion is not None:
                # Normalizar valores
                asistencia_cap = str(asistencia_capacitacion).strip()
                if asistencia_cap.upper() in ['SI', 'SÍ', 'YES', 'TRUE', '1']:
                    asistencia_cap = 'SI'
                elif asistencia_cap.upper() in ['NO', 'NOT', 'FALSE', '0']:
                    asistencia_cap = 'No'
                elif asistencia_cap.upper() in ['NO APLICA', 'N/A', 'NA', 'NONE']:
                    asistencia_cap = 'No aplica'
                
                if asistencia_cap not in ['SI', 'No', 'No aplica']:
                    return jsonify({
                        'success': False,
                        'error': f'Valor inválido para asistencia_capacitacion: {asistencia_capacitacion}. Debe ser "SI", "No" o "No aplica"'
                    }), 400
                
                postulacion.asistencia_capacitacion = asistencia_cap
            
            if asistencia_actividad is not None:
                # Normalizar valores
                asistencia_act = str(asistencia_actividad).strip()
                if asistencia_act.upper() in ['SI', 'SÍ', 'YES', 'TRUE', '1']:
                    asistencia_act = 'SI'
                elif asistencia_act.upper() in ['NO', 'NOT', 'FALSE', '0']:
                    asistencia_act = 'No'
                elif asistencia_act.upper() in ['NO APLICA', 'N/A', 'NA', 'NONE']:
                    asistencia_act = 'No aplica'
                
                if asistencia_act not in ['SI', 'No', 'No aplica']:
                    return jsonify({
                        'success': False,
                        'error': f'Valor inválido para asistencia_actividad: {asistencia_actividad}. Debe ser "SI", "No" o "No aplica"'
                    }), 400
                
                postulacion.asistencia_actividad = asistencia_act
        
        # Formato antiguo (compatibilidad hacia atrás)
        elif tipo and asistencia:
            if tipo not in ['capacitacion', 'actividad']:
                return jsonify({
                    'success': False,
                    'error': 'Tipo inválido. Debe ser "capacitacion" o "actividad"'
                }), 400
            
            # Normalizar asistencia
            asistencia_norm = str(asistencia).strip()
            if asistencia_norm.upper() in ['SI', 'SÍ', 'YES', 'TRUE', '1']:
                asistencia_norm = 'SI'
            elif asistencia_norm.upper() in ['NO', 'NOT', 'FALSE', '0']:
                asistencia_norm = 'No'
            elif asistencia_norm.upper() in ['NO APLICA', 'N/A', 'NA', 'NONE']:
                asistencia_norm = 'No aplica'
            
            if asistencia_norm not in ['SI', 'No', 'No aplica']:
                return jsonify({
                    'success': False,
                    'error': f'Valor de asistencia inválido: {asistencia}'
                }), 400
            
            if tipo == 'capacitacion':
                postulacion.asistencia_capacitacion = asistencia_norm
            else:
                postulacion.asistencia_actividad = asistencia_norm
        else:
            return jsonify({
                'success': False,
                'error': 'Se requiere enviar asistencia_capacitacion y/o asistencia_actividad, o tipo y asistencia'
            }), 400
        
        # Actualizar horas de voluntariado de la postulación si se proporciona
        if horas_voluntariado is not None:
            try:
                horas_int = int(horas_voluntariado)
                if horas_int < 0:
                    return jsonify({
                        'success': False,
                        'error': 'Las horas no pueden ser negativas'
                    }), 400
                postulacion.horas_voluntariado = horas_int
                
                # Calcular y actualizar el total de horas del usuario (suma de todas sus postulaciones)
                from sqlalchemy import func
                total_horas = db.session.query(func.coalesce(func.sum(Postulacion.horas_voluntariado), 0)).filter(
                    Postulacion.usuario_id == postulacion.usuario_id
                ).scalar() or 0
                
                # Actualizar el total en el usuario
                usuario = Usuario.query.get(postulacion.usuario_id)
                if usuario:
                    usuario.hora_voluntariado = int(total_horas)
                    print(f"✅ Horas totales del usuario {postulacion.usuario_id} actualizadas a: {total_horas} (suma de todas sus postulaciones)")
            except (ValueError, TypeError):
                return jsonify({
                    'success': False,
                    'error': 'Las horas deben ser un número entero válido'
                }), 400
        
        postulacion.updated_at = datetime.now()
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Asistencia actualizada exitosamente',
            'asistencia_capacitacion': postulacion.asistencia_capacitacion,
            'asistencia_actividad': postulacion.asistencia_actividad,
            'horas_voluntariado': postulacion.horas_voluntariado
        }), 200
        
    except Exception as e:
        db.session.rollback()
        import traceback
        print(f"Error en actualizar_asistencia: {str(e)}")
        print(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Generar certificado
@app.route("/api/postulaciones/<int:postulacion_id>/certificado", methods=["PUT"])
def generar_certificado(postulacion_id):
    try:
        data = request.json
        generar = data.get('generar', False)  # True para generar, False para no
        
        postulacion = Postulacion.query.get(postulacion_id)
        if not postulacion:
            return jsonify({
                'success': False,
                'error': 'Postulación no encontrada'
            }), 404
        
        if generar:
            # TODO: Generar PDF del certificado
            # ruta_certificado = generar_certificado_pdf(postulacion)
            # Por ahora, solo marcamos que tiene certificado
            postulacion.tiene_certificado = True
            postulacion.ruta_certificado_pdf = f'/certificados/certificado_{postulacion_id}.pdf'  # Placeholder
        else:
            postulacion.tiene_certificado = False
            postulacion.ruta_certificado_pdf = None
        
        postulacion.updated_at = datetime.now()
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Certificado actualizado exitosamente',
            'tiene_certificado': postulacion.tiene_certificado,
            'ruta_certificado': postulacion.ruta_certificado_pdf
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Agregar reseña de organización sobre voluntario
@app.route("/api/postulaciones/<int:postulacion_id>/resena", methods=["PUT"])
def agregar_resena(postulacion_id):
    try:
        data = request.json
        resena = data.get('resena') or data.get('resena_org_sobre_voluntario')  # Aceptar ambos nombres
        es_publica = data.get('es_publica', False)
        calificacion = data.get('calificacion') or data.get('calificacion_org')  # Aceptar ambos nombres
        
        # Reseña ya no es requerida (puede ser opcional)
        
        # Validar calificación si se proporciona
        if calificacion is not None:
            try:
                calificacion_float = float(calificacion)
                if calificacion_float < 0 or calificacion_float > 5:
                    return jsonify({
                        'success': False,
                        'error': 'La calificación debe estar entre 0 y 5'
                    }), 400
            except (ValueError, TypeError):
                return jsonify({
                    'success': False,
                    'error': 'La calificación debe ser un número válido'
                }), 400
        
        postulacion = Postulacion.query.get(postulacion_id)
        if not postulacion:
            return jsonify({
                'success': False,
                'error': 'Postulación no encontrada'
            }), 404
        
        if resena is not None:
            postulacion.resena_org_sobre_voluntario = resena
        if es_publica is not None:
            postulacion.resena_org_publica = es_publica
        # Solo actualizar calificacion_org si el atributo existe en el modelo
        if hasattr(postulacion, 'calificacion_org') and calificacion is not None:
            postulacion.calificacion_org = float(calificacion) if calificacion else None
        postulacion.updated_at = datetime.now()
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Reseña agregada exitosamente'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Endpoint para que el usuario deje una reseña sobre la organización
@app.route("/api/postulaciones/<int:postulacion_id>/resena-usuario", methods=["PUT"])
def agregar_resena_usuario(postulacion_id):
    try:
        # Importar text al inicio para que esté disponible en toda la función
        from sqlalchemy import text
        
        data = request.json
        resena = data.get('resena') or data.get('resena_usuario_sobre_org') or data.get('reseña_org') or data.get('reseña_organizacion')
        calificacion = data.get('calificacion') or data.get('calificacion_usuario_org')
        es_publica = data.get('es_publica', True)  # Por defecto es pública
        
        if not resena or not resena.strip():
            return jsonify({
                'success': False,
                'error': 'La reseña es requerida'
            }), 400
        
        # Validar calificación
        if calificacion is None:
            return jsonify({
                'success': False,
                'error': 'La calificación es requerida'
            }), 400
        
        try:
            calificacion_float = float(calificacion)
            if calificacion_float < 0 or calificacion_float > 5:
                return jsonify({
                    'success': False,
                    'error': 'La calificación debe estar entre 0 y 5'
                }), 400
        except (ValueError, TypeError):
            return jsonify({
                'success': False,
                'error': 'La calificación debe ser un número válido'
            }), 400
        
        # Verificar que la postulación existe usando SQL directo para evitar problemas con columnas faltantes
        db.session.rollback()  # Limpiar sesión primero
        check_post_query = text("""
            SELECT id FROM postulaciones WHERE id = :postulacion_id
        """)
        post_exists = db.session.execute(check_post_query, {'postulacion_id': postulacion_id}).fetchone()
        if not post_exists:
            return jsonify({
                'success': False,
                'error': 'Postulación no encontrada'
            }), 404
        
        # Verificar que el usuario está autenticado (opcional, depende de tu sistema de autenticación)
        # Aquí puedes agregar validación adicional si es necesario
        
        # Actualizar los campos usando la columna reseña_org (nombre correcto)
        try:
            # Verificar qué columnas existen realmente (priorizar reseña_org)
            db.session.rollback()  # Limpiar sesión antes de verificar
            check_col_query = text("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name = 'postulaciones' 
                AND (column_name = 'reseña_org' 
                     OR column_name = 'resena_org'
                     OR column_name = 'reseña_organizacion' 
                     OR column_name = 'resena_organizacion'
                     OR column_name = 'resena_usuario_sobre_org')
            """)
            existing_cols = db.session.execute(check_col_query).fetchall()
            existing_col_names = [row[0] for row in existing_cols]
            
            # Obtener el nombre real de la columna (priorizar reseña_org)
            resena_col_name = None
            for name in ['reseña_org', 'resena_org', 'reseña_organizacion', 'resena_organizacion', 'resena_usuario_sobre_org']:
                if name in existing_col_names:
                    resena_col_name = name
                    break
            
            # Si tenemos el nombre de la columna, usar SQL directo (más seguro que ORM)
            if resena_col_name:
                # Usar SQL directo con el nombre real de la columna
                if 'ñ' in resena_col_name or 'ó' in resena_col_name:
                    db.session.execute(
                        text(f'UPDATE postulaciones SET "{resena_col_name}" = :resena WHERE id = :id'),
                        {'resena': resena.strip(), 'id': postulacion_id}
                    )
                else:
                    db.session.execute(
                        text(f'UPDATE postulaciones SET {resena_col_name} = :resena WHERE id = :id'),
                        {'resena': resena.strip(), 'id': postulacion_id}
                    )
            else:
                # Si no existe ninguna columna, retornar error informativo
                return jsonify({
                    'success': False,
                    'error': 'La columna de reseña no existe en la base de datos. Por favor crea la columna reseña_org en la tabla postulaciones usando: ALTER TABLE postulaciones ADD COLUMN "reseña_org" TEXT;'
                }), 500
            
            # Verificar si existe la columna de calificación
            check_calif_query = text("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name = 'postulaciones' 
                AND column_name = 'calificacion_usuario_org'
            """)
            calif_col_exists = db.session.execute(check_calif_query).fetchone()
            
            if calif_col_exists:
                # Usar SQL directo para actualizar la calificación
                db.session.execute(
                    text("UPDATE postulaciones SET calificacion_usuario_org = :calificacion WHERE id = :id"),
                    {'calificacion': calificacion_float, 'id': postulacion_id}
                )
            # Si no existe la columna, simplemente no la actualizamos (no es crítico)
            
            # Verificar si existe la columna de resena_usuario_publica
            check_publica_query = text("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name = 'postulaciones' 
                AND column_name = 'resena_usuario_publica'
            """)
            publica_col_exists = db.session.execute(check_publica_query).fetchone()
            
            if publica_col_exists:
                # Actualizar el campo es_publica
                db.session.execute(
                    text("UPDATE postulaciones SET resena_usuario_publica = :es_publica WHERE id = :id"),
                    {'es_publica': bool(es_publica), 'id': postulacion_id}
                )
            
            # Actualizar updated_at usando SQL directo
            db.session.execute(
                text("UPDATE postulaciones SET updated_at = :updated_at WHERE id = :id"),
                {'updated_at': datetime.now(), 'id': postulacion_id}
            )
            db.session.commit()
        except Exception as db_error:
            # Si hay un error porque las columnas no existen, informar al usuario
            error_str = str(db_error).lower()
            if 'no existe la columna' in error_str or ('column' in error_str and 'does not exist' in error_str):
                db.session.rollback()
                return jsonify({
                    'success': False,
                    'error': f'Error al guardar: {str(db_error)}'
                }), 500
            raise
        
        return jsonify({
            'success': True,
            'message': 'Reseña agregada exitosamente'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        import traceback
        print(f"Error al agregar reseña de usuario: {str(e)}")
        print(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Endpoint para cambiar la visibilidad de una reseña de usuario (pública/privada)
# Solo la organización puede cambiar esto
@app.route("/api/postulaciones/<int:postulacion_id>/resena-visibilidad", methods=["PUT"])
def cambiar_visibilidad_resena(postulacion_id):
    try:
        from sqlalchemy import text
        
        data = request.json or {}
        raw_pub = data.get('es_publica', True)

        def _normalizar_es_publica(val):
            if isinstance(val, bool):
                return val
            if isinstance(val, (int, float)):
                return val != 0
            if isinstance(val, str):
                s = val.strip().lower()
                if s in ('false', 'f', '0', 'no', 'n'):
                    return False
                if s in ('true', 't', '1', 'si', 'sí', 'yes', 'y'):
                    return True
            return bool(val)

        es_publica = _normalizar_es_publica(raw_pub)
        organizacion_id = data.get('organizacion_id')
        organizacion_id_int = None
        if organizacion_id is not None and organizacion_id != '':
            try:
                organizacion_id_int = int(organizacion_id)
            except (TypeError, ValueError):
                return jsonify({
                    'success': False,
                    'error': 'organizacion_id inválido'
                }), 400
        
        # Verificar que la postulación existe
        db.session.rollback()
        check_post_query = text("""
            SELECT p.id, p.oportunidad_id, o.organizacion_id
            FROM postulaciones p
            INNER JOIN oportunidades o ON p.oportunidad_id = o.id
            WHERE p.id = :postulacion_id
        """)
        post_info = db.session.execute(check_post_query, {'postulacion_id': postulacion_id}).fetchone()
        
        if not post_info:
            return jsonify({
                'success': False,
                'error': 'Postulación no encontrada'
            }), 404
        
        # Verificar que la organización que solicita el cambio es la dueña de la oportunidad
        if organizacion_id_int is not None and int(post_info[2]) != organizacion_id_int:
            return jsonify({
                'success': False,
                'error': 'No tienes permisos para cambiar la visibilidad de esta reseña'
            }), 403
        
        # Verificar si existe la columna resena_usuario_publica
        check_publica_query = text("""
            SELECT column_name 
            FROM information_schema.columns 
            WHERE table_name = 'postulaciones' 
            AND column_name = 'resena_usuario_publica'
        """)
        publica_col_exists = db.session.execute(check_publica_query).fetchone()
        
        if not publica_col_exists:
            return jsonify({
                'success': False,
                'error': 'La columna resena_usuario_publica no existe en la base de datos. Por favor ejecuta la migración correspondiente.'
            }), 500
        
        # Actualizar solo el campo de visibilidad
        db.session.execute(
            text("UPDATE postulaciones SET resena_usuario_publica = :es_publica, updated_at = :updated_at WHERE id = :id"),
            {'es_publica': es_publica, 'updated_at': datetime.now(), 'id': postulacion_id}
        )
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Visibilidad de la reseña actualizada exitosamente',
            'es_publica': es_publica
        }), 200
        
    except Exception as e:
        db.session.rollback()
        import traceback
        print(f"Error al cambiar visibilidad de reseña: {str(e)}")
        print(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Obtener todas las reseñas de usuarios sobre la organización
@app.route("/api/organizaciones/<int:organizacion_id>/reseñas", methods=["GET"])
def obtener_reseñas_organizacion(organizacion_id):
    try:
        from sqlalchemy import text
        
        def normalizar_bool_api(valor, default=True):
            if valor is None:
                return default
            if isinstance(valor, bool):
                return valor
            if isinstance(valor, (int, float)):
                return valor != 0
            if isinstance(valor, str):
                v = valor.strip().lower()
                if v in ['true', 't', '1', 'si', 'sí', 'yes', 'y']:
                    return True
                if v in ['false', 'f', '0', 'no', 'n']:
                    return False
            return bool(valor)
        
        # Verificar si se solicitan solo reseñas públicas (desde el perfil público)
        solo_publicas = request.args.get('solo_publicas', 'false').lower() == 'true'
        
        # Verificar que la organización existe
        organizacion = Organizacion.query.get(organizacion_id)
        if not organizacion:
            return jsonify({
                'success': False,
                'error': 'Organización no encontrada'
            }), 404
        
        # Limpiar sesión
        db.session.rollback()
        
        # Verificar si existen las columnas de reseña
        try:
            check_col_query = text("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name = 'postulaciones' 
                AND (column_name = 'reseña_org' 
                     OR column_name = 'resena_org'
                     OR column_name = 'reseña_organizacion'
                     OR column_name = 'resena_organizacion'
                     OR column_name = 'resena_usuario_sobre_org'
                     OR column_name = 'calificacion_usuario_org'
                     OR column_name = 'resena_usuario_publica')
            """)
            existing_cols = db.session.execute(check_col_query).fetchall()
            existing_col_names = [row[0] for row in existing_cols]
            
            resena_col_candidates = [
                'reseña_org',
                'resena_org',
                'reseña_organizacion',
                'resena_organizacion',
                'resena_usuario_sobre_org'
            ]
            resena_cols_present = [n for n in resena_col_candidates if n in existing_col_names]
            has_resena_col = len(resena_cols_present) > 0
            
            has_calif_col = 'calificacion_usuario_org' in existing_col_names
            has_publica_col = 'resena_usuario_publica' in existing_col_names
        except:
            has_resena_col = False
            has_calif_col = False
            has_publica_col = False
            resena_cols_present = []
        
        # Construir consulta SQL para obtener postulaciones con reseñas
        # Primero obtener todas las oportunidades de la organización
        oportunidades_query = text("""
            SELECT id, titulo FROM oportunidades 
            WHERE organizacion_id = :organizacion_id
        """)
        oportunidades = db.session.execute(oportunidades_query, {'organizacion_id': organizacion_id}).fetchall()
        
        if not oportunidades:
            return jsonify({
                'success': True,
                'reseñas_por_voluntariado': []
            }), 200
        
        oportunidad_ids = [op[0] for op in oportunidades]
        
        if not oportunidad_ids:
            return jsonify({
                'success': True,
                'reseñas_por_voluntariado': []
            }), 200
        
        if not has_resena_col and not has_calif_col:
            return jsonify({
                'success': True,
                'reseñas_por_voluntariado': []
            }), 200

        def quote_col(name: str) -> str:
            return f'p."{name}"' if ('ñ' in name or 'ó' in name) else f'p.{name}'

        if has_resena_col:
            null_trim_parts = [f"NULLIF(TRIM({quote_col(n)}), '')" for n in resena_cols_present]
            resena_sql = 'COALESCE(' + ', '.join(null_trim_parts) + ')'
        else:
            resena_sql = 'NULL'

        calif_sql = 'p.calificacion_usuario_org' if has_calif_col else 'NULL'
        public_sql = 'COALESCE(p.resena_usuario_publica, true)' if has_publica_col else 'true'

        text_cond = f'({resena_sql}) IS NOT NULL' if has_resena_col else 'FALSE'
        calif_cond = 'p.calificacion_usuario_org IS NOT NULL' if has_calif_col else 'FALSE'
        where_review = f'({text_cond} OR ({calif_cond}))'

        base_query = f"""
            SELECT 
                p.id, p.oportunidad_id, p.usuario_id,
                p.created_at,
                ({resena_sql}) AS texto_resena,
                {calif_sql} AS calificacion_usuario,
                {public_sql} AS es_publica_db
            FROM postulaciones p
            WHERE p.oportunidad_id IN ({','.join(map(str, oportunidad_ids))})
            AND {where_review}
            ORDER BY p.created_at DESC
        """
        
        # Ejecutar consulta
        sql_query = text(base_query)
        results = db.session.execute(sql_query).fetchall()
        
        # Organizar por oportunidad
        reseñas_por_voluntariado = {}

        for row in results:
            post_id = row[0]
            oportunidad_id = row[1]
            usuario_id = row[2]
            fecha_postulacion = row[3]
            resena_text = row[4] if len(row) > 4 else None
            calificacion = float(row[5]) if len(row) > 5 and row[5] is not None else None
            es_publica_val = row[6] if len(row) > 6 else None
            es_publica = normalizar_bool_api(es_publica_val, True)
            
            # Si se solicitan solo públicas y esta no lo es, saltarla
            if solo_publicas and not es_publica:
                continue
            
            # Obtener información de la oportunidad
            oportunidad_info = next((op for op in oportunidades if op[0] == oportunidad_id), None)
            if not oportunidad_info:
                continue
            
            oportunidad_titulo = oportunidad_info[1]
            
            # Obtener información del usuario
            usuario_query = text("""
                SELECT id, nombre, apellido 
                FROM usuarios 
                WHERE id = :usuario_id
            """)
            usuario_row = db.session.execute(usuario_query, {'usuario_id': usuario_id}).fetchone()
            
            usuario_nombre = 'Usuario'
            if usuario_row:
                nombre_parts = [usuario_row[1], usuario_row[2]] if usuario_row[1] and usuario_row[2] else []
                usuario_nombre = ' '.join(nombre_parts) if nombre_parts else 'Usuario'
            
            # Agregar a la estructura organizada
            if oportunidad_id not in reseñas_por_voluntariado:
                reseñas_por_voluntariado[oportunidad_id] = {
                    'oportunidad_id': oportunidad_id,
                    'oportunidad_titulo': oportunidad_titulo,
                    'reseñas': []
                }
            
            reseña_data = {
                'postulacion_id': post_id,
                'usuario_id': usuario_id,
                'usuario_nombre': usuario_nombre,
                'fecha_postulacion': fecha_postulacion.strftime('%Y-%m-%d %H:%M:%S') if fecha_postulacion else None,
                'reseña': resena_text,
                'calificacion': calificacion,
                'es_publica': es_publica
            }
            
            reseñas_por_voluntariado[oportunidad_id]['reseñas'].append(reseña_data)
        
        # Convertir a lista
        resultado = list(reseñas_por_voluntariado.values())
        
        # Ordenar por número de reseñas (descendente)
        resultado.sort(key=lambda x: len(x['reseñas']), reverse=True)
        
        return jsonify({
            'success': True,
            'reseñas_por_voluntariado': resultado
        }), 200
        
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"Error en obtener_reseñas_organizacion: {str(e)}")
        print(error_trace)
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': f'Error al obtener reseñas: {str(e)}'
        }), 500

# Obtener reseñas públicas recientes para la página principal
@app.route("/api/reseñas/publicas", methods=["GET"])
def obtener_reseñas_publicas():
    try:
        from sqlalchemy import text
        
        # Parámetros opcionales
        limite = request.args.get('limite', default=3, type=int)
        
        # Limpiar sesión
        db.session.rollback()
        
        # Verificar si existen las columnas de reseña
        try:
            check_col_query = text("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name = 'postulaciones' 
                AND (column_name = 'reseña_org' 
                     OR column_name = 'resena_org'
                     OR column_name = 'calificacion_usuario_org'
                     OR column_name = 'calificacion_org'
                     OR column_name = 'resena_usuario_publica')
            """)
            existing_cols = db.session.execute(check_col_query).fetchall()
            existing_col_names = [row[0] for row in existing_cols]
            
            has_resena_col = any(name in existing_col_names for name in [
                'reseña_org', 'resena_org'
            ])
            resena_col_name = None
            for name in ['reseña_org', 'resena_org']:
                if name in existing_col_names:
                    resena_col_name = name
                    break
            
            has_calif_col = 'calificacion_usuario_org' in existing_col_names or 'calificacion_org' in existing_col_names
            calif_col_name = 'calificacion_usuario_org' if 'calificacion_usuario_org' in existing_col_names else ('calificacion_org' if 'calificacion_org' in existing_col_names else None)
            has_publica_col = 'resena_usuario_publica' in existing_col_names
        except:
            has_resena_col = False
            has_calif_col = False
            has_publica_col = False
            resena_col_name = None
            calif_col_name = None
        
        if not has_resena_col:
            return jsonify({
                'success': True,
                'reseñas': []
            }), 200
        
        # Construir consulta SQL para obtener postulaciones con reseñas públicas
        base_query = """
            SELECT 
                p.id as postulacion_id,
                p.usuario_id,
                p.oportunidad_id,
                p.created_at,
        """
        
        if resena_col_name:
            if 'ñ' in resena_col_name or 'ó' in resena_col_name:
                base_query += f'p."{resena_col_name}" as reseña,'
            else:
                base_query += f'p.{resena_col_name} as reseña,'
        else:
            base_query += 'NULL as reseña,'
        
        if calif_col_name:
            base_query += f'p.{calif_col_name} as calificacion,'
        else:
            base_query += 'NULL as calificacion,'
        
        if has_publica_col:
            base_query += 'COALESCE(p.resena_usuario_publica, true) as es_publica,'
        else:
            base_query += 'true as es_publica,'
        
        base_query += """
                u.nombre as usuario_nombre,
                u.apellido as usuario_apellido,
                o.id as organizacion_id,
                o.nombre as organizacion_nombre,
                op.titulo as oportunidad_titulo
            FROM postulaciones p
            INNER JOIN usuarios u ON p.usuario_id = u.id
            INNER JOIN oportunidades op ON p.oportunidad_id = op.id
            INNER JOIN organizaciones o ON op.organizacion_id = o.id
            WHERE 
        """
        
        if resena_col_name:
            if 'ñ' in resena_col_name or 'ó' in resena_col_name:
                base_query += f'p."{resena_col_name}" IS NOT NULL AND p."{resena_col_name}" != \'\''
            else:
                base_query += f'p.{resena_col_name} IS NOT NULL AND p.{resena_col_name} != \'\''
        
        if has_publica_col:
            base_query += ' AND COALESCE(p.resena_usuario_publica, true) = true'
        
        base_query += """
            ORDER BY p.created_at DESC
            LIMIT :limite
        """
        
        # Ejecutar consulta
        sql_query = text(base_query)
        results = db.session.execute(sql_query, {'limite': limite}).fetchall()
        
        # Procesar resultados
        reseñas = []
        for row in results:
            postulacion_id = row[0]
            usuario_id = row[1]
            oportunidad_id = row[2]
            fecha_creacion = row[3]
            resena_text = row[4]
            calificacion = float(row[5]) if row[5] is not None else None
            es_publica = row[6] if len(row) > 6 else True
            usuario_nombre = row[7] or ''
            usuario_apellido = row[8] or ''
            organizacion_id = row[9]
            organizacion_nombre = row[10] or 'Organización'
            oportunidad_titulo = row[11] or 'Oportunidad'
            
            nombre_completo = f"{usuario_nombre} {usuario_apellido}".strip() or 'Usuario'
            
            reseña_data = {
                'postulacion_id': postulacion_id,
                'usuario_id': usuario_id,
                'usuario_nombre': nombre_completo,
                'organizacion_id': organizacion_id,
                'organizacion_nombre': organizacion_nombre,
                'oportunidad_titulo': oportunidad_titulo,
                'reseña': resena_text,
                'calificacion': calificacion,
                'fecha': fecha_creacion.strftime('%Y-%m-%d %H:%M:%S') if fecha_creacion else None
            }
            
            reseñas.append(reseña_data)
        
        return jsonify({
            'success': True,
            'reseñas': reseñas
        }), 200
        
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"Error en obtener_reseñas_publicas: {str(e)}")
        print(error_trace)
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': f'Error al obtener reseñas públicas: {str(e)}'
        }), 500


# ----- Repositorio de documentos (Biblioteca y Academia) -----

def _documento_a_json(doc, seccion, include_organizacion_nombre=False):
    """Convierte BibliotecaDocumento o AcademiaDocumento a dict con sección."""
    d = {
        'id': doc.id,
        'nombre_archivo': doc.nombre_archivo,
        'archivo_filename': doc.archivo_filename,
        'archivo_mime': getattr(doc, 'archivo_mime', None) or None,
        'archivo_tamano_bytes': getattr(doc, 'archivo_tamano_bytes', None),
        'autor': doc.autor,
        'fecha_edicion': doc.fecha_edicion.isoformat() if doc.fecha_edicion else None,
        'descripcion': doc.descripcion,
        'created_at': doc.created_at.isoformat() if doc.created_at else None,
        'seccion': seccion,
    }
    if seccion == 'biblioteca' and hasattr(doc, 'tematica_id'):
        d['tematica_id'] = doc.tematica_id
    if seccion == 'academia':
        if hasattr(doc, 'organizacion_id'):
            d['organizacion_id'] = doc.organizacion_id
        if hasattr(doc, 'estado'):
            d['estado'] = doc.estado or 'pendiente'
        if hasattr(doc, 'categoria'):
            d['categoria'] = doc.categoria
        if include_organizacion_nombre and getattr(doc, 'organizacion_id', None):
            org = Organizacion.query.get(doc.organizacion_id)
            d['organizacion_nombre'] = org.nombre or org.siglas_nombre if org else None
    return d


@app.route("/api/repositorio/documentos", methods=["GET"])
def listar_repositorio_documentos():
    """Lista todos los documentos de Biblioteca y Academia (para panel admin)."""
    try:
        biblioteca = []
        academia = []
        if hasattr(db.session, 'query'):
            try:
                for doc in BibliotecaDocumento.query.order_by(BibliotecaDocumento.created_at.desc()).all():
                    biblioteca.append(_documento_a_json(doc, 'biblioteca'))
            except Exception as e:
                err_msg = str(e).lower()
                is_undefined_col = 'column' in err_msg or 'no existe' in err_msg or 'undefined' in err_msg
                if is_undefined_col and 'biblioteca_documentos' in err_msg:
                    db.session.rollback()
                    try:
                        # Añadir columnas que puede esperar el modelo (commit por columna para no deshacer las que sí se aplicaron)
                        for sql in [
                            "ALTER TABLE biblioteca_documentos ADD COLUMN IF NOT EXISTS archivo_tamano_bytes BIGINT",
                            "ALTER TABLE biblioteca_documentos ADD COLUMN IF NOT EXISTS descripcion VARCHAR(500)",
                            "ALTER TABLE biblioteca_documentos ADD COLUMN IF NOT EXISTS archivo_mime VARCHAR(120)",
                            "ALTER TABLE biblioteca_documentos ADD COLUMN IF NOT EXISTS autor VARCHAR(150)",
                            "ALTER TABLE biblioteca_documentos ADD COLUMN IF NOT EXISTS fecha_edicion TIMESTAMP",
                            "ALTER TABLE biblioteca_documentos ADD COLUMN IF NOT EXISTS tematica_id INTEGER",
                            "ALTER TABLE biblioteca_documentos ADD COLUMN IF NOT EXISTS created_at TIMESTAMP DEFAULT NOW()",
                            "ALTER TABLE biblioteca_documentos ADD COLUMN IF NOT EXISTS updated_at TIMESTAMP DEFAULT NOW()",
                        ]:
                            try:
                                db.session.execute(text(sql))
                                db.session.commit()
                            except Exception:
                                db.session.rollback()
                        for doc in BibliotecaDocumento.query.order_by(BibliotecaDocumento.created_at.desc()).all():
                            biblioteca.append(_documento_a_json(doc, 'biblioteca'))
                    except Exception as e2:
                        db.session.rollback()
                        return jsonify({
                            'success': False,
                            'error': f'Biblioteca: {str(e2)}. Ejecuta en tu BD: ALTER TABLE biblioteca_documentos ADD COLUMN descripcion VARCHAR(500);'
                        }), 500
                else:
                    import traceback
                    traceback.print_exc()
                    return jsonify({'success': False, 'error': f'Biblioteca: {str(e)}'}), 500
            try:
                for doc in AcademiaDocumento.query.order_by(AcademiaDocumento.created_at.desc()).all():
                    academia.append(_documento_a_json(doc, 'academia', include_organizacion_nombre=True))
            except Exception as e:
                err_msg = str(e).lower()
                # Si falla por columna 'estado' inexistente, intentar crearla y reintentar
                if 'estado' in err_msg and ('column' in err_msg or 'no existe' in err_msg or 'undefined' in err_msg):
                    db.session.rollback()
                    try:
                        db.session.execute(text(
                            "ALTER TABLE academia_documentos ADD COLUMN IF NOT EXISTS estado VARCHAR(20) NOT NULL DEFAULT 'pendiente'"
                        ))
                        db.session.execute(text(
                            "UPDATE academia_documentos SET estado = 'aprobado' WHERE estado = 'pendiente' OR estado IS NULL"
                        ))
                        db.session.commit()
                        for doc in AcademiaDocumento.query.order_by(AcademiaDocumento.created_at.desc()).all():
                            academia.append(_documento_a_json(doc, 'academia', include_organizacion_nombre=True))
                    except Exception as e2:
                        db.session.rollback()
                        return jsonify({
                            'success': False,
                            'error': f'Academia: {str(e2)}. Ejecuta en tu BD: ALTER TABLE academia_documentos ADD COLUMN estado VARCHAR(20) NOT NULL DEFAULT \'pendiente\';'
                        }), 500
                elif 'categoria' in err_msg and ('column' in err_msg or 'no existe' in err_msg or 'undefined' in err_msg):
                    db.session.rollback()
                    try:
                        db.session.execute(text(
                            "ALTER TABLE academia_documentos ADD COLUMN IF NOT EXISTS categoria VARCHAR(80)"
                        ))
                        db.session.commit()
                        for doc in AcademiaDocumento.query.order_by(AcademiaDocumento.created_at.desc()).all():
                            academia.append(_documento_a_json(doc, 'academia', include_organizacion_nombre=True))
                    except Exception as e2:
                        db.session.rollback()
                        return jsonify({
                            'success': False,
                            'error': f'Academia categoria: {str(e2)}. Ejecuta: ALTER TABLE academia_documentos ADD COLUMN IF NOT EXISTS categoria VARCHAR(80);'
                        }), 500
                else:
                    import traceback
                    traceback.print_exc()
                    return jsonify({'success': False, 'error': f'Academia: {str(e)}'}), 500
        return jsonify({
            'success': True,
            'biblioteca': biblioteca,
            'academia': academia,
        }), 200
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route("/api/biblioteca/documentos", methods=["GET"])
def listar_biblioteca_documentos():
    """Lista documentos de la Biblioteca (público o admin)."""
    try:
        docs = BibliotecaDocumento.query.order_by(BibliotecaDocumento.created_at.desc()).all()
        return jsonify({
            'success': True,
            'documentos': [_documento_a_json(d, 'biblioteca') for d in docs]
        }), 200
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route("/api/academia/documentos", methods=["GET"])
def listar_academia_documentos():
    """Lista documentos de la Academia públicos (solo aprobados), con nombre de organización."""
    try:
        docs = AcademiaDocumento.query.filter(
            AcademiaDocumento.estado == 'aprobado'
        ).order_by(AcademiaDocumento.created_at.desc()).all()
        return jsonify({
            'success': True,
            'documentos': [_documento_a_json(d, 'academia', include_organizacion_nombre=True) for d in docs]
        }), 200
    except Exception as e:
        err_msg = str(e).lower()
        if 'categoria' in err_msg and ('column' in err_msg or 'no existe' in err_msg or 'undefined' in err_msg):
            db.session.rollback()
            try:
                db.session.execute(text(
                    "ALTER TABLE academia_documentos ADD COLUMN IF NOT EXISTS categoria VARCHAR(80)"
                ))
                db.session.commit()
                docs = AcademiaDocumento.query.filter(
                    AcademiaDocumento.estado == 'aprobado'
                ).order_by(AcademiaDocumento.created_at.desc()).all()
                return jsonify({
                    'success': True,
                    'documentos': [_documento_a_json(d, 'academia', include_organizacion_nombre=True) for d in docs]
                }), 200
            except Exception as e2:
                db.session.rollback()
                return jsonify({
                    'success': False,
                    'error': 'Columna categoria no existe. Ejecuta en PostgreSQL: ALTER TABLE academia_documentos ADD COLUMN IF NOT EXISTS categoria VARCHAR(80);'
                }), 500
        return jsonify({'success': False, 'error': str(e)}), 500


def _allowed_document_extensions():
    return {'.pdf', '.doc', '.docx', '.xls', '.xlsx', '.ppt', '.pptx', '.txt', '.odt', '.ods', '.odp'}


def _allowed_academia_extensions():
    """Extensiones permitidas para Academia: documentos + videos."""
    return {'.pdf', '.doc', '.docx', '.xls', '.xlsx', '.ppt', '.pptx', '.txt', '.odt', '.ods', '.odp',
            '.mp4', '.avi', '.mov', '.wmv', '.flv', '.webm', '.mkv', '.m4v'}


def _subir_documento_repositorio(seccion, request):
    """Guarda archivo en disco y devuelve (filename, mime, size_bytes)."""
    from werkzeug.utils import secure_filename
    key = 'archivo'
    if key not in request.files:
        return None, 'No se proporcionó el archivo'
    file = request.files[key]
    if not file or file.filename == '':
        return None, 'No se seleccionó ningún archivo'
    ext = os.path.splitext(file.filename)[1].lower()
    # Academia permite documentos y videos; Biblioteca solo documentos
    allowed_exts = _allowed_academia_extensions() if seccion == 'academia' else _allowed_document_extensions()
    if ext not in allowed_exts:
        return None, f'Tipo no permitido. Permitidos: {", ".join(sorted(allowed_exts))}'
    file.seek(0, os.SEEK_END)
    size = file.tell()
    file.seek(0)
    # Límite de tamaño: Academia permite hasta 500MB (videos), Biblioteca 25MB
    max_size = 500 * 1024 * 1024 if seccion == 'academia' else 25 * 1024 * 1024
    max_size_mb = 500 if seccion == 'academia' else 25
    if size > max_size:
        return None, f'Tamaño máximo {max_size_mb}MB'
    folder = 'repositorio_biblioteca' if seccion == 'biblioteca' else 'repositorio_academia'
    base_dir = os.path.join(os.getcwd(), folder)
    if not os.path.exists(base_dir):
        os.makedirs(base_dir)
    safe_name = secure_filename(f"{seccion}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{os.urandom(4).hex()}{ext}")
    path = os.path.join(base_dir, safe_name)
    file.save(path)
    mime = file.content_type or None
    return (safe_name, mime, size), None


@app.route("/api/biblioteca/documentos", methods=["POST"])
def subir_biblioteca_documento():
    """Sube un documento a la Biblioteca (multipart: archivo, nombre_archivo?, descripcion?, autor?, tematica_id?)."""
    try:
        payload, err = _subir_documento_repositorio('biblioteca', request)
        if err:
            return jsonify({'success': False, 'error': err}), 400
        filename, mime, size = payload
        nombre_archivo = request.form.get('nombre_archivo', '').strip() or request.files.get('archivo').filename
        descripcion = (request.form.get('descripcion') or '')[:500]
        autor = (request.form.get('autor') or '')[:150]
        tematica_id = request.form.get('tematica_id', type=int) or None
        doc = BibliotecaDocumento(
            nombre_archivo=nombre_archivo[:255],
            archivo_filename=filename,
            archivo_mime=mime,
            archivo_tamano_bytes=size,
            autor=autor or None,
            fecha_edicion=datetime.utcnow(),
            descripcion=descripcion or None,
            tematica_id=tematica_id,
        )
        db.session.add(doc)
        db.session.commit()
        return jsonify({'success': True, 'documento': _documento_a_json(doc, 'biblioteca')}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route("/api/academia/documentos", methods=["POST"])
def subir_academia_documento():
    """Sube un documento a la Academia (multipart: archivo, nombre_archivo?, descripcion obligatoria, autor?)."""
    try:
        descripcion = (request.form.get('descripcion') or '').strip()[:500]
        if not descripcion:
            return jsonify({'success': False, 'error': 'La descripción breve es obligatoria para documentos de Academia.'}), 400
        payload, err = _subir_documento_repositorio('academia', request)
        if err:
            return jsonify({'success': False, 'error': err}), 400
        filename, mime, size = payload
        nombre_archivo = request.form.get('nombre_archivo', '').strip() or request.files.get('archivo').filename
        autor = (request.form.get('autor') or '')[:150]
        categoria = (request.form.get('categoria') or '').strip().lower()
        if categoria and categoria not in ACADEMIA_CATEGORIAS:
            categoria = None
        doc = AcademiaDocumento(
            nombre_archivo=nombre_archivo[:255],
            archivo_filename=filename,
            archivo_mime=mime,
            archivo_tamano_bytes=size,
            autor=autor or None,
            fecha_edicion=datetime.utcnow(),
            descripcion=descripcion,
            estado='aprobado',  # Admin sube directo como público
            categoria=categoria or None,
        )
        db.session.add(doc)
        db.session.commit()
        return jsonify({'success': True, 'documento': _documento_a_json(doc, 'academia')}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route("/api/repositorio/documento/<seccion>/<int:doc_id>/descargar", methods=["GET"])
def descargar_repositorio_documento(seccion, doc_id):
    """Descarga un documento por sección (biblioteca|academia) e id."""
    if seccion not in ('biblioteca', 'academia'):
        return jsonify({'success': False, 'error': 'Sección no válida'}), 400
    try:
        if seccion == 'biblioteca':
            doc = BibliotecaDocumento.query.get(doc_id)
        else:
            doc = AcademiaDocumento.query.get(doc_id)
        if not doc:
            return jsonify({'success': False, 'error': 'Documento no encontrado'}), 404
        folder = 'repositorio_biblioteca' if seccion == 'biblioteca' else 'repositorio_academia'
        filepath = os.path.join(os.getcwd(), folder, doc.archivo_filename)
        if not os.path.exists(filepath):
            return jsonify({'success': False, 'error': 'Archivo no encontrado en disco'}), 404
        return send_file(filepath, as_attachment=True, download_name=doc.nombre_archivo or doc.archivo_filename)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route("/api/biblioteca/documentos/<int:doc_id>", methods=["DELETE"])
def eliminar_biblioteca_documento(doc_id):
    """Elimina un documento de la Biblioteca."""
    try:
        doc = BibliotecaDocumento.query.get(doc_id)
        if not doc:
            return jsonify({'success': False, 'error': 'Documento no encontrado'}), 404
        filepath = os.path.join(os.getcwd(), 'repositorio_biblioteca', doc.archivo_filename)
        if os.path.exists(filepath):
            try:
                os.remove(filepath)
            except OSError:
                pass
        db.session.delete(doc)
        db.session.commit()
        return jsonify({'success': True}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route("/api/academia/documentos/<int:doc_id>", methods=["DELETE"])
def eliminar_academia_documento(doc_id):
    """Elimina un documento de la Academia."""
    try:
        doc = AcademiaDocumento.query.get(doc_id)
        if not doc:
            return jsonify({'success': False, 'error': 'Documento no encontrado'}), 404
        filepath = os.path.join(os.getcwd(), 'repositorio_academia', doc.archivo_filename)
        if os.path.exists(filepath):
            try:
                os.remove(filepath)
            except OSError:
                pass
        db.session.delete(doc)
        db.session.commit()
        return jsonify({'success': True}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ----- Academia: documentos por organización (panel de organización) -----

@app.route("/api/organizaciones/<int:org_id>/academia/documentos", methods=["GET"])
def listar_academia_documentos_organizacion(org_id):
    """Lista documentos educativos de la organización (panel Academia)."""
    try:
        Organizacion.query.get_or_404(org_id)
        docs = AcademiaDocumento.query.filter_by(organizacion_id=org_id).order_by(AcademiaDocumento.created_at.desc()).all()
        return jsonify({
            'success': True,
            'documentos': [_documento_a_json(d, 'academia') for d in docs]
        }), 200
    except Exception as e:
        err_msg = str(e).lower()
        if 'categoria' in err_msg and ('column' in err_msg or 'no existe' in err_msg or 'undefined' in err_msg):
            db.session.rollback()
            try:
                db.session.execute(text(
                    "ALTER TABLE academia_documentos ADD COLUMN IF NOT EXISTS categoria VARCHAR(80)"
                ))
                db.session.commit()
                docs = AcademiaDocumento.query.filter_by(organizacion_id=org_id).order_by(AcademiaDocumento.created_at.desc()).all()
                return jsonify({
                    'success': True,
                    'documentos': [_documento_a_json(d, 'academia') for d in docs]
                }), 200
            except Exception as e2:
                db.session.rollback()
                return jsonify({
                    'success': False,
                    'error': 'Columna categoria no existe. Ejecuta en PostgreSQL: ALTER TABLE academia_documentos ADD COLUMN IF NOT EXISTS categoria VARCHAR(80);'
                }), 500
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route("/api/organizaciones/<int:org_id>/academia/documentos", methods=["POST"])
def subir_academia_documento_organizacion(org_id):
    """Sube un documento educativo de la organización (multipart: archivo, nombre_archivo, descripcion obligatoria, fecha_creacion?)."""
    try:
        Organizacion.query.get_or_404(org_id)
        descripcion = (request.form.get('descripcion') or '').strip()[:500]
        if not descripcion:
            return jsonify({'success': False, 'error': 'La descripción breve es obligatoria.'}), 400
        categoria = (request.form.get('categoria') or '').strip().lower()
        if categoria not in ACADEMIA_CATEGORIAS:
            return jsonify({'success': False, 'error': 'Selecciona una categoría válida: Guías y Manuales, Plantillas y Formatos o Videos Tutoriales.'}), 400
        payload, err = _subir_documento_repositorio('academia', request)
        if err:
            return jsonify({'success': False, 'error': err}), 400
        filename, mime, size = payload
        nombre_archivo = (request.form.get('nombre_archivo') or '').strip() or (request.files.get('archivo') and request.files.get('archivo').filename) or 'Documento'
        fecha_creacion_str = request.form.get('fecha_creacion', '').strip()
        fecha_edicion = datetime.utcnow()
        if fecha_creacion_str:
            try:
                fecha_edicion = datetime.strptime(fecha_creacion_str[:10], '%Y-%m-%d')
            except ValueError:
                pass
        doc = AcademiaDocumento(
            organizacion_id=org_id,
            nombre_archivo=nombre_archivo[:255],
            archivo_filename=filename,
            archivo_mime=mime,
            archivo_tamano_bytes=size,
            autor=None,
            fecha_edicion=fecha_edicion,
            descripcion=descripcion,
            estado='pendiente',  # Requiere validación del admin para ser público
            categoria=categoria,
        )
        db.session.add(doc)
        db.session.commit()
        return jsonify({'success': True, 'documento': _documento_a_json(doc, 'academia')}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route("/api/organizaciones/<int:org_id>/academia/documentos/<int:doc_id>", methods=["PATCH"])
def actualizar_academia_documento_organizacion(org_id, doc_id):
    """La organización puede editar un documento solo si está pendiente o rechazado. Aprobado no se puede editar."""
    try:
        doc = AcademiaDocumento.query.filter_by(id=doc_id, organizacion_id=org_id).first()
        if not doc:
            return jsonify({'success': False, 'error': 'Documento no encontrado'}), 404
        estado = (doc.estado or '').lower()
        if estado == 'aprobado':
            return jsonify({
                'success': False,
                'error': 'No se puede editar un documento aprobado. Solo se pueden editar documentos en estado Pendiente o Rechazado.'
            }), 403
        # Actualizar campos del formulario
        nombre_archivo = (request.form.get('nombre_archivo') or '').strip()[:255]
        if nombre_archivo:
            doc.nombre_archivo = nombre_archivo
        descripcion = (request.form.get('descripcion') or '').strip()[:500]
        if descripcion:
            doc.descripcion = descripcion
        categoria = (request.form.get('categoria') or '').strip().lower()
        if categoria in ACADEMIA_CATEGORIAS:
            doc.categoria = categoria
        fecha_creacion_str = (request.form.get('fecha_creacion') or '').strip()
        if fecha_creacion_str:
            try:
                doc.fecha_edicion = datetime.strptime(fecha_creacion_str[:10], '%Y-%m-%d')
            except ValueError:
                pass
        # Archivo opcional: si envían uno nuevo, reemplazar
        if request.files and request.files.get('archivo') and request.files.get('archivo').filename:
            payload, err = _subir_documento_repositorio('academia', request)
            if err:
                return jsonify({'success': False, 'error': err}), 400
            filename_new, mime_new, size_new = payload
            # Borrar archivo antiguo
            old_path = os.path.join(os.getcwd(), 'repositorio_academia', doc.archivo_filename)
            if os.path.exists(old_path):
                try:
                    os.remove(old_path)
                except OSError:
                    pass
            doc.archivo_filename = filename_new
            doc.archivo_mime = mime_new
            doc.archivo_tamano_bytes = size_new
        db.session.commit()
        return jsonify({'success': True, 'documento': _documento_a_json(doc, 'academia')}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route("/api/organizaciones/<int:org_id>/academia/documentos/<int:doc_id>", methods=["DELETE"])
def eliminar_academia_documento_organizacion(org_id, doc_id):
    """Elimina un documento educativo de la organización (solo si pertenece a esa org)."""
    try:
        doc = AcademiaDocumento.query.filter_by(id=doc_id, organizacion_id=org_id).first()
        if not doc:
            return jsonify({'success': False, 'error': 'Documento no encontrado'}), 404
        filepath = os.path.join(os.getcwd(), 'repositorio_academia', doc.archivo_filename)
        if os.path.exists(filepath):
            try:
                os.remove(filepath)
            except OSError:
                pass
        db.session.delete(doc)
        db.session.commit()
        return jsonify({'success': True}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route("/api/admin/academia/documentos/<int:doc_id>/estado", methods=["PATCH"])
def actualizar_estado_academia_documento(doc_id):
    """Admin: cambiar estado de un documento de Academia (pendiente, aprobado, rechazado)."""
    try:
        data = request.get_json() or {}
        nuevo_estado = (data.get('estado') or '').strip().lower()
        if nuevo_estado not in ('pendiente', 'aprobado', 'rechazado'):
            return jsonify({'success': False, 'error': 'estado debe ser "pendiente", "aprobado" o "rechazado"'}), 400
        doc = AcademiaDocumento.query.get(doc_id)
        if not doc:
            return jsonify({'success': False, 'error': 'Documento no encontrado'}), 404
        doc.estado = nuevo_estado
        db.session.commit()
        return jsonify({
            'success': True,
            'documento': _documento_a_json(doc, 'academia', include_organizacion_nombre=True)
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# Obtener todas las reseñas de todas las organizaciones (para admin)
@app.route("/api/admin/reseñas/todas", methods=["GET"])
def obtener_todas_reseñas_admin():
    try:
        from sqlalchemy import text
        
        # Parámetro para agrupar por organización u oportunidad
        agrupar_por = request.args.get('agrupar_por', default='oportunidad', type=str)  # 'oportunidad' o 'organizacion'
        
        # Limpiar sesión
        db.session.rollback()
        
        # Verificar si existen las columnas de reseña
        try:
            check_col_query = text("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name = 'postulaciones' 
                AND (column_name = 'reseña_org' 
                     OR column_name = 'resena_org'
                     OR column_name = 'calificacion_usuario_org'
                     OR column_name = 'calificacion_org'
                     OR column_name = 'resena_usuario_publica')
            """)
            existing_cols = db.session.execute(check_col_query).fetchall()
            existing_col_names = [row[0] for row in existing_cols]
            
            has_resena_col = any(name in existing_col_names for name in [
                'reseña_org', 'resena_org'
            ])
            resena_col_name = None
            for name in ['reseña_org', 'resena_org']:
                if name in existing_col_names:
                    resena_col_name = name
                    break
            
            has_calif_col = 'calificacion_usuario_org' in existing_col_names or 'calificacion_org' in existing_col_names
            calif_col_name = 'calificacion_usuario_org' if 'calificacion_usuario_org' in existing_col_names else ('calificacion_org' if 'calificacion_org' in existing_col_names else None)
            has_publica_col = 'resena_usuario_publica' in existing_col_names
        except:
            has_resena_col = False
            has_calif_col = False
            has_publica_col = False
            resena_col_name = None
            calif_col_name = None
        
        if not has_resena_col:
            return jsonify({
                'success': True,
                'reseñas': []
            }), 200
        
        # Construir consulta SQL para obtener todas las reseñas
        base_query = """
            SELECT 
                p.id as postulacion_id,
                p.usuario_id,
                p.oportunidad_id,
                p.created_at,
        """
        
        if resena_col_name:
            if 'ñ' in resena_col_name or 'ó' in resena_col_name:
                base_query += f'p."{resena_col_name}" as reseña,'
            else:
                base_query += f'p.{resena_col_name} as reseña,'
        else:
            base_query += 'NULL as reseña,'
        
        if calif_col_name:
            base_query += f'p.{calif_col_name} as calificacion,'
        else:
            base_query += 'NULL as calificacion,'
        
        if has_publica_col:
            base_query += 'COALESCE(p.resena_usuario_publica, true) as es_publica,'
        else:
            base_query += 'true as es_publica,'
        
        base_query += """
                u.nombre as usuario_nombre,
                u.apellido as usuario_apellido,
                o.id as organizacion_id,
                o.nombre as organizacion_nombre,
                op.titulo as oportunidad_titulo
            FROM postulaciones p
            INNER JOIN usuarios u ON p.usuario_id = u.id
            INNER JOIN oportunidades op ON p.oportunidad_id = op.id
            INNER JOIN organizaciones o ON op.organizacion_id = o.id
            WHERE 
        """
        
        if resena_col_name:
            if 'ñ' in resena_col_name or 'ó' in resena_col_name:
                base_query += f'p."{resena_col_name}" IS NOT NULL AND p."{resena_col_name}" != \'\''
            else:
                base_query += f'p.{resena_col_name} IS NOT NULL AND p.{resena_col_name} != \'\''
        
        base_query += """
            ORDER BY p.created_at DESC
        """
        
        # Ejecutar consulta
        sql_query = text(base_query)
        results = db.session.execute(sql_query).fetchall()
        
        # Organizar según el parámetro agrupar_por
        if agrupar_por == 'organizacion':
            # Agrupar por organización, luego por oportunidad
            reseñas_agrupadas = {}
            
            for row in results:
                postulacion_id = row[0]
                usuario_id = row[1]
                oportunidad_id = row[2]
                fecha_creacion = row[3]
                resena_text = row[4]
                calificacion = float(row[5]) if row[5] is not None else None
                es_publica = row[6] if len(row) > 6 else True
                usuario_nombre = row[7] or ''
                usuario_apellido = row[8] or ''
                organizacion_id = row[9]
                organizacion_nombre = row[10] or 'Organización'
                oportunidad_titulo = row[11] or 'Oportunidad'
                
                nombre_completo = f"{usuario_nombre} {usuario_apellido}".strip() or 'Usuario'
                
                if organizacion_id not in reseñas_agrupadas:
                    reseñas_agrupadas[organizacion_id] = {
                        'organizacion_id': organizacion_id,
                        'organizacion_nombre': organizacion_nombre,
                        'oportunidades': {}
                    }
                
                if oportunidad_id not in reseñas_agrupadas[organizacion_id]['oportunidades']:
                    reseñas_agrupadas[organizacion_id]['oportunidades'][oportunidad_id] = {
                        'oportunidad_id': oportunidad_id,
                        'oportunidad_titulo': oportunidad_titulo,
                        'reseñas': []
                    }
                
                reseña_data = {
                    'postulacion_id': postulacion_id,
                    'usuario_id': usuario_id,
                    'usuario_nombre': nombre_completo,
                    'reseña': resena_text,
                    'calificacion': calificacion,
                    'fecha': fecha_creacion.strftime('%Y-%m-%d %H:%M:%S') if fecha_creacion else None,
                    'es_publica': es_publica
                }
                
                reseñas_agrupadas[organizacion_id]['oportunidades'][oportunidad_id]['reseñas'].append(reseña_data)
            
            # Convertir a lista
            resultado = []
            for org_id, org_data in reseñas_agrupadas.items():
                oportunidades_lista = list(org_data['oportunidades'].values())
                resultado.append({
                    'organizacion_id': org_data['organizacion_id'],
                    'organizacion_nombre': org_data['organizacion_nombre'],
                    'oportunidades': oportunidades_lista
                })
            
        else:
            # Agrupar solo por oportunidad (por defecto)
            reseñas_por_oportunidad = {}
            
            for row in results:
                postulacion_id = row[0]
                usuario_id = row[1]
                oportunidad_id = row[2]
                fecha_creacion = row[3]
                resena_text = row[4]
                calificacion = float(row[5]) if row[5] is not None else None
                es_publica = row[6] if len(row) > 6 else True
                usuario_nombre = row[7] or ''
                usuario_apellido = row[8] or ''
                organizacion_id = row[9]
                organizacion_nombre = row[10] or 'Organización'
                oportunidad_titulo = row[11] or 'Oportunidad'
                
                nombre_completo = f"{usuario_nombre} {usuario_apellido}".strip() or 'Usuario'
                
                if oportunidad_id not in reseñas_por_oportunidad:
                    reseñas_por_oportunidad[oportunidad_id] = {
                        'oportunidad_id': oportunidad_id,
                        'oportunidad_titulo': oportunidad_titulo,
                        'organizacion_id': organizacion_id,
                        'organizacion_nombre': organizacion_nombre,
                        'reseñas': []
                    }
                
                reseña_data = {
                    'postulacion_id': postulacion_id,
                    'usuario_id': usuario_id,
                    'usuario_nombre': nombre_completo,
                    'reseña': resena_text,
                    'calificacion': calificacion,
                    'fecha': fecha_creacion.strftime('%Y-%m-%d %H:%M:%S') if fecha_creacion else None,
                    'es_publica': es_publica
                }
                
                reseñas_por_oportunidad[oportunidad_id]['reseñas'].append(reseña_data)
            
            resultado = list(reseñas_por_oportunidad.values())
        
        return jsonify({
            'success': True,
            'agrupar_por': agrupar_por,
            'reseñas': resultado
        }), 200
        
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"Error en obtener_todas_reseñas_admin: {str(e)}")
        print(error_trace)
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': f'Error al obtener todas las reseñas: {str(e)}'
        }), 500

# Endpoint para enviar correo masivo a postulantes
@app.route("/api/oportunidades/<int:oportunidad_id>/correo-masivo", methods=["POST"])
def enviar_correo_masivo(oportunidad_id):
    try:
        data = request.json
        asunto = data.get('asunto')
        mensaje = data.get('mensaje')
        filtro_estado = data.get('filtro_estado', 'todos')  # 'todos' o un estado específico
        
        if not asunto or not mensaje:
            return jsonify({
                'success': False,
                'error': 'El asunto y el mensaje son requeridos'
            }), 400
        
        # Obtener la oportunidad
        oportunidad = Oportunidad.query.get(oportunidad_id)
        if not oportunidad:
            return jsonify({
                'success': False,
                'error': 'Oportunidad no encontrada'
            }), 404
        
        # Obtener postulaciones según el filtro
        query = Postulacion.query.filter_by(oportunidad_id=oportunidad_id)
        if filtro_estado != 'todos':
            query = query.filter_by(estado=filtro_estado)
        
        postulaciones = query.all()
        
        if not postulaciones:
            return jsonify({
                'success': False,
                'error': 'No hay postulantes para enviar el correo'
            }), 400
        
        # Obtener información de la organización
        organizacion = Organizacion.query.get(oportunidad.organizacion_id)
        organizacion_nombre = organizacion.nombre if organizacion else 'la organización'
        
        # Enviar correo a cada postulante
        emails_enviados = 0
        emails_fallidos = []
        
        for postulacion in postulaciones:
            usuario = Usuario.query.get(postulacion.usuario_id)
            if not usuario or not usuario.email:
                emails_fallidos.append(f"Usuario {usuario.id if usuario else postulacion.usuario_id} sin email")
                continue
            
            try:
                msg = Message(
                    subject=asunto,
                    recipients=[usuario.email],
                    html=f"""
                    <!DOCTYPE html>
                    <html>
                    <head>
                        <meta charset="UTF-8">
                        <style>
                            body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
                            .container {{ max-width: 600px; margin: 0 auto; padding: 20px; background-color: #f9f9f9; }}
                            .header {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 20px; text-align: center; border-radius: 8px 8px 0 0; }}
                            .content {{ background: white; padding: 20px; border-radius: 0 0 8px 8px; }}
                            .footer {{ text-align: center; margin-top: 20px; color: #666; font-size: 12px; }}
                        </style>
                    </head>
                    <body>
                        <div class="container">
                            <div class="header">
                                <h1>INJUV - Comunicación de Organización</h1>
                            </div>
                            <div class="content">
                                <p>Hola <strong>{usuario.nombre or 'Usuario'}</strong>,</p>
                                
                                <p>{mensaje.replace(chr(10), '<br>')}</p>
                                
                                <h3>Detalles:</h3>
                                <ul>
                                    <li><strong>Oportunidad:</strong> {oportunidad.titulo}</li>
                                    <li><strong>Organización:</strong> {organizacion_nombre}</li>
                                </ul>
                                
                                <p>Saludos cordiales,<br>
                                <strong>{organizacion_nombre}</strong><br>
                                <small>Equipo INJUV</small></p>
                            </div>
                            <div class="footer">
                                <p>Este es un email automático enviado por {organizacion_nombre} a través de la plataforma INJUV.</p>
                            </div>
                        </div>
                    </body>
                    </html>
                    """
                )
                mail.send(msg)
                emails_enviados += 1
            except Exception as e:
                print(f"Error enviando email a {usuario.email}: {str(e)}")
                emails_fallidos.append(usuario.email)
        
        return jsonify({
            'success': True,
            'message': f'Correos enviados exitosamente',
            'emails_enviados': emails_enviados,
            'emails_fallidos': emails_fallidos,
            'total': len(postulaciones)
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Endpoint para actualizar horas de voluntariado de un usuario
@app.route("/api/usuarios/<int:usuario_id>/horas-voluntariado", methods=["PUT"])
def actualizar_horas_voluntariado(usuario_id):
    try:
        data = request.json
        horas = data.get('hora_voluntariado') or data.get('horas_voluntariado')
        
        if horas is None:
            return jsonify({
                'success': False,
                'error': 'Las horas de voluntariado son requeridas'
            }), 400
        
        try:
            horas_int = int(horas)
            if horas_int < 0:
                return jsonify({
                    'success': False,
                    'error': 'Las horas no pueden ser negativas'
                }), 400
        except (ValueError, TypeError):
            return jsonify({
                'success': False,
                'error': 'Las horas deben ser un número entero válido'
            }), 400
        
        usuario = Usuario.query.get(usuario_id)
        if not usuario:
            return jsonify({
                'success': False,
                'error': 'Usuario no encontrado'
            }), 404
        
        usuario.hora_voluntariado = horas_int
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Horas de voluntariado actualizadas exitosamente',
            'hora_voluntariado': usuario.hora_voluntariado
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Subir certificado de voluntario
@app.route("/api/postulaciones/<int:postulacion_id>/certificado", methods=["POST"])
def subir_certificado_voluntario(postulacion_id):
    try:
        postulacion = Postulacion.query.get(postulacion_id)
        if not postulacion:
            return jsonify({
                'success': False,
                'error': 'Postulación no encontrada'
            }), 404
        
        if 'certificado' not in request.files:
            return jsonify({
                'success': False,
                'error': 'No se proporcionó ningún archivo'
            }), 400
        
        file = request.files['certificado']
        
        if file.filename == '':
            return jsonify({
                'success': False,
                'error': 'No se seleccionó ningún archivo'
            }), 400
        
        # Validar que sea PDF
        if not file.filename.lower().endswith('.pdf'):
            return jsonify({
                'success': False,
                'error': 'El archivo debe ser un PDF'
            }), 400
        
        # Crear directorio de certificados si no existe
        import os
        certificados_dir = os.path.join(os.getcwd(), 'certificados_voluntarios')
        if not os.path.exists(certificados_dir):
            os.makedirs(certificados_dir)
        
        # Generar nombre único para el archivo
        from werkzeug.utils import secure_filename
        filename = f"certificado_{postulacion_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        filename = secure_filename(filename)
        filepath = os.path.join(certificados_dir, filename)
        
        # Guardar el archivo
        file.save(filepath)
        
        # Guardar la ruta en la base de datos
        postulacion.ruta_certificado_pdf = filepath
        postulacion.tiene_certificado = True
        postulacion.updated_at = datetime.now()
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Certificado subido exitosamente',
            'ruta': filepath
        }), 200
        
    except Exception as e:
        db.session.rollback()
        import traceback
        print(f"Error al subir certificado: {str(e)}")
        print(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Enviar certificado por correo
@app.route("/api/postulaciones/<int:postulacion_id>/enviar-certificado", methods=["POST"])
def enviar_certificado_por_correo(postulacion_id):
    try:
        data = request.json
        email_destino = data.get('email')
        
        if not email_destino:
            return jsonify({
                'success': False,
                'error': 'El correo electrónico es requerido'
            }), 400
        
        postulacion = Postulacion.query.get(postulacion_id)
        if not postulacion:
            return jsonify({
                'success': False,
                'error': 'Postulación no encontrada'
            }), 404
        
        if not postulacion.ruta_certificado_pdf:
            return jsonify({
                'success': False,
                'error': 'No hay certificado subido para esta postulación'
            }), 400
        
        # Obtener información del usuario y la oportunidad
        usuario = Usuario.query.get(postulacion.usuario_id)
        oportunidad = Oportunidad.query.get(postulacion.oportunidad_id)
        organizacion = Organizacion.query.get(oportunidad.organizacion_id) if oportunidad else None
        
        # Leer el archivo PDF
        import os
        if not os.path.exists(postulacion.ruta_certificado_pdf):
            return jsonify({
                'success': False,
                'error': 'El archivo del certificado no existe'
            }), 404
        
        with open(postulacion.ruta_certificado_pdf, 'rb') as f:
            certificado_data = f.read()
        
        # Enviar correo con el certificado adjunto usando Flask-Mail
        try:
            msg = Message(
                subject=f'Certificado de Voluntariado - {oportunidad.titulo if oportunidad else "Voluntariado"}',
                recipients=[email_destino],
                html=f"""
                <html>
                <body>
                    <h2>Certificado de Voluntariado</h2>
                    <p>Estimado/a {usuario.nombre if usuario else 'Voluntario'},</p>
                    <p>Te enviamos adjunto tu certificado de participación en el voluntariado:</p>
                    <p><strong>{oportunidad.titulo if oportunidad else 'Voluntariado'}</strong></p>
                    <p>Organización: <strong>{organizacion.nombre if organizacion else 'Organización'}</strong></p>
                    <p>Gracias por tu participación y compromiso con el voluntariado.</p>
                    <p>Saludos cordiales,<br>Equipo INJUV</p>
                </body>
                </html>
                """
            )
            
            # Adjuntar el certificado PDF
            import os
            if os.path.exists(postulacion.ruta_certificado_pdf):
                with open(postulacion.ruta_certificado_pdf, 'rb') as f:
                    msg.attach(
                        'certificado_voluntariado.pdf',
                        'application/pdf',
                        f.read()
                    )
            
            mail.send(msg)
            
            return jsonify({
                'success': True,
                'message': f'Certificado enviado exitosamente a {email_destino}'
            }), 200
            
        except Exception as email_error:
            # Si Flask-Mail no está configurado o hay un error, retornar un mensaje informativo
            print(f"Error al enviar correo: {str(email_error)}")
            return jsonify({
                'success': False,
                'error': f'Error al enviar el correo: {str(email_error)}. Verifica la configuración de correo en el servidor.'
            }), 500
        
    except Exception as e:
        db.session.rollback()
        import traceback
        print(f"Error al enviar certificado: {str(e)}")
        print(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Descargar certificado de voluntario
@app.route("/api/postulaciones/<int:postulacion_id>/certificado/descargar", methods=["GET"])
def descargar_certificado_voluntario(postulacion_id):
    try:
        postulacion = Postulacion.query.get(postulacion_id)
        if not postulacion:
            return jsonify({
                'success': False,
                'error': 'Postulación no encontrada'
            }), 404
        
        if not postulacion.ruta_certificado_pdf:
            return jsonify({
                'success': False,
                'error': 'No hay certificado disponible para esta postulación'
            }), 404
        
        import os
        if not os.path.exists(postulacion.ruta_certificado_pdf):
            return jsonify({
                'success': False,
                'error': 'El archivo del certificado no existe'
            }), 404
        
        from flask import send_file
        return send_file(
            postulacion.ruta_certificado_pdf,
            as_attachment=True,
            download_name=f'certificado_voluntariado_{postulacion_id}.pdf',
            mimetype='application/pdf'
        )
        
    except Exception as e:
        import traceback
        print(f"Error al descargar certificado: {str(e)}")
        print(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# ==================== ENDPOINTS DE ADMINISTRACIÓN ====================

# Endpoint para obtener todos los usuarios con sus roles
@app.route("/api/admin/usuarios", methods=["GET"])
def obtener_usuarios_admin():
    try:
        usuarios = Usuario.query.all()
        
        usuarios_data = []
        for usuario in usuarios:
            # Obtener organización si es admin u organizacion
            organizacion = None
            if usuario.rol in ['admin', 'organizacion']:
                organizacion = Organizacion.query.filter_by(id_usuario_org=usuario.id).first()
            
            usuarios_data.append({
                'id': usuario.id,
                'nombre': usuario.nombre or '',
                'apellido': usuario.apellido or '',
                'email': usuario.email or '',
                'rut': usuario.rut or '',
                'rol': usuario.rol or 'user',
                'telefono': usuario.telefono or '',
                'region': usuario.region or '',
                'comuna': usuario.comuna or '',
                'organizacion': organizacion.nombre if organizacion else None,
                'organizacion_id': organizacion.id if organizacion else None,
                'created_at': usuario.created_at.isoformat() if usuario.created_at else None
            })
        
        return jsonify({
            'success': True,
            'usuarios': usuarios_data,
            'total': len(usuarios_data)
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Endpoint para generar reporte Excel de usuarios
@app.route("/api/admin/usuarios/generar-reporte", methods=["POST"])
def generar_reporte_usuarios():
    try:
        if not OPENPYXL_AVAILABLE:
            return jsonify({
                'success': False,
                'error': 'openpyxl no está instalado. Instala con: pip install openpyxl'
            }), 500
        
        data = request.json or {}
        rol_filter = data.get('rol')
        region_filter = data.get('region')
        
        # Obtener todos los usuarios
        query = Usuario.query
        if rol_filter:
            query = query.filter(Usuario.rol == rol_filter)
        if region_filter:
            query = query.filter(Usuario.region == region_filter)
        
        usuarios = query.all()
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Usuarios"
        
        # Estilos
        header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=16)
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws['A1'] = "REPORTE DE USUARIOS REGISTRADOS"
        ws['A1'].font = title_font
        ws.merge_cells('A1:O1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Información del reporte
        row = 3
        ws[f'A{row}'] = f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws[f'A{row}'].font = Font(size=10, italic=True)
        row += 1
        if rol_filter:
            ws[f'A{row}'] = f"Rol filtrado: {rol_filter}"
            ws[f'A{row}'].font = Font(size=10, italic=True)
            row += 1
        if region_filter:
            ws[f'A{row}'] = f"Región filtrada: {region_filter}"
            ws[f'A{row}'].font = Font(size=10, italic=True)
            row += 1
        ws[f'A{row}'] = f"Total de usuarios: {len(usuarios)}"
        ws[f'A{row}'].font = Font(bold=True, size=11)
        row += 2
        
        # Encabezados
        headers = [
            'ID', 'Nombre', 'Apellido', 'RUT', 'Email',
            'Teléfono', 'Rol', 'Región', 'Ciudad', 'Comuna',
            'Sexo', 'Fecha Nacimiento', 'Horas Voluntariado',
            'Organización', 'Fecha Registro'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border_style
        
        row += 1
        
        # Datos de usuarios
        for usuario in usuarios:
            try:
                # Obtener organización si es admin u organizacion
                organizacion_nombre = None
                if usuario.rol in ['admin', 'organizacion']:
                    org = Organizacion.query.filter_by(id_usuario_org=usuario.id).first()
                    organizacion_nombre = org.nombre if org else None
                
                # Formatear fecha de nacimiento
                fecha_nacimiento = ''
                if usuario.fecha_nacimiento:
                    fecha_nacimiento = usuario.fecha_nacimiento.strftime('%d/%m/%Y')
                
                # Formatear fecha de registro
                fecha_registro = ''
                if usuario.created_at:
                    fecha_registro = usuario.created_at.strftime('%d/%m/%Y')
                
                # Obtener horas de voluntariado
                horas_voluntariado = usuario.hora_voluntariado if usuario.hora_voluntariado else 0
                
                # Escribir datos
                datos = [
                    usuario.id,
                    usuario.nombre or '',
                    usuario.apellido or '',
                    usuario.rut or '',
                    usuario.email or '',
                    usuario.telefono or '',
                    usuario.rol or 'user',
                    usuario.region or '',
                    usuario.ciudad or '',
                    usuario.comuna or '',
                    usuario.sexo or '',
                    fecha_nacimiento,
                    horas_voluntariado,
                    organizacion_nombre or '',
                    fecha_registro
                ]
                
                for col, valor in enumerate(datos, 1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = valor
                    cell.border = border_style
                    cell.alignment = Alignment(
                        horizontal='left' if col in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15] else 'left',
                        vertical='center',
                        wrap_text=True
                    )
                
                row += 1
            except Exception as user_error:
                print(f"Error al procesar usuario ID {usuario.id}: {user_error}")
                import traceback
                print(traceback.format_exc())
                continue
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 8   # ID
        ws.column_dimensions['B'].width = 20  # Nombre
        ws.column_dimensions['C'].width = 20  # Apellido
        ws.column_dimensions['D'].width = 15  # RUT
        ws.column_dimensions['E'].width = 30  # Email
        ws.column_dimensions['F'].width = 15  # Teléfono
        ws.column_dimensions['G'].width = 15  # Rol
        ws.column_dimensions['H'].width = 25  # Región
        ws.column_dimensions['I'].width = 20  # Ciudad
        ws.column_dimensions['J'].width = 20  # Comuna
        ws.column_dimensions['K'].width = 12  # Sexo
        ws.column_dimensions['L'].width = 15  # Fecha Nacimiento
        ws.column_dimensions['M'].width = 18  # Horas Voluntariado
        ws.column_dimensions['N'].width = 30  # Organización
        ws.column_dimensions['O'].width = 15  # Fecha Registro
        
        # Congelar primera fila de encabezados
        if row > len(usuarios) + 6:
            ws.freeze_panes = f'A{row - len(usuarios)}'
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        rol_suffix = f"_{rol_filter}" if rol_filter else ""
        region_suffix = f"_{region_filter.replace(' ', '_')}" if region_filter else ""
        filename = f'reporte_usuarios{rol_suffix}{region_suffix}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"Error al generar reporte de usuarios: {str(e)}")
        print(error_trace)
        
        return jsonify({
            'success': False,
            'error': f"Error al generar el reporte: {str(e)}"
        }), 500

# Endpoint para obtener formularios de creación de organización (admin)
@app.route("/api/admin/solicitudes-organizacion", methods=["GET"])
def obtener_solicitudes_organizacion_admin():
    try:
        estado = request.args.get('estado')
        query = SolicitudOrganizacion.query
        if estado:
            query = query.filter_by(estado=estado)

        solicitudes = query.order_by(SolicitudOrganizacion.created_at.desc()).all()
        data = []
        for solicitud in solicitudes:
            usuario = Usuario.query.get(solicitud.id_usuario_org)
            revisor = Usuario.query.get(solicitud.revisado_por_admin_id) if solicitud.revisado_por_admin_id else None
            data.append({
                'id': solicitud.id,
                'nombre': solicitud.nombre,
                'rut': solicitud.rut or '',
                'email_contacto': solicitud.email_contacto or '',
                'fecha_creacion': solicitud.fecha_creacion.isoformat() if solicitud.fecha_creacion else None,
                'region': solicitud.region,
                'ciudad': solicitud.ciudad,
                'comuna': solicitud.comuna,
                'descripcion': solicitud.descripcion,
                'sitio_web': solicitud.sitio_web or '',
                'redes_sociales': solicitud.redes_sociales or [],
                'estado': solicitud.estado,
                'comentario_revision': solicitud.comentario_revision or '',
                'id_usuario_org': solicitud.id_usuario_org,
                'usuario_nombre': f"{usuario.nombre} {usuario.apellido}".strip() if usuario else '',
                'usuario_email': usuario.email if usuario else '',
                'revisado_por': f"{revisor.nombre} {revisor.apellido}".strip() if revisor else '',
                'created_at': solicitud.created_at.isoformat() if solicitud.created_at else None,
                'reviewed_at': solicitud.reviewed_at.isoformat() if solicitud.reviewed_at else None
            })

        return jsonify({'success': True, 'solicitudes': data, 'total': len(data)}), 200
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route("/api/admin/solicitudes-organizacion/<int:solicitud_id>/estado", methods=["PATCH"])
def actualizar_estado_solicitud_organizacion(solicitud_id):
    try:
        data = request.json or {}
        nuevo_estado = data.get('estado')
        comentario = data.get('comentario', '')
        admin_id = data.get('admin_id')

        if nuevo_estado not in ['aprobada', 'rechazada']:
            return jsonify({'success': False, 'error': 'Estado inválido. Debe ser aprobada o rechazada.'}), 400

        solicitud = SolicitudOrganizacion.query.get(solicitud_id)
        if not solicitud:
            return jsonify({'success': False, 'error': 'Solicitud no encontrada'}), 404

        if solicitud.estado != 'pendiente':
            return jsonify({'success': False, 'error': f'La solicitud ya fue {solicitud.estado}.'}), 400

        usuario = Usuario.query.get(solicitud.id_usuario_org)
        if not usuario:
            return jsonify({'success': False, 'error': 'Usuario asociado no encontrado'}), 404

        solicitud.estado = nuevo_estado
        solicitud.comentario_revision = comentario if comentario else None
        solicitud.revisado_por_admin_id = admin_id if admin_id else None
        solicitud.reviewed_at = datetime.now()

        organizacion_creada = None
        if nuevo_estado == 'aprobada':
            if solicitud.rut:
                rut_existente = Organizacion.query.filter_by(rut=solicitud.rut).first()
                if rut_existente:
                    return jsonify({'success': False, 'error': 'No se puede aprobar: el RUT ya existe en otra organización.'}), 400

            org_existente_usuario = Organizacion.query.filter_by(id_usuario_org=solicitud.id_usuario_org).first()
            if org_existente_usuario:
                return jsonify({'success': False, 'error': 'No se puede aprobar: el usuario ya administra una organización.'}), 400

            organizacion_creada = Organizacion(
                nombre=solicitud.nombre,
                rut=solicitud.rut if solicitud.rut else None,
                email_contacto=solicitud.email_contacto if solicitud.email_contacto else usuario.email,
                telefono_contacto=None,
                fecha_creacion=solicitud.fecha_creacion,
                region=solicitud.region,
                ciudad=solicitud.ciudad,
                comuna=solicitud.comuna,
                descripcion=solicitud.descripcion,
                sitio_web=solicitud.sitio_web if solicitud.sitio_web else None,
                redes_sociales=solicitud.redes_sociales if solicitud.redes_sociales else [],
                id_usuario_org=solicitud.id_usuario_org,
                created_at=datetime.now()
            )
            db.session.add(organizacion_creada)

            if usuario.rol != 'admin':
                usuario.rol = 'organizacion'

        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'Solicitud {nuevo_estado} exitosamente',
            'solicitud_id': solicitud.id,
            'estado': solicitud.estado,
            'organizacion_id': organizacion_creada.id if organizacion_creada else None
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

# Endpoint para obtener todas las organizaciones con sus administradores
@app.route("/api/admin/organizaciones", methods=["GET"])
def obtener_organizaciones_admin():
    try:
        organizaciones = Organizacion.query.all()
        
        organizaciones_data = []
        for org in organizaciones:
            try:
                # Obtener usuario administrador
                usuario_org = Usuario.query.get(org.id_usuario_org) if org.id_usuario_org else None
                
                # Contar oportunidades de esta organización
                num_oportunidades = Oportunidad.query.filter_by(organizacion_id=org.id).count()
                
                # Obtener ciudad de forma segura usando getattr
                ciudad_val = getattr(org, 'ciudad', None) or ''
                
                organizaciones_data.append({
                    'id': org.id,
                    'nombre': org.nombre or '',
                    'rut': org.rut or '',
                    'email_contacto': org.email_contacto or '',
                    'telefono_contacto': org.telefono_contacto or '',
                    'region': org.region or '',
                    'ciudad': ciudad_val,
                    'comuna': org.comuna or '',
                    'descripcion': org.descripcion or '',
                    'area_trabajo': org.area_trabajo or '',
                    'fecha_creacion': org.fecha_creacion.isoformat() if org.fecha_creacion else None,
                    'usuario_org_id': org.id_usuario_org,
                    'admin_nombre': f"{usuario_org.nombre} {usuario_org.apellido}".strip() if usuario_org else None,
                    'admin_email': usuario_org.email if usuario_org else None,
                    'admin_rol': usuario_org.rol if usuario_org else None,
                    'num_oportunidades': num_oportunidades,
                    'created_at': org.created_at.isoformat() if org.created_at else None
                })
            except Exception as org_error:
                print(f"Error al procesar organización ID {org.id}: {org_error}")
                import traceback
                print(traceback.format_exc())
                # Continuar con la siguiente organización
                continue
        
        return jsonify({
            'success': True,
            'organizaciones': organizaciones_data,
            'total': len(organizaciones_data)
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Endpoint para generar reporte de impacto de una organización (formato INJUV)
@app.route("/api/organizaciones/<int:organizacion_id>/reporte-impacto", methods=["GET"])
def generar_reporte_impacto_organizacion(organizacion_id):
    try:
        if not OPENPYXL_AVAILABLE:
            return jsonify({
                'success': False,
                'error': 'openpyxl no está instalado. Instala con: pip install openpyxl'
            }), 500
        
        # Obtener parámetros de filtro
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        oportunidad_id = request.args.get('oportunidad_id')
        
        # Obtener la organización
        organizacion = Organizacion.query.get(organizacion_id)
        if not organizacion:
            return jsonify({
                'success': False,
                'error': 'Organización no encontrada'
            }), 404
        
        # Obtener oportunidades de la organización
        query = Oportunidad.query.filter_by(organizacion_id=organizacion_id)
        if oportunidad_id:
            query = query.filter(Oportunidad.id == oportunidad_id)
        
        oportunidades = query.all()
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Impacto"
        
        # Estilos
        header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=16)
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws['A1'] = f"REPORTE DE IMPACTO - {organizacion.nombre.upper()}"
        ws['A1'].font = title_font
        ws.merge_cells('A1:H1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Información del reporte
        row = 3
        ws[f'A{row}'] = f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws[f'A{row}'].font = Font(size=10, italic=True)
        row += 1
        if fecha_inicio:
            ws[f'A{row}'] = f"Período desde: {fecha_inicio}"
            ws[f'A{row}'].font = Font(size=10, italic=True)
            row += 1
        if fecha_fin:
            ws[f'A{row}'] = f"Período hasta: {fecha_fin}"
            ws[f'A{row}'].font = Font(size=10, italic=True)
            row += 1
        row += 1
        
        # Encabezados
        headers = ['Oportunidad', 'Voluntario', 'Estado', 'Asistencia Capacitación', 
                   'Asistencia Actividad', 'Calificación', 'Certificado', 'Fecha Postulación']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border_style
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row += 1
        
        # Datos de postulaciones
        total_voluntarios = 0
        total_horas = 0
        total_actividades = 0
        total_certificados = 0
        
        for oportunidad in oportunidades:
            # Filtrar por fecha si se especifica
            postulaciones = Postulacion.query.filter_by(oportunidad_id=oportunidad.id).all()
            
            for post in postulaciones:
                # Filtrar por fecha si se especifica
                if fecha_inicio or fecha_fin:
                    post_date = post.created_at.date() if post.created_at else None
                    if fecha_inicio and post_date and post_date < datetime.strptime(fecha_inicio, '%Y-%m-%d').date():
                        continue
                    if fecha_fin and post_date and post_date > datetime.strptime(fecha_fin, '%Y-%m-%d').date():
                        continue
                
                usuario = Usuario.query.get(post.usuario_id)
                if not usuario:
                    continue
                
                # Solo incluir seleccionados para el reporte de impacto
                if post.estado != 'Seleccionado':
                    continue
                
                total_voluntarios += 1
                if post.asistencia_actividad:
                    total_actividades += 1
                    total_horas += 8  # Estimación: 8 horas por actividad
                if post.tiene_certificado:
                    total_certificados += 1
                
                # Escribir datos
                data_row = [
                    oportunidad.titulo,
                    f"{usuario.nombre} {usuario.apellido}".strip(),
                    post.estado,
                    'Sí' if post.asistencia_capacitacion else 'No',
                    'Sí' if post.asistencia_actividad else 'No',
                    post.calificacion_org if post.calificacion_org else 'Sin calificar',
                    'Sí' if post.tiene_certificado else 'No',
                    post.created_at.strftime('%d/%m/%Y') if post.created_at else 'N/A'
                ]
                
                for col, value in enumerate(data_row, 1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = value
                    cell.border = border_style
                
                row += 1
        
        # Resumen
        row += 1
        ws[f'A{row}'] = "RESUMEN DE IMPACTO"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        ws.merge_cells(f'A{row}:H{row}')
        row += 1
        
        resumen_data = [
            ['Total de Voluntarios', total_voluntarios],
            ['Total de Horas de Voluntariado', total_horas],
            ['Total de Actividades Completadas', total_actividades],
            ['Total de Certificados Emitidos', total_certificados]
        ]
        
        for item in resumen_data:
            ws[f'A{row}'] = item[0]
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = item[1]
            row += 1
        
        # Ajustar ancho de columnas
        column_widths = [30, 25, 15, 20, 20, 15, 15, 18]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width
        
        # Guardar en memoria
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f'reporte_impacto_{organizacion.nombre.replace(" ", "_")}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"Error al generar reporte de impacto: {str(e)}")
        print(error_trace)
        
        return jsonify({
            'success': False,
            'error': f"Error al generar el reporte: {str(e)}"
        }), 500

# Endpoint para generar reporte Excel de organizaciones
@app.route("/api/admin/organizaciones/generar-reporte", methods=["POST"])
def generar_reporte_organizaciones():
    try:
        if not OPENPYXL_AVAILABLE:
            return jsonify({
                'success': False,
                'error': 'openpyxl no está instalado. Instala con: pip install openpyxl'
            }), 500
        
        data = request.json or {}
        region_filter = data.get('region')
        
        # Obtener todas las organizaciones
        query = Organizacion.query
        if region_filter:
            query = query.filter(Organizacion.region == region_filter)
        
        organizaciones = query.all()
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Organizaciones"
        
        # Estilos
        header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=16)
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws['A1'] = "REPORTE DE ORGANIZACIONES REGISTRADAS"
        ws['A1'].font = title_font
        ws.merge_cells('A1:J1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Información del reporte
        row = 3
        ws[f'A{row}'] = f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws[f'A{row}'].font = Font(size=10, italic=True)
        row += 1
        if region_filter:
            ws[f'A{row}'] = f"Región filtrada: {region_filter}"
            ws[f'A{row}'].font = Font(size=10, italic=True)
            row += 1
        ws[f'A{row}'] = f"Total de organizaciones: {len(organizaciones)}"
        ws[f'A{row}'].font = Font(bold=True, size=11)
        row += 2
        
        # Encabezados
        headers = [
            'ID', 'Nombre', 'RUT', 'Email Contacto', 'Teléfono',
            'Región', 'Ciudad', 'Comuna', 'Área de Trabajo',
            'Administrador', 'Email Admin', 'Oportunidades Creadas',
            'Fecha Creación', 'Descripción'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border_style
        
        row += 1
        
        # Datos de organizaciones
        for org in organizaciones:
            try:
                # Obtener usuario administrador
                usuario_org = Usuario.query.get(org.id_usuario_org) if org.id_usuario_org else None
                admin_nombre = f"{usuario_org.nombre} {usuario_org.apellido}".strip() if usuario_org else 'Sin administrador'
                admin_email = usuario_org.email if usuario_org else ''
                
                # Contar oportunidades
                num_oportunidades = Oportunidad.query.filter_by(organizacion_id=org.id).count()
                
                # Obtener ciudad
                ciudad_val = getattr(org, 'ciudad', None) or ''
                
                # Formatear fecha
                fecha_creacion = ''
                if org.fecha_creacion:
                    fecha_creacion = org.fecha_creacion.strftime('%d/%m/%Y')
                elif org.created_at:
                    fecha_creacion = org.created_at.strftime('%d/%m/%Y')
                
                # Descripción truncada
                descripcion = (org.descripcion or '')[:100] + ('...' if org.descripcion and len(org.descripcion) > 100 else '')
                
                # Escribir datos
                datos = [
                    org.id,
                    org.nombre or '',
                    org.rut or '',
                    org.email_contacto or '',
                    org.telefono_contacto or '',
                    org.region or '',
                    ciudad_val,
                    org.comuna or '',
                    org.area_trabajo or '',
                    admin_nombre,
                    admin_email,
                    num_oportunidades,
                    fecha_creacion,
                    descripcion
                ]
                
                for col, valor in enumerate(datos, 1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = valor
                    cell.border = border_style
                    cell.alignment = Alignment(
                        horizontal='left' if col in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14] else 'left',
                        vertical='center',
                        wrap_text=True
                    )
                
                row += 1
            except Exception as org_error:
                print(f"Error al procesar organización ID {org.id}: {org_error}")
                continue
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 8   # ID
        ws.column_dimensions['B'].width = 30  # Nombre
        ws.column_dimensions['C'].width = 15  # RUT
        ws.column_dimensions['D'].width = 30  # Email Contacto
        ws.column_dimensions['E'].width = 15  # Teléfono
        ws.column_dimensions['F'].width = 25  # Región
        ws.column_dimensions['G'].width = 20  # Ciudad
        ws.column_dimensions['H'].width = 20  # Comuna
        ws.column_dimensions['I'].width = 25  # Área de Trabajo
        ws.column_dimensions['J'].width = 25  # Administrador
        ws.column_dimensions['K'].width = 30  # Email Admin
        ws.column_dimensions['L'].width = 18  # Oportunidades
        ws.column_dimensions['M'].width = 15  # Fecha Creación
        ws.column_dimensions['N'].width = 50  # Descripción
        
        # Congelar primera fila de encabezados
        ws.freeze_panes = f'A{row - len(organizaciones)}'
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        region_suffix = f"_{region_filter.replace(' ', '_')}" if region_filter else ""
        filename = f'reporte_organizaciones{region_suffix}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"Error al generar reporte de organizaciones: {str(e)}")
        print(error_trace)
        
        return jsonify({
            'success': False,
            'error': f"Error al generar el reporte: {str(e)}"
        }), 500

# Endpoint para actualizar el rol de un usuario y asignar organización
@app.route("/api/admin/usuarios/<int:user_id>/rol", methods=["PUT"])
def actualizar_rol_usuario(user_id):
    try:
        data = request.json
        nuevo_rol = data.get('rol')
        organizacion_id = data.get('organizacion_id')  # Nueva: ID de organización a asignar
        
        if not nuevo_rol:
            return jsonify({
                'success': False,
                'error': 'El rol es requerido'
            }), 400
        
        usuario = Usuario.query.get(user_id)
        if not usuario:
            return jsonify({
                'success': False,
                'error': 'Usuario no encontrado'
            }), 404
        
        # Si se está asignando rol "organizacion"
        if nuevo_rol == 'organizacion':
            if organizacion_id:
                # Verificar que la organización existe
                organizacion = Organizacion.query.get(organizacion_id)
                if not organizacion:
                    return jsonify({
                        'success': False,
                        'error': 'La organización especificada no existe'
                    }), 404
                
                # Si la organización ya tiene otro usuario asignado, desasignarla primero
                if organizacion.id_usuario_org and organizacion.id_usuario_org != user_id:
                    usuario_anterior_id = organizacion.id_usuario_org
                    usuario_anterior = Usuario.query.get(usuario_anterior_id)
                    
                    # Si el usuario anterior solo tiene rol "organizacion" (sin otras organizaciones), cambiar su rol a "user"
                    if usuario_anterior and usuario_anterior.rol == 'organizacion':
                        # Verificar si tiene otras organizaciones
                        otras_orgs = Organizacion.query.filter_by(id_usuario_org=usuario_anterior_id).filter(Organizacion.id != organizacion_id).first()
                        if not otras_orgs:
                            # No tiene otras organizaciones, cambiar su rol a "user"
                            usuario_anterior.rol = 'user'
                            print(f"Rol del usuario anterior {usuario_anterior.email} cambiado a 'user' porque ya no tiene organizaciones asignadas")
                
                # Asignar la organización al nuevo usuario
                organizacion.id_usuario_org = user_id
            else:
                # Si organizacion_id es None/null (opción "Ninguna organización"), desasignar cualquier organización actual
                organizacion_existente = Organizacion.query.filter_by(id_usuario_org=user_id).first()
                if organizacion_existente:
                    organizacion_existente.id_usuario_org = None
                    print(f"Organización '{organizacion_existente.nombre}' desasignada del usuario {user_id}")
        
        # Si se está cambiando de admin/organizacion a otro rol, verificar si tiene organización
        if usuario.rol in ['admin', 'organizacion'] and nuevo_rol not in ['admin', 'organizacion']:
            organizacion_existente = Organizacion.query.filter_by(id_usuario_org=user_id).first()
            if organizacion_existente:
                # Desasignar la organización (pero no eliminarla)
                organizacion_existente.id_usuario_org = None
        
        # Si se está cambiando de "organizacion" a otro rol sin organización asignada, está bien
        usuario.rol = nuevo_rol
        db.session.commit()
        
        # Obtener información de la organización asignada si existe
        organizacion_asignada = None
        if nuevo_rol == 'organizacion':
            org = Organizacion.query.filter_by(id_usuario_org=user_id).first()
            if org:
                organizacion_asignada = {
                    'id': org.id,
                    'nombre': org.nombre
                }
        
        return jsonify({
            'success': True,
            'message': 'Rol actualizado exitosamente',
            'usuario': {
                'id': usuario.id,
                'email': usuario.email,
                'nombre': usuario.nombre,
                'apellido': usuario.apellido,
                'rol': usuario.rol,
                'organizacion': organizacion_asignada
            }
        }), 200
        
    except Exception as e:
        db.session.rollback()
        import traceback
        print(f"Error al actualizar rol de usuario: {str(e)}")
        print(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Endpoint para eliminar el rol de un usuario (poner en 'user' por defecto)
@app.route("/api/admin/usuarios/<int:user_id>/rol", methods=["DELETE"])
def eliminar_rol_usuario(user_id):
    try:
        usuario = Usuario.query.get(user_id)
        if not usuario:
            return jsonify({
                'success': False,
                'error': 'Usuario no encontrado'
            }), 404
        
        # Verificar si es admin de una organización
        if usuario.rol in ['admin', 'organizacion']:
            organizacion = Organizacion.query.filter_by(id_usuario_org=user_id).first()
            if organizacion:
                return jsonify({
                    'success': False,
                    'error': 'No se puede eliminar el rol de un administrador de organización. Primero debe asignar otro administrador a la organización.'
                }), 400
        
        # Establecer rol por defecto
        usuario.rol = 'user'
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Rol eliminado exitosamente (establecido como usuario por defecto)',
            'usuario': {
                'id': usuario.id,
                'email': usuario.email,
                'nombre': usuario.nombre,
                'apellido': usuario.apellido,
                'rol': usuario.rol
            }
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Endpoint para que el administrador elimine usuarios
@app.route("/api/admin/usuarios/<int:user_id>", methods=["DELETE"])
def admin_eliminar_usuario(user_id):
    try:
        # Buscar el usuario
        usuario = Usuario.query.get(user_id)
        
        if not usuario:
            return jsonify({
                'success': False,
                'error': 'Usuario no encontrado'
            }), 404
        
        # Obtener todas las postulaciones del usuario
        postulaciones = Postulacion.query.filter_by(usuario_id=user_id).all()
        
        # Verificar si el usuario es administrador de una organización
        organizacion = Organizacion.query.filter_by(id_usuario_org=user_id).first()
        
        if organizacion:
            # Si tiene una organización, eliminar también las oportunidades y postulaciones asociadas
            oportunidades = Oportunidad.query.filter_by(organizacion_id=organizacion.id).all()
            
            for oportunidad in oportunidades:
                # Eliminar postulaciones de cada oportunidad
                postulaciones_oportunidad = Postulacion.query.filter_by(oportunidad_id=oportunidad.id).all()
                for post in postulaciones_oportunidad:
                    db.session.delete(post)
                # Eliminar la oportunidad
                db.session.delete(oportunidad)
            
            # Eliminar la organización
            db.session.delete(organizacion)
        
        # Eliminar todas las postulaciones del usuario
        for postulacion in postulaciones:
            db.session.delete(postulacion)
        
        # Eliminar el usuario
        db.session.delete(usuario)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Usuario eliminado exitosamente'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        import traceback
        print(f"Error al eliminar usuario: {str(e)}")
        print(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Endpoint para permitir NULL en id_usuario_org
@app.route("/api/admin/migrate/allow-null-usuario-org", methods=["POST"])
def allow_null_usuario_org():
    try:
        with app.app_context():
            # Ejecutar ALTER TABLE para permitir NULL
            alter_query = text("ALTER TABLE organizaciones ALTER COLUMN id_usuario_org DROP NOT NULL")
            db.session.execute(alter_query)
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': 'Columna id_usuario_org ahora permite valores NULL'
            }), 200
    except Exception as e:
        db.session.rollback()
        error_msg = str(e)
        # Si el error es que ya no tiene NOT NULL, considerarlo como éxito
        if 'does not have a NOT NULL constraint' in error_msg or 'does not exist' in error_msg.lower():
            return jsonify({
                'success': True,
                'message': 'La columna ya permite valores NULL o no existe la restricción'
            }), 200
        
        return jsonify({
            'success': False,
            'error': f'Error al ejecutar migración: {error_msg}'
        }), 500

# Endpoint para obtener estadísticas del dashboard
@app.route("/api/admin/estadisticas", methods=["GET"])
def obtener_estadisticas():
    try:
        # Contar usuarios registrados
        total_usuarios = Usuario.query.count()
        
        # Contar oportunidades (voluntariados creados)
        total_oportunidades = Oportunidad.query.count()
        
        # Contar organizaciones creadas
        total_organizaciones = Organizacion.query.count()
        
        # Contar postulaciones
        total_postulaciones = Postulacion.query.count()
        
        # Contar voluntariados por estado
        voluntariados_cerrados = Oportunidad.query.filter_by(estado='cerrada').count()
        voluntariados_activos = Oportunidad.query.filter_by(estado='activa').count()
        voluntariados_abiertos = Oportunidad.query.filter_by(estado='abierta').count()
        voluntariados_en_proceso = voluntariados_activos + voluntariados_abiertos
        
        # Contar noticias activas (verificar si existe la tabla)
        total_noticias = 0
        try:
            from sqlalchemy import text
            # Intentar contar noticias si la tabla existe
            result = db.session.execute(text("""
                SELECT COUNT(*) FROM information_schema.tables 
                WHERE table_name = 'noticias'
            """)).fetchone()
            if result and result[0] > 0:
                # La tabla existe, contar noticias activas
                noticias_result = db.session.execute(text("""
                    SELECT COUNT(*) FROM noticias 
                    WHERE estado = 'activa' OR estado = 'publicada' OR estado IS NULL
                """)).fetchone()
                total_noticias = noticias_result[0] if noticias_result else 0
        except:
            # Si no existe la tabla o hay error, devolver 0
            total_noticias = 0
        
        # Contar documentos de Academia
        total_academia = 0
        try:
            total_academia = AcademiaDocumento.query.count()
        except:
            total_academia = 0
        
        # Contar documentos de Biblioteca
        total_biblioteca = 0
        try:
            total_biblioteca = BibliotecaDocumento.query.count()
        except:
            total_biblioteca = 0
        
        return jsonify({
            'success': True,
            'usuarios_registrados': total_usuarios,
            'voluntariados_creados': total_oportunidades,
            'voluntariados_cerrados': voluntariados_cerrados,
            'voluntariados_en_proceso': voluntariados_en_proceso,
            'organizaciones_creadas': total_organizaciones,
            'noticias_activas': total_noticias,
            'postulaciones': total_postulaciones,
            'documentos_academia': total_academia,
            'documentos_biblioteca': total_biblioteca,
            'estadisticas': {
                'usuarios_registrados': total_usuarios,
                'voluntariados_creados': total_oportunidades,
                'voluntariados_cerrados': voluntariados_cerrados,
                'voluntariados_en_proceso': voluntariados_en_proceso,
                'organizaciones_creadas': total_organizaciones,
                'noticias_activas': total_noticias,
                'postulaciones': total_postulaciones,
                'documentos_academia': total_academia,
                'documentos_biblioteca': total_biblioteca
            }
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Endpoint para obtener estadísticas de voluntarios
@app.route("/api/admin/estadisticas/voluntarios", methods=["GET"])
def obtener_estadisticas_voluntarios():
    try:
        from sqlalchemy import extract, func
        from datetime import datetime
        
        mes_filtro = request.args.get('mes')
        año_filtro = request.args.get('año')
        
        query = db.session.query(
            extract('year', Usuario.created_at).label('año'),
            extract('month', Usuario.created_at).label('mes'),
            func.count(Usuario.id).label('cantidad')
        ).filter(Usuario.created_at.isnot(None))
        
        if año_filtro:
            query = query.filter(extract('year', Usuario.created_at) == int(año_filtro))
        if mes_filtro:
            query = query.filter(extract('month', Usuario.created_at) == int(mes_filtro))
        
        query = query.group_by(
            extract('year', Usuario.created_at),
            extract('month', Usuario.created_at)
        ).order_by(
            extract('year', Usuario.created_at).desc(),
            extract('month', Usuario.created_at).asc()
        )
        
        resultados = query.all()
        
        meses_nombres = {
            1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
            5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
            9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
        }
        
        datos = []
        for r in resultados:
            año = int(r.año) if r.año else None
            mes = int(r.mes) if r.mes else None
            if año and mes:
                datos.append({
                    'label': f"{meses_nombres.get(mes, f'Mes {mes}')} {año}",
                    'mes': meses_nombres.get(mes, f'Mes {mes}'),
                    'año': año,
                    'cantidad': int(r.cantidad) if r.cantidad else 0
                })
        
        return jsonify({
            'success': True,
            'datos': datos,
            'tipo': 'line'
        }), 200
        
    except Exception as e:
        import traceback
        print(f"Error en obtener_estadisticas_voluntarios: {str(e)}")
        print(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Endpoint para obtener estadísticas de organizaciones
@app.route("/api/admin/estadisticas/organizaciones", methods=["GET"])
def obtener_estadisticas_organizaciones():
    try:
        from sqlalchemy import func
        
        tipo = request.args.get('tipo', 'area_trabajo')  # 'area_trabajo' o 'voluntariados_por_org'
        
        if tipo == 'area_trabajo':
            # Estadísticas por área de trabajo
            query = db.session.query(
                Organizacion.area_trabajo.label('area'),
                func.count(Organizacion.id).label('cantidad')
            ).filter(Organizacion.area_trabajo.isnot(None)).filter(Organizacion.area_trabajo != '')
            
            query = query.group_by(Organizacion.area_trabajo).order_by(func.count(Organizacion.id).desc())
            
            resultados = query.all()
            
            datos = []
            for r in resultados:
                datos.append({
                    'label': r.area or 'Sin área',
                    'cantidad': int(r.cantidad) if r.cantidad else 0
                })
            
            return jsonify({
                'success': True,
                'datos': datos,
                'tipo': 'pie'
            }), 200
            
        elif tipo == 'voluntariados_por_org':
            # Voluntariados creados por organización
            query = db.session.query(
                Organizacion.nombre.label('organizacion'),
                func.count(Oportunidad.id).label('cantidad')
            ).join(Oportunidad, Organizacion.id == Oportunidad.organizacion_id)
            
            query = query.group_by(Organizacion.nombre).order_by(func.count(Oportunidad.id).desc()).limit(10)
            
            resultados = query.all()
            
            datos = []
            for r in resultados:
                datos.append({
                    'label': r.organizacion or 'Sin nombre',
                    'cantidad': int(r.cantidad) if r.cantidad else 0
                })
            
            return jsonify({
                'success': True,
                'datos': datos,
                'tipo': 'bar'
            }), 200
        
        else:
            return jsonify({
                'success': False,
                'error': 'Tipo no válido'
            }), 400
        
    except Exception as e:
        import traceback
        print(f"Error en obtener_estadisticas_organizaciones: {str(e)}")
        print(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Endpoint para obtener estadísticas de voluntariados
@app.route("/api/admin/estadisticas/voluntariados", methods=["GET"])
def obtener_estadisticas_voluntariados():
    try:
        from sqlalchemy import func
        
        tipo = request.args.get('tipo', 'estado')  # 'estado' o 'mensual'
        
        if tipo == 'estado':
            # Estadísticas por estado (creados, cerrados, en proceso)
            estados = {
                'cerrada': Oportunidad.query.filter_by(estado='cerrada').count(),
                'activa': Oportunidad.query.filter_by(estado='activa').count(),
                'abierta': Oportunidad.query.filter_by(estado='abierta').count()
            }
            
            datos = [
                {'label': 'Cerrados', 'cantidad': estados['cerrada']},
                {'label': 'En Proceso', 'cantidad': estados['activa'] + estados['abierta']},
                {'label': 'Total Creados', 'cantidad': Oportunidad.query.count()}
            ]
            
            return jsonify({
                'success': True,
                'datos': datos,
                'tipo': 'pie'
            }), 200
            
        elif tipo == 'mensual':
            # Voluntariados creados por mes
            from sqlalchemy import extract
            
            mes_filtro = request.args.get('mes')
            año_filtro = request.args.get('año')
            
            query = db.session.query(
                extract('year', Oportunidad.created_at).label('año'),
                extract('month', Oportunidad.created_at).label('mes'),
                func.count(Oportunidad.id).label('cantidad')
            ).filter(Oportunidad.created_at.isnot(None))
            
            if año_filtro:
                query = query.filter(extract('year', Oportunidad.created_at) == int(año_filtro))
            if mes_filtro:
                query = query.filter(extract('month', Oportunidad.created_at) == int(mes_filtro))
            
            query = query.group_by(
                extract('year', Oportunidad.created_at),
                extract('month', Oportunidad.created_at)
            ).order_by(
                extract('year', Oportunidad.created_at).desc(),
                extract('month', Oportunidad.created_at).asc()
            )
            
            resultados = query.all()
            
            meses_nombres = {
                1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
                5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
                9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
            }
            
            datos = []
            for r in resultados:
                año = int(r.año) if r.año else None
                mes = int(r.mes) if r.mes else None
                if año and mes:
                    datos.append({
                        'label': f"{meses_nombres.get(mes, f'Mes {mes}')} {año}",
                        'mes': meses_nombres.get(mes, f'Mes {mes}'),
                        'año': año,
                        'cantidad': int(r.cantidad) if r.cantidad else 0
                    })
            
            return jsonify({
                'success': True,
                'datos': datos,
                'tipo': 'line'
            }), 200
        
        else:
            return jsonify({
                'success': False,
                'error': 'Tipo no válido'
            }), 400
        
    except Exception as e:
        import traceback
        print(f"Error en obtener_estadisticas_voluntariados: {str(e)}")
        print(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Endpoint para obtener estadísticas mensuales
@app.route("/api/admin/estadisticas/mensuales", methods=["GET"])
def obtener_estadisticas_mensuales():
    try:
        from sqlalchemy import extract, func
        from datetime import datetime
        
        # Obtener parámetros de filtro
        mes_filtro = request.args.get('mes')
        año_filtro = request.args.get('año')
        
        # Obtener todos los años disponibles
        años_usuarios = db.session.query(extract('year', Usuario.created_at).label('año')).distinct().all()
        años_oportunidades = db.session.query(extract('year', Oportunidad.created_at).label('año')).distinct().all()
        años_disponibles = sorted(set([int(a[0]) for a in años_usuarios + años_oportunidades if a[0] is not None]), reverse=True)
        
        # Construir query base para usuarios
        query_usuarios = db.session.query(
            extract('year', Usuario.created_at).label('año'),
            extract('month', Usuario.created_at).label('mes'),
            func.count(Usuario.id).label('cantidad')
        ).filter(Usuario.created_at.isnot(None))
        
        # Construir query base para oportunidades
        query_oportunidades = db.session.query(
            extract('year', Oportunidad.created_at).label('año'),
            extract('month', Oportunidad.created_at).label('mes'),
            func.count(Oportunidad.id).label('cantidad')
        ).filter(Oportunidad.created_at.isnot(None))
        
        # Aplicar filtros si existen
        if año_filtro:
            año_int = int(año_filtro)
            query_usuarios = query_usuarios.filter(extract('year', Usuario.created_at) == año_int)
            query_oportunidades = query_oportunidades.filter(extract('year', Oportunidad.created_at) == año_int)
        
        if mes_filtro:
            mes_int = int(mes_filtro)
            query_usuarios = query_usuarios.filter(extract('month', Usuario.created_at) == mes_int)
            query_oportunidades = query_oportunidades.filter(extract('month', Oportunidad.created_at) == mes_int)
        
        # Agrupar por año y mes
        query_usuarios = query_usuarios.group_by(
            extract('year', Usuario.created_at),
            extract('month', Usuario.created_at)
        )
        
        query_oportunidades = query_oportunidades.group_by(
            extract('year', Oportunidad.created_at),
            extract('month', Oportunidad.created_at)
        )
        
        # Ejecutar queries
        usuarios_stats = query_usuarios.all()
        oportunidades_stats = query_oportunidades.all()
        
        # Obtener voluntarios activos (usuarios con postulaciones) por mes
        query_activos = db.session.query(
            extract('year', Postulacion.created_at).label('año'),
            extract('month', Postulacion.created_at).label('mes'),
            func.count(func.distinct(Postulacion.usuario_id)).label('activos')
        ).filter(Postulacion.created_at.isnot(None))
        
        if año_filtro:
            año_int = int(año_filtro)
            query_activos = query_activos.filter(extract('year', Postulacion.created_at) == año_int)
        
        if mes_filtro:
            mes_int = int(mes_filtro)
            query_activos = query_activos.filter(extract('month', Postulacion.created_at) == mes_int)
        
        query_activos = query_activos.group_by(
            extract('year', Postulacion.created_at),
            extract('month', Postulacion.created_at)
        )
        
        activos_stats = query_activos.all()
        
        # Mapear nombres de meses
        meses_nombres = {
            1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
            5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
            9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
        }
        
        # Procesar datos de usuarios
        voluntarios_data = []
        for stat in usuarios_stats:
            año = int(stat.año) if stat.año else None
            mes = int(stat.mes) if stat.mes else None
            if año and mes:
                voluntarios_data.append({
                    'mes': meses_nombres.get(mes, f'Mes {mes}'),
                    'año': año,
                    'cantidad': int(stat.cantidad) if stat.cantidad else 0
                })
        
        # Procesar datos de oportunidades
        voluntariados_data = []
        for stat in oportunidades_stats:
            año = int(stat.año) if stat.año else None
            mes = int(stat.mes) if stat.mes else None
            if año and mes:
                voluntariados_data.append({
                    'mes': meses_nombres.get(mes, f'Mes {mes}'),
                    'año': año,
                    'cantidad': int(stat.cantidad) if stat.cantidad else 0
                })
        
        # Procesar datos de activos
        activos_map = {}
        for stat in activos_stats:
            año = int(stat.año) if stat.año else None
            mes = int(stat.mes) if stat.mes else None
            if año and mes:
                key = f"{meses_nombres.get(mes, f'Mes {mes}')}-{año}"
                activos_map[key] = int(stat.activos) if stat.activos else 0
        
        # Combinar datos
        combined_data = []
        all_keys = set()
        
        for item in voluntarios_data:
            key = f"{item['mes']}-{item['año']}"
            all_keys.add(key)
        
        for item in voluntariados_data:
            key = f"{item['mes']}-{item['año']}"
            all_keys.add(key)
        
        for key in all_keys:
            mes, año = key.rsplit('-', 1)
            año_int = int(año)
            
            voluntarios_item = next((x for x in voluntarios_data if x['mes'] == mes and x['año'] == año_int), None)
            voluntariados_item = next((x for x in voluntariados_data if x['mes'] == mes and x['año'] == año_int), None)
            
            combined_data.append({
                'mes': mes,
                'año': año_int,
                'voluntariados': voluntariados_item['cantidad'] if voluntariados_item else 0,
                'voluntarios': voluntarios_item['cantidad'] if voluntarios_item else 0,
                'activos': activos_map.get(key, 0)
            })
        
        # Ordenar por año y mes
        meses_orden = list(meses_nombres.values())
        combined_data.sort(key=lambda x: (x['año'], meses_orden.index(x['mes']) if x['mes'] in meses_orden else 999), reverse=True)
        
        return jsonify({
            'success': True,
            'estadisticas': combined_data,
            'años_disponibles': años_disponibles
        }), 200
        
    except Exception as e:
        import traceback
        print(f"Error en obtener_estadisticas_mensuales: {str(e)}")
        print(traceback.format_exc())
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Endpoint temporal para actualizar la tabla oportunidades
@app.route("/api/admin/update-oportunidades-table", methods=["POST"])
def update_oportunidades_table():
    """Endpoint temporal para agregar columnas faltantes a la tabla oportunidades"""
    try:
        from sqlalchemy import text
        
        messages = []
        
        # Verificar y agregar organizacion_id si no existe
        check_query = text("""
            SELECT 1 FROM information_schema.columns 
            WHERE table_name = 'oportunidades' AND column_name = 'organizacion_id'
        """)
        result = db.session.execute(check_query).fetchone()
        
        if not result:
            # Agregar la columna
            alter_query = text("ALTER TABLE oportunidades ADD COLUMN organizacion_id INTEGER")
            db.session.execute(alter_query)
            messages.append('Columna organizacion_id agregada')
            
            # Verificar y agregar foreign key
            fk_check = text("""
                SELECT 1 FROM information_schema.table_constraints 
                WHERE constraint_name = 'fk_oportunidades_organizacion' 
                AND table_name = 'oportunidades'
            """)
            fk_exists = db.session.execute(fk_check).fetchone()
            
            if not fk_exists:
                fk_query = text("""
                    ALTER TABLE oportunidades 
                    ADD CONSTRAINT fk_oportunidades_organizacion 
                    FOREIGN KEY (organizacion_id) REFERENCES organizaciones(id)
                """)
                db.session.execute(fk_query)
                messages.append('Foreign key agregada')
        else:
            messages.append('Columna organizacion_id ya existe')
        
        # Verificar y agregar area_voluntariado si no existe
        check_area = text("""
            SELECT 1 FROM information_schema.columns 
            WHERE table_name = 'oportunidades' AND column_name = 'area_voluntariado'
        """)
        result_area = db.session.execute(check_area).fetchone()
        
        if not result_area:
            # Agregar la columna
            alter_area = text("ALTER TABLE oportunidades ADD COLUMN area_voluntariado VARCHAR(100)")
            db.session.execute(alter_area)
            messages.append('Columna area_voluntariado agregada')
        else:
            messages.append('Columna area_voluntariado ya existe')
        
        db.session.commit()
        return jsonify({
            'success': True,
            'message': '; '.join(messages)
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Endpoint temporal para corregir el tipo de fecha_creacion
@app.route("/api/admin/fix-fecha-creacion-type", methods=["POST"])
def fix_fecha_creacion_type():
    """Endpoint temporal para corregir el tipo de la columna fecha_creacion de INTEGER a DATE"""
    try:
        from sqlalchemy import text
        
        # Verificar el tipo actual de la columna
        check_query = text("""
            SELECT data_type FROM information_schema.columns 
            WHERE table_name = 'organizaciones' AND column_name = 'fecha_creacion'
        """)
        result = db.session.execute(check_query).fetchone()
        
        if not result:
            return jsonify({
                'success': False,
                'error': 'La columna fecha_creacion no existe en la tabla organizaciones'
            }), 404
        
        if result[0] == 'date':
            return jsonify({
                'success': True,
                'message': 'La columna fecha_creacion ya es de tipo DATE'
            }), 200
        
        # Cambiar el tipo de INTEGER a DATE
        alter_query = text("""
            ALTER TABLE organizaciones 
            ALTER COLUMN fecha_creacion TYPE DATE USING 
                CASE 
                    WHEN fecha_creacion IS NULL THEN NULL
                    WHEN fecha_creacion::text ~ '^\\d{4}-\\d{2}-\\d{2}$' THEN fecha_creacion::text::DATE
                    ELSE NULL
                END
        """)
        db.session.execute(alter_query)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Tipo de columna fecha_creacion corregido de INTEGER a DATE'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Endpoint temporal para actualizar la tabla organizaciones
@app.route("/api/admin/update-organizaciones-table", methods=["POST"])
def update_organizaciones_table():
    """Endpoint temporal para agregar columnas faltantes a la tabla organizaciones"""
    try:
        from sqlalchemy import text
        
        columns_to_add = [
            ('id_usuario_org', 'INTEGER'),
            ('ciudad', 'VARCHAR(100)'),
            ('siglas_nombre', 'VARCHAR(100)'),
            ('documentos_legales', 'JSONB DEFAULT \'[]\'::jsonb'),
            ('fecha_creacion', 'DATE'),
            ('descripcion_breve', 'VARCHAR(500)'),
            ('area_trabajo', 'VARCHAR(100)'),
            ('tipo_org', 'VARCHAR(100)'),
            ('sitio_web', 'TEXT'),
            ('redes_sociales', 'JSONB DEFAULT \'[]\'::jsonb'),
            ('experiencia_anios', 'INTEGER'),
            ('voluntarios_anuales', 'VARCHAR(100)'),
            ('certificacion', 'JSONB DEFAULT \'[]\'::jsonb'),
            ('created_at', 'TIMESTAMP'),
        ]
        
        added_columns = []
        existing_columns = []
        
        for column_name, column_def in columns_to_add:
            # Verificar si la columna existe
            check_query = text("""
                SELECT 1, data_type FROM information_schema.columns 
                WHERE table_name = 'organizaciones' AND column_name = :col_name
            """)
            result = db.session.execute(check_query, {'col_name': column_name}).fetchone()
            
            if not result:
                # Agregar la columna
                alter_query = text(f"ALTER TABLE organizaciones ADD COLUMN {column_name} {column_def}")
                db.session.execute(alter_query)
                added_columns.append(column_name)
            else:
                # Si la columna existe pero tiene el tipo incorrecto, corregirlo
                if column_name == 'fecha_creacion' and result[1] != 'date':
                    # Cambiar el tipo de INTEGER a DATE
                    alter_type_query = text("""
                        ALTER TABLE organizaciones 
                        ALTER COLUMN fecha_creacion TYPE DATE USING 
                            CASE 
                                WHEN fecha_creacion IS NULL THEN NULL
                                WHEN fecha_creacion::text ~ '^\\d{4}-\\d{2}-\\d{2}$' THEN fecha_creacion::text::DATE
                                ELSE NULL
                            END
                    """)
                    db.session.execute(alter_type_query)
                    added_columns.append(f'{column_name} (tipo corregido)')
                else:
                    existing_columns.append(column_name)
        
        # Agregar foreign key si no existe
        fk_check = text("""
            SELECT 1 FROM information_schema.table_constraints 
            WHERE constraint_name = 'fk_organizaciones_usuario_org' 
            AND table_name = 'organizaciones'
        """)
        fk_exists = db.session.execute(fk_check).fetchone()
        
        if not fk_exists:
            col_check = text("""
                SELECT 1 FROM information_schema.columns 
                WHERE table_name = 'organizaciones' AND column_name = 'id_usuario_org'
            """)
            if db.session.execute(col_check).fetchone():
                fk_query = text("""
                    ALTER TABLE organizaciones 
                    ADD CONSTRAINT fk_organizaciones_usuario_org 
                    FOREIGN KEY (id_usuario_org) REFERENCES usuarios(id)
                """)
                db.session.execute(fk_query)
        
        # Corregir el tipo de fecha_creacion si es necesario
        fecha_check = text("""
            SELECT data_type FROM information_schema.columns 
            WHERE table_name = 'organizaciones' AND column_name = 'fecha_creacion'
        """)
        fecha_result = db.session.execute(fecha_check).fetchone()
        
        if fecha_result and fecha_result[0] != 'date':
            # Cambiar el tipo de INTEGER a DATE
            alter_fecha_query = text("""
                ALTER TABLE organizaciones 
                ALTER COLUMN fecha_creacion TYPE DATE USING 
                    CASE 
                        WHEN fecha_creacion IS NULL THEN NULL
                        WHEN fecha_creacion::text ~ '^\\d{4}-\\d{2}-\\d{2}$' THEN fecha_creacion::text::DATE
                        ELSE NULL
                    END
            """)
            db.session.execute(alter_fecha_query)
            print("✅ Tipo de columna 'fecha_creacion' corregido de INTEGER a DATE")
        
        # Verificar y agregar organizacion_id a oportunidades si no existe
        check_oportunidades = text("""
            SELECT 1 FROM information_schema.columns 
            WHERE table_name = 'oportunidades' AND column_name = 'organizacion_id'
        """)
        result_oportunidades = db.session.execute(check_oportunidades).fetchone()
        
        if not result_oportunidades:
            alter_oportunidades = text("ALTER TABLE oportunidades ADD COLUMN organizacion_id INTEGER")
            db.session.execute(alter_oportunidades)
            added_columns.append('oportunidades.organizacion_id')
            
            # Agregar foreign key si no existe
            fk_check_oportunidades = text("""
                SELECT 1 FROM information_schema.table_constraints 
                WHERE constraint_name = 'fk_oportunidades_organizacion' 
                AND table_name = 'oportunidades'
            """)
            fk_exists_oportunidades = db.session.execute(fk_check_oportunidades).fetchone()
            
            if not fk_exists_oportunidades:
                fk_query_oportunidades = text("""
                    ALTER TABLE oportunidades 
                    ADD CONSTRAINT fk_oportunidades_organizacion 
                    FOREIGN KEY (organizacion_id) REFERENCES organizaciones(id)
                """)
                db.session.execute(fk_query_oportunidades)
        else:
            existing_columns.append('oportunidades.organizacion_id')
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Tablas actualizadas exitosamente',
            'added_columns': added_columns,
            'existing_columns': existing_columns
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Endpoint para generar reporte Excel de estadísticas
@app.route("/api/admin/estadisticas/generar-reporte", methods=["POST"])
def generar_reporte_excel():
    try:
        # Intentar importar openpyxl dinámicamente (por si se instaló después de iniciar el servidor)
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            openpyxl_available = True
            import_error_msg = None
        except ImportError as e:
            openpyxl_available = False
            import_error_msg = str(e)
            import sys
            print(f"❌ Error al importar openpyxl: {import_error_msg}")
            print(f"❌ Python ejecutándose: {sys.executable}")
            print(f"❌ Asegúrate de que el servidor Flask esté corriendo con el entorno virtual activado")
        
        if not openpyxl_available:
            error_message = f'openpyxl no está instalado. Instala con: pip install openpyxl'
            if import_error_msg:
                error_message += f'\nError detallado: {import_error_msg}'
            return jsonify({
                'success': False,
                'error': error_message
            }), 500
        
        data = request.json
        categoria = data.get('categoria', 'voluntarios')
        titulo = data.get('titulo', 'Estadísticas')
        datos_grafico = data.get('datos', [])
        tipo_grafico = data.get('tipo_grafico', 'bar')
        labels = data.get('labels', [])
        values = data.get('values', [])
        imagen_grafico = data.get('imagen_grafico', '')
        filtros = data.get('filtros', {})
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Estadísticas"
        
        # Estilos
        header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=14)
        title_font = Font(bold=True, size=16)
        subtitle_font = Font(size=12)
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws['A1'] = f"Reporte de Estadísticas - {titulo}"
        ws['A1'].font = title_font
        ws.merge_cells('A1:D1')
        
        # Información del reporte
        row = 3
        ws[f'A{row}'] = f"Categoría: {categoria.capitalize()}"
        ws[f'A{row}'].font = subtitle_font
        row += 1
        
        fecha_reporte = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        ws[f'A{row}'] = f"Fecha de generación: {fecha_reporte}"
        ws[f'A{row}'].font = Font(size=10, italic=True)
        row += 2
        
        # Filtros aplicados
        if filtros:
            ws[f'A{row}'] = "Filtros aplicados:"
            ws[f'A{row}'].font = Font(bold=True, size=11)
            row += 1
            if filtros.get('mes'):
                meses = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 
                        'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
                ws[f'A{row}'] = f"Mes: {meses[int(filtros['mes'])]}"
                row += 1
            if filtros.get('año'):
                ws[f'A{row}'] = f"Año: {filtros['año']}"
                row += 1
            if filtros.get('tipo'):
                ws[f'A{row}'] = f"Tipo: {filtros['tipo']}"
                row += 1
            row += 1
        
        # Tabla de datos
        ws[f'A{row}'] = "Datos del Gráfico"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        # Encabezados de la tabla
        headers = ['Categoría', 'Cantidad']
        if tipo_grafico == 'pie':
            headers.append('Porcentaje')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_style
        
        row += 1
        
        # Datos de la tabla
        total = sum(values) if values else 0
        for i, label in enumerate(labels):
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=values[i] if i < len(values) else 0)
            
            if tipo_grafico == 'pie' and total > 0:
                porcentaje = (values[i] / total * 100) if i < len(values) else 0
                ws.cell(row=row, column=3, value=f"{porcentaje:.2f}%")
            
            # Aplicar bordes
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).border = border_style
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='left' if col == 1 else 'center', vertical='center')
            
            row += 1
        
        # Fila de total
        row += 1
        ws.cell(row=row, column=1, value="TOTAL")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=total)
        ws.cell(row=row, column=2).font = Font(bold=True)
        for col in range(1, 3):
            ws.cell(row=row, column=col).border = border_style
            ws.cell(row=row, column=col).fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        
        row += 3
        
        # Imagen del gráfico (si está disponible)
        if imagen_grafico:
            try:
                # Intentar importar Image de openpyxl.drawing
                # Nota: openpyxl.drawing.image requiere Pillow para funcionar
                try:
                    from openpyxl.drawing.image import Image
                    # Verificar si Pillow está disponible
                    try:
                        import PIL
                    except ImportError:
                        raise ImportError("Pillow no está instalado. La imagen no se puede insertar sin Pillow. Instala con: pip install Pillow")
                    
                    # Decodificar imagen base64
                    imagen_data = base64.b64decode(imagen_grafico.split(',')[1])
                    img = Image(io.BytesIO(imagen_data))
                    
                    # Ajustar tamaño de la imagen
                    img.width = 600
                    img.height = 400
                    
                    # Insertar imagen
                    ws.add_image(img, f'A{row}')
                    row += 25  # Espacio después de la imagen
                except ImportError as import_err:
                    # Si no se puede importar Image, simplemente omitir la imagen
                    print(f"Advertencia: No se pudo importar Image de openpyxl.drawing: {import_err}")
                    print("El reporte se generará sin la imagen del gráfico.")
                    ws[f'A{row}'] = "Nota: La imagen del gráfico no está disponible (Pillow no está instalado)"
                    ws[f'A{row}'].font = Font(italic=True, size=10, color="666666")
                    row += 2
            except Exception as img_error:
                print(f"Error al insertar imagen: {img_error}")
                print(f"Tipo de error: {type(img_error).__name__}")
                import traceback
                print(traceback.format_exc())
                # Continuar sin la imagen en lugar de fallar
                ws[f'A{row}'] = f"Nota: No se pudo incluir la imagen del gráfico"
                ws[f'A{row}'].font = Font(italic=True, size=10, color="666666")
                row += 2
        
        # Conclusión
        row += 1
        ws[f'A{row}'] = "CONCLUSIÓN"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        # Generar conclusión basada en los datos
        conclusion = generar_conclusion(categoria, datos_grafico, labels, values, tipo_grafico, total)
        ws[f'A{row}'] = conclusion
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.row_dimensions[row].height = 100
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        if tipo_grafico == 'pie':
            ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 50
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'reporte_estadisticas_{categoria}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"Error al generar reporte Excel: {str(e)}")
        print(error_trace)
        
        # Proporcionar un mensaje de error más útil
        error_message = f"Error al generar el reporte: {str(e)}"
        if "openpyxl" in str(e).lower() or "ImportError" in str(type(e).__name__):
            error_message += "\n\nAsegúrate de que openpyxl esté instalado y reinicia el servidor Flask."
            error_message += "\nInstala con: pip install openpyxl"
        
        return jsonify({
            'success': False,
            'error': error_message
        }), 500

# Función para generar conclusión basada en los datos
def generar_conclusion(categoria, datos, labels, values, tipo_grafico, total):
    conclusion = ""
    
    if categoria == 'voluntarios':
        conclusion += "ANÁLISIS DE VOLUNTARIOS REGISTRADOS\n\n"
        if total > 0:
            max_value = max(values) if values else 0
            max_index = values.index(max_value) if values else 0
            max_label = labels[max_index] if max_index < len(labels) else ""
            
            conclusion += f"El análisis muestra un total de {total:,} voluntarios registrados. "
            if tipo_grafico == 'line':
                conclusion += f"La tendencia indica que {max_label} presenta el mayor número de registros con {max_value:,} voluntarios. "
                conclusion += "Esto sugiere un crecimiento constante en la participación de voluntarios en la plataforma. "
            else:
                conclusion += f"El período con mayor registro es {max_label} con {max_value:,} voluntarios ({max_value/total*100:.1f}% del total). "
                conclusion += "Se observa una distribución variada en los registros a lo largo del tiempo. "
            
            conclusion += "\n\nRECOMENDACIONES:\n"
            conclusion += "- Continuar promoviendo la participación en períodos de menor registro.\n"
            conclusion += "- Analizar las estrategias exitosas en períodos de mayor registro.\n"
            conclusion += "- Implementar campañas dirigidas para mantener el crecimiento."
    
    elif categoria == 'organizaciones':
        conclusion += "ANÁLISIS DE ORGANIZACIONES\n\n"
        if total > 0:
            max_value = max(values) if values else 0
            max_index = values.index(max_value) if values else 0
            max_label = labels[max_index] if max_index < len(labels) else ""
            
            conclusion += f"El análisis muestra un total de {total:,} organizaciones. "
            conclusion += f"El área de trabajo más representativa es '{max_label}' con {max_value:,} organizaciones ({max_value/total*100:.1f}% del total). "
            conclusion += "Esto refleja la diversidad de áreas de trabajo en la plataforma. "
            
            conclusion += "\n\nRECOMENDACIONES:\n"
            conclusion += "- Fomentar la participación de organizaciones en áreas menos representadas.\n"
            conclusion += "- Fortalecer las áreas con mayor presencia para maximizar su impacto.\n"
            conclusion += "- Desarrollar estrategias específicas por área de trabajo."
    
    elif categoria == 'voluntariados':
        conclusion += "ANÁLISIS DE VOLUNTARIADOS\n\n"
        if total > 0:
            max_value = max(values) if values else 0
            max_index = values.index(max_value) if values else 0
            max_label = labels[max_index] if max_index < len(labels) else ""
            
            conclusion += f"El análisis muestra un total de {total:,} voluntariados. "
            if 'estado' in str(labels[0] if labels else '').lower():
                conclusion += f"El estado más común es '{max_label}' con {max_value:,} voluntariados ({max_value/total*100:.1f}% del total). "
                conclusion += "Esto indica el estado general de las oportunidades de voluntariado en la plataforma. "
            else:
                conclusion += f"El período con mayor actividad es {max_label} con {max_value:,} voluntariados ({max_value/total*100:.1f}% del total). "
                conclusion += "Se observa una distribución temporal de las oportunidades creadas. "
            
            conclusion += "\n\nRECOMENDACIONES:\n"
            conclusion += "- Mantener un equilibrio entre voluntariados activos y cerrados.\n"
            conclusion += "- Promover la creación de nuevas oportunidades en períodos de menor actividad.\n"
            conclusion += "- Analizar las causas de cierre para mejorar la gestión."
    
    else:
        conclusion += "ANÁLISIS DE ESTADÍSTICAS\n\n"
        conclusion += f"El análisis muestra un total de {total:,} registros. "
        conclusion += "Los datos presentados reflejan la actividad general de la plataforma. "
        conclusion += "\n\nSe recomienda revisar periódicamente estas estadísticas para identificar tendencias y oportunidades de mejora."
    
    return conclusion
  

if __name__ == "__main__":
    app.run(debug=True, host='0.0.0.0', port=5000)
